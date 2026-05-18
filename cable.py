from collections import OrderedDict
from datetime import datetime
from io import BytesIO
import os
import re
import shutil
import uuid

from flask import request, redirect, url_for, render_template_string, send_file, session
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_

from app import db
from config import Config


ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
RECYCLE_SUBDIR = "recycle"
CABLE_SPECS = ["高频", "低频", "其它"]
CABLE_STATUSES = ["在库", "借出", "计量", "其它"]
CABLE_ADD_TYPES = ["电缆"]
SCAN_PARAM_NAMES = ["keyword", "code", "barcode", "text", "result", "data", "content", "q"]
EMPTY_CABLE_NO_PREFIX = "__EMPTY_CABLE_NO__"


class Cable(db.Model):
    __tablename__ = "cable"

    id = db.Column(db.Integer, primary_key=True)
    cable_no = db.Column(db.String(128), unique=True, nullable=False, index=True)
    name = db.Column(db.String(255), nullable=True)
    spec = db.Column(db.String(32), nullable=True)
    owner = db.Column(db.String(128), nullable=True, index=True)
    location = db.Column(db.String(255), nullable=True, index=True)
    status = db.Column(db.String(32), nullable=True, index=True)
    remark = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class CableImage(db.Model):
    __tablename__ = "cable_image"

    id = db.Column(db.Integer, primary_key=True)
    cable_id = db.Column(db.Integer, db.ForeignKey("cable.id"), nullable=False, index=True)
    image_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class CableShelf(db.Model):
    __tablename__ = "cable_shelf"

    id = db.Column(db.Integer, primary_key=True)
    shelf_name = db.Column(db.String(255), unique=True, nullable=False, index=True)
    remark = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class CableShelfImage(db.Model):
    __tablename__ = "cable_shelf_image"

    id = db.Column(db.Integer, primary_key=True)
    shelf_id = db.Column(db.Integer, db.ForeignKey("cable_shelf.id"), nullable=False, index=True)
    image_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_empty_to_none(value):
    value = normalize_text(value)
    return value if value else None


def normalize_cable_no(value):
    return normalize_text(value).upper()


def is_valid_cable_no(value):
    return bool(re.fullmatch(r"DMDL\d{6}", normalize_cable_no(value)))


def normalize_cable_status(value):
    value = normalize_text(value)
    if value == "维修":
        return "计量"
    return value


def normalize_location(value):
    if value is None:
        return ""

    if isinstance(value, (list, tuple)):
        if not value:
            return ""
        return normalize_location(value[0])

    value = normalize_text(value)
    if not value:
        return ""

    tuple_match = re.fullmatch(r"\(\s*['\"](.+?)['\"]\s*,\s*\)", value)
    if tuple_match:
        value = normalize_text(tuple_match.group(1))

    return value


def is_virtual_empty_cable_no(value):
    value = normalize_text(value)
    return bool(value) and value.startswith(EMPTY_CABLE_NO_PREFIX)


def get_display_cable_no(value):
    value = normalize_text(value)
    return "" if is_virtual_empty_cable_no(value) else value


def make_stored_cable_no(value="", existing_value=""):
    value = normalize_text(value)
    existing_value = normalize_text(existing_value)
    if value:
        return value
    if is_virtual_empty_cable_no(existing_value):
        return existing_value
    return f"{EMPTY_CABLE_NO_PREFIX}{uuid.uuid4().hex}"


def build_cable_filename_prefix(cable_no="", name="", location=""):
    return (
        get_display_cable_no(cable_no)
        or normalize_text(name)
        or normalize_location(location)
        or "cable"
    )


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def sanitize_image_prefix(value):
    value = normalize_text(value)
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("._-")
    return value or "cable"


def get_visitor_role():
    return normalize_text(session.get("visitor_role"))


def has_manage_access():
    return bool(getattr(current_user, "is_authenticated", False)) or get_visitor_role() == "editor"


def get_user_display_name():
    if getattr(current_user, "is_authenticated", False):
        return current_user.username
    role = get_visitor_role()
    if role == "editor":
        return "高级访客"
    if role == "viewer":
        return "普通访客"
    return "访客模式"


def ensure_manage_access():
    if has_manage_access():
        return None
    return ("当前访客链接仅支持检索", 403)


def save_uploaded_image(file_storage, subdir, filename_prefix="cable"):
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        return None

    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    time_part = datetime.now().strftime("%Y%m%d%H%M%S")
    random_part = uuid.uuid4().hex[:6]
    filename = f"{time_part}_{random_part}.{ext}"

    folder = os.path.join(Config.UPLOAD_FOLDER, subdir)
    os.makedirs(folder, exist_ok=True)

    abs_path = os.path.join(folder, filename)
    file_storage.save(abs_path)
    return f"{subdir}/{filename}"


def move_image_file_to_recycle(relative_path):
    if not relative_path:
        return

    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return
    if safe_relative_path == RECYCLE_SUBDIR or safe_relative_path.startswith(f"{RECYCLE_SUBDIR}/"):
        return

    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if not os.path.exists(abs_path) or not os.path.isfile(abs_path):
        return

    recycle_dir = os.path.join(Config.UPLOAD_FOLDER, RECYCLE_SUBDIR, os.path.dirname(safe_relative_path))
    os.makedirs(recycle_dir, exist_ok=True)

    recycle_filename = os.path.basename(safe_relative_path)
    target_path = os.path.join(recycle_dir, recycle_filename)
    if os.path.exists(target_path):
        name, ext = os.path.splitext(recycle_filename)
        target_path = os.path.join(recycle_dir, f"{name}.{uuid.uuid4().hex[:8]}{ext}")

    shutil.move(abs_path, target_path)


def delete_image_file(relative_path):
    move_image_file_to_recycle(relative_path)


def permanent_delete_image_file(relative_path):
    if not relative_path:
        return

    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return

    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if os.path.exists(abs_path) and os.path.isfile(abs_path):
        os.remove(abs_path)


def trim_cable_images(cable):
    images = CableImage.query.filter_by(cable_id=cable.id).order_by(CableImage.created_at.asc(), CableImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def trim_shelf_images(shelf):
    images = CableShelfImage.query.filter_by(shelf_id=shelf.id).order_by(CableShelfImage.created_at.asc(), CableShelfImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def delete_cable_with_files(cable):
    if not cable:
        return

    image_rows = CableImage.query.filter_by(cable_id=cable.id).all()
    for img in image_rows:
        delete_image_file(img.image_path)
        db.session.delete(img)

    if image_rows:
        db.session.flush()

    db.session.delete(cable)


def delete_shelf_with_files(shelf):
    if not shelf:
        return

    image_rows = CableShelfImage.query.filter_by(shelf_id=shelf.id).all()
    for img in image_rows:
        delete_image_file(img.image_path)
        db.session.delete(img)

    if image_rows:
        db.session.flush()

    db.session.delete(shelf)


def get_cable_images(cable_id):
    return CableImage.query.filter_by(cable_id=cable_id).order_by(CableImage.created_at.asc(), CableImage.id.asc()).all()


def get_shelf_images(shelf_id):
    return CableShelfImage.query.filter_by(shelf_id=shelf_id).order_by(CableShelfImage.created_at.asc(), CableShelfImage.id.asc()).all()


def get_shelf_by_location(location):
    location = normalize_location(location)
    if not location:
        return None
    return CableShelf.query.filter_by(shelf_name=location).first()


def ensure_shelf(location, auto_create=False):
    location = normalize_location(location)
    if not location:
        return None
    shelf = CableShelf.query.filter_by(shelf_name=location).first()
    if shelf or not auto_create:
        return shelf
    shelf = CableShelf(shelf_name=location)
    db.session.add(shelf)
    db.session.flush()
    return shelf


def get_all_shelf_locations():
    rows = CableShelf.query.order_by(CableShelf.shelf_name.asc(), CableShelf.id.asc()).all()
    values = []
    seen = set()
    for row in rows:
        location = normalize_location(row.shelf_name)
        if not location or location in seen:
            continue
        values.append(location)
        seen.add(location)
    return values


def build_location_shelf_lookup(locations):
    normalized_locations = [normalize_location(item) for item in locations if normalize_location(item)]
    if not normalized_locations:
        return {}
    shelves = CableShelf.query.filter(CableShelf.shelf_name.in_(normalized_locations)).all()
    return {normalize_location(item.shelf_name): item for item in shelves}


def backfill_cable_shelves():
    changed = False

    for cable in Cable.query.filter(Cable.location.isnot(None)).all():
        normalized = normalize_location(cable.location)
        if cable.location != normalized:
            cable.location = normalized or None
            changed = True

    shelf_rows = CableShelf.query.order_by(CableShelf.id.asc()).all()
    existing = {}
    for shelf in shelf_rows:
        normalized = normalize_location(shelf.shelf_name)
        if not normalized:
            continue

        keeper = existing.get(normalized)
        if keeper and keeper.id != shelf.id:
            for img in CableShelfImage.query.filter_by(shelf_id=shelf.id).all():
                img.shelf_id = keeper.id
                changed = True
            if not normalize_text(keeper.remark) and normalize_text(shelf.remark):
                keeper.remark = shelf.remark
                changed = True
            db.session.delete(shelf)
            changed = True
            continue

        if shelf.shelf_name != normalized:
            shelf.shelf_name = normalized
            changed = True
        existing[normalized] = shelf

    locations = (
        db.session.query(Cable.location)
        .filter(Cable.location.isnot(None))
        .distinct()
        .all()
    )
    for row in locations:
        location = normalize_location(row[0] if isinstance(row, (list, tuple)) else row)
        if location and location not in existing:
            shelf = CableShelf(shelf_name=location)
            db.session.add(shelf)
            existing[location] = shelf
            changed = True

    if changed:
        db.session.commit()


def extract_scan_prefill_code(scan_code=""):
    scan_code = normalize_text(scan_code)
    if scan_code:
        return scan_code
    for name in SCAN_PARAM_NAMES:
        value = normalize_text(request.args.get(name))
        if value:
            return value
    return ""


def extract_scan_keyword(scan_code=""):
    scan_code = normalize_text(scan_code)
    if scan_code:
        return scan_code
    keyword = normalize_text(request.args.get("keyword"))
    if keyword:
        return keyword
    return extract_scan_prefill_code("")


def build_cable_rows(keyword="", searched=False):
    if not searched:
        return []

    keyword = normalize_text(keyword)
    query = Cable.query
    if keyword:
        query = query.filter(
            or_(
                Cable.cable_no.like(f"%{keyword}%"),
                Cable.owner.like(f"%{keyword}%"),
                Cable.location.like(f"%{keyword}%"),
                Cable.remark.like(f"%{keyword}%"),
            )
        )

    items = query.order_by(Cable.location.asc(), Cable.cable_no.asc(), Cable.id.asc()).all()
    shelf_lookup = build_location_shelf_lookup([item.location for item in items])

    rows = []
    for item in items:
        location = item.location or ""
        shelf = shelf_lookup.get(normalize_location(location)) if location else None
        rows.append({
            "row_type": "cable",
            "id": item.id,
            "selected_value": f"cable:{item.id}",
            "cable_no": get_display_cable_no(item.cable_no),
            "name": item.name or "",
            "spec": item.spec or "",
            "owner": item.owner or "",
            "location": location,
            "status": normalize_cable_status(item.status),
            "remark": item.remark or "",
            "detail_url": url_for("cable_detail", cable_id=item.id),
            "location_detail_url": url_for("cable_location_detail", shelf_id=shelf.id) if shelf else "",
        })

    occupied_locations = {
        normalize_location(row[0])
        for row in db.session.query(Cable.location).filter(Cable.location.isnot(None)).distinct().all()
        if normalize_location(row[0])
    }

    shelf_query = CableShelf.query
    if keyword:
        shelf_query = shelf_query.filter(
            or_(
                CableShelf.shelf_name.like(f"%{keyword}%"),
                CableShelf.remark.like(f"%{keyword}%")
            )
        )

    for shelf in shelf_query.order_by(CableShelf.shelf_name.asc(), CableShelf.id.asc()).all():
        shelf_name = normalize_location(shelf.shelf_name)
        if not shelf_name or shelf_name in occupied_locations:
            continue
        rows.append({
            "row_type": "shelf",
            "id": shelf.id,
            "selected_value": f"shelf:{shelf.id}",
            "cable_no": "",
            "name": "空货架",
            "spec": "",
            "owner": "",
            "location": normalize_location(shelf.shelf_name),
            "status": "",
            "remark": shelf.remark or "",
            "detail_url": url_for("cable_location_detail", shelf_id=shelf.id),
            "location_detail_url": url_for("cable_location_detail", shelf_id=shelf.id),
        })

    rows.sort(key=lambda x: (normalize_location(x.get("location")), 0 if x.get("row_type") == "shelf" else 1, normalize_text(x.get("cable_no")), x.get("id") or 0))
    return rows


def import_cables_from_excel(file_storage):
    wb = load_workbook(file_storage)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cable_no = normalize_cable_no(row[0] if len(row) > 0 else "")
        name = normalize_text(row[1] if len(row) > 1 else "")
        spec = normalize_text(row[2] if len(row) > 2 else "")
        owner = normalize_text(row[3] if len(row) > 3 else "")
        location = normalize_text(row[4] if len(row) > 4 else "")
        status = normalize_cable_status(row[5] if len(row) > 5 else "")
        remark = normalize_text(row[6] if len(row) > 6 else "")

        if not cable_no and not name and not spec and not owner and not location and not status and not remark:
            continue

        if cable_no and not is_valid_cable_no(cable_no):
            raise ValueError(f"第{row_idx}行电缆编号格式不正确，应为 DMDL+6位数字，例如 DMDL123456")
        if spec and spec not in CABLE_SPECS:
            raise ValueError(f"第{row_idx}行规格不合法：{spec}")
        if status and status not in CABLE_STATUSES:
            raise ValueError(f"第{row_idx}行状态不合法：{status}")

        if location:
            ensure_shelf(location, auto_create=True)

        obj = Cable.query.filter_by(cable_no=cable_no).first() if cable_no else None
        if obj:
            obj.name = name or ""
            obj.spec = normalize_empty_to_none(spec)
            obj.owner = normalize_empty_to_none(owner)
            obj.location = normalize_empty_to_none(location)
            obj.status = normalize_empty_to_none(status)
            obj.remark = normalize_empty_to_none(remark)
        else:
            obj = Cable(
                cable_no=make_stored_cable_no(cable_no),
                name=name or "",
                spec=normalize_empty_to_none(spec),
                owner=normalize_empty_to_none(owner),
                location=normalize_empty_to_none(location),
                status=normalize_empty_to_none(status),
                remark=normalize_empty_to_none(remark),
            )
            db.session.add(obj)


CABLE_SEARCH_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>电缆查询</title>
<style>
:root{--bg:#edf1f5;--card:#ffffff;--line:#d9e1ea;--text:#233243;--muted:#677383;--primary:#5f6f82;--primary-soft:#eef2f6;--cyan:#5e8e9f;--cyan-soft:#eef6f8;--orange:#b88347;--orange-soft:#faf2e8;}
*{box-sizing:border-box;}
body{font-family:Arial,sans-serif;margin:0;color:var(--text);background:linear-gradient(180deg,#edf1f5 0%,#f7f8fa 34%,#f1f4f8 100%);}.wrap{max-width:1240px;margin:auto;padding:18px 14px 26px;}.card{background:rgba(255,255,255,.94);border:1px solid rgba(210,218,227,.98);border-radius:22px;padding:18px 18px 16px;margin-bottom:16px;box-shadow:0 16px 40px rgba(27,39,53,.07);backdrop-filter:blur(10px);}.topbar{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;}.switch-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;}.switch-row a{flex:1 1 160px;}.switch-row a button{width:100%;}.switch-row .tab-btn{background:linear-gradient(180deg,#f9fbfd 0%,#eef3f7 100%);color:#677383;border:1px solid #d8e0e8;box-shadow:inset 0 1px 0 rgba(255,255,255,.86),0 5px 12px rgba(63,78,96,.06);font-weight:800;letter-spacing:.2px;}.switch-row .tab-btn:hover{background:linear-gradient(180deg,#ffffff 0%,#edf3f8 100%);color:#4d5b6b;border-color:#c8d4df;box-shadow:inset 0 1px 0 rgba(255,255,255,.92),0 10px 18px rgba(63,78,96,.1);}.switch-row .tab-btn.tab-active-cable{background:#ead8aa;color:#6b5630;border-color:#d8c18a;box-shadow:inset 0 1px 0 rgba(255,255,255,.62),0 10px 18px rgba(161,138,84,.16);text-shadow:none;}.switch-row .tab-btn.tab-active-cable:hover{background:#e4cf9a;color:#5f4c2b;border-color:#ccb277;box-shadow:inset 0 1px 0 rgba(255,255,255,.62),0 12px 20px rgba(161,138,84,.2);transform:translateY(-1px);}h2{margin:0 0 14px;font-size:26px;color:#263646;letter-spacing:.4px;}a{color:#5a6675;}input,select,button{padding:11px 12px;font-size:16px;border-radius:12px;border:1px solid #cfd6df;box-sizing:border-box;transition:all .2s ease;}input[type=text]{min-width:260px;background:#fbfdff;}input[type=text]:focus,select:focus{outline:none;border-color:#9ba8b8;box-shadow:0 0 0 4px rgba(95,111,130,.12);background:#fff;}input[type=file]{background:#fff;}input[type=checkbox]{accent-color:#5f6f82;width:16px;height:16px;}button{background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;border:none;cursor:pointer;font-weight:700;box-shadow:0 8px 18px rgba(46,58,74,.14);}button:hover{transform:translateY(-1px);box-shadow:0 12px 22px rgba(46,58,74,.16);}button:disabled{cursor:not-allowed;transform:none;box-shadow:none;}.btn-green{background:linear-gradient(135deg,#4f8d78 0%,#2f6b58 100%);}.btn-gray{background:linear-gradient(135deg,#8b95a1 0%,#697380 100%);}.btn-orange{background:linear-gradient(135deg,#c39a68 0%,#a9753a 100%);}.btn-red{background:linear-gradient(135deg,#d06f72 0%,#b34f55 100%);}
.btn-search-home{background:linear-gradient(135deg,#67c9ff 0%,#3388ff 100%);box-shadow:0 10px 18px rgba(54,132,255,.2);}
.btn-search-home:hover{background:linear-gradient(135deg,#7ad1ff 0%,#247dff 100%);}
.btn-back{background:linear-gradient(135deg,#8b2f45 0%,#6b1f33 100%);box-shadow:0 10px 18px rgba(139,47,69,.22);}
.btn-back:hover{background:linear-gradient(135deg,#9c3850 0%,#77243a 100%);}
.power-link{display:inline-flex;text-decoration:none;}
.btn-power{width:28px;min-width:28px;height:28px;padding:0;border:none;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:radial-gradient(circle at 32% 28%,#f47d82 0%,#e45f65 34%,#cb4049 70%,#aa2733 100%);box-shadow:inset 0 2px 3px rgba(255,255,255,.38),inset 0 -6px 10px rgba(120,10,20,.18),0 2px 0 #8f1f2b,0 8px 16px rgba(146,24,36,.28);color:#fff;position:relative;}
.btn-power::before{content:'';position:absolute;inset:3px;border-radius:999px;border:1.2px solid rgba(120,18,28,.28);box-shadow:inset 0 1px 0 rgba(255,255,255,.18);}
.btn-power svg{width:14px;height:14px;stroke:currentColor;fill:none;stroke-width:2.5;stroke-linecap:round;stroke-linejoin:round;filter:drop-shadow(0 1px 0 rgba(120,10,20,.18));position:relative;z-index:1;}
.btn-power:hover{background:radial-gradient(circle at 32% 28%,#fb8d92 0%,#ea676d 34%,#d24750 70%,#b12c37 100%);box-shadow:inset 0 2px 3px rgba(255,255,255,.42),inset 0 -6px 10px rgba(120,10,20,.22),0 2px 0 #8f1f2b,0 12px 22px rgba(146,24,36,.36);}
.power-link{display:inline-flex;text-decoration:none;}
.btn-reset-home{background:linear-gradient(135deg,#c5ccd4 0%,#97a2af 100%);box-shadow:0 8px 16px rgba(99,116,135,.16);}
.btn-reset-home:hover{background:linear-gradient(135deg,#d0d7df 0%,#8d98a6 100%);}
.btn-add-home{background:linear-gradient(135deg,#54ddb1 0%,#159f76 100%);box-shadow:0 10px 18px rgba(21,159,118,.2);}
.btn-add-home:hover{background:linear-gradient(135deg,#64e4ba 0%,#0f956d 100%);}.btn-disabled{background:#adb5bd !important;color:#fff !important;cursor:not-allowed;pointer-events:none;box-shadow:none !important;}.icon-only-btn{width:44px;min-width:44px;height:44px;padding:0;display:inline-flex;align-items:center;justify-content:center;}.icon-only-btn svg{width:22px;height:22px;stroke:currentColor;fill:none;stroke-width:2.2;stroke-linecap:round;stroke-linejoin:round;}.table-wrap{overflow-x:auto;border:1px solid #d7dee7;border-radius:16px;background:#fff;box-shadow:inset 0 1px 0 rgba(255,255,255,.8),0 8px 22px rgba(27,39,53,.04);}table{width:100%;border-collapse:separate;border-spacing:0;min-width:920px;}.cable-search-table{table-layout:fixed;}th,td{padding:11px 10px;text-align:left;border:none;border-right:1px solid #e7eef7;border-bottom:1px solid #e7eef7;vertical-align:middle;font-size:15px;font-weight:400;color:#6f7b88;}th:last-child,td:last-child{border-right:none;}thead th{position:sticky;top:0;z-index:1;background:linear-gradient(180deg,#fafbfd 0%,#eff2f6 100%);color:#5e6977;font-weight:600;white-space:nowrap;}tbody tr:last-child td{border-bottom:none;}.cable-search-table th:nth-child(1),.cable-search-table td:nth-child(1){width:48px;}.cable-search-table th:nth-child(2),.cable-search-table td:nth-child(2){width:18%;}.cable-search-table th:nth-child(3),.cable-search-table td:nth-child(3){width:16%;}.cable-search-table th:nth-child(4),.cable-search-table td:nth-child(4){width:10%;}.cable-search-table th:nth-child(5),.cable-search-table td:nth-child(5){width:8%;}.cable-search-table th:nth-child(6),.cable-search-table td:nth-child(6){width:9%;}.cable-search-table th:nth-child(7),.cable-search-table td:nth-child(7){width:8%;}.cable-search-table th:nth-child(2),.cable-search-table td:nth-child(2),.cable-search-table th:nth-child(3),.cable-search-table td:nth-child(3),.cable-search-table th:nth-child(4),.cable-search-table td:nth-child(4){white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}.cable-search-table .num-link{font-size:16px;}tbody tr.result-row{transition:transform .16s ease,box-shadow .16s ease;}tbody tr.result-row td:first-child{border-left:4px solid transparent;}tbody tr.result-row:nth-child(4n+1) td{background:#fbfcfd;}tbody tr.result-row:nth-child(4n+2) td{background:#f8fafb;}tbody tr.result-row:nth-child(4n+3) td{background:#fcfbf8;}tbody tr.result-row:nth-child(4n+4) td{background:#f9f8fb;}tbody tr.result-row:hover td{background:#eff3f7;}tbody tr.result-row:hover{transform:translateY(-1px);box-shadow:0 10px 22px rgba(42,56,74,.08);}tbody tr.result-row-cable td:first-child{border-left-color:#8ca9b3;}tbody tr.result-row-shelf td:first-child{border-left-color:#c3a06d;}tbody tr.empty-row td{background:#fbfdff;color:#6f7b88;text-align:center;padding:20px 12px;}.err{color:#dc2626;margin-top:10px;font-weight:700;}.muted{color:var(--muted);font-size:14px;line-height:1.6;}.num-link{color:#4e5865;text-decoration:underline;text-decoration-thickness:1px;text-underline-offset:2px;font-weight:400;font-size:18px;}.num-link:hover{color:#3f4853;text-decoration:underline;}.action-bar{display:flex;gap:10px;flex-wrap:wrap;align-items:center;}.pagination{margin-top:14px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;}.pagination a,.pagination span{display:inline-flex;align-items:center;justify-content:center;min-height:38px;padding:0 14px;border-radius:999px;border:1px solid #d7e3f1;background:#fff;color:#334155;text-decoration:none;}.pagination a:hover{background:#f1f4f7;border-color:#c7d0db;}.summary-badge{display:inline-flex;align-items:center;padding:8px 14px;border-radius:999px;background:var(--primary-soft);color:#4a5767;font-weight:700;border:1px solid #d4dbe4;}.simple-modal{position:fixed;inset:0;background:rgba(15,23,42,.52);display:none;align-items:center;justify-content:center;z-index:10000;padding:18px;}.simple-modal.show{display:flex;}.simple-modal-card{width:100%;max-width:380px;background:#fff;border-radius:18px;padding:22px 20px;box-shadow:0 18px 50px rgba(15,23,42,.28);}.simple-modal-title{font-size:22px;font-weight:800;color:#111827;text-align:center;margin-bottom:10px;}.simple-modal-text{font-size:18px;font-weight:600;line-height:1.7;color:#1f2937;text-align:center;margin-bottom:14px;}.simple-modal-input-wrap{display:none;margin-bottom:12px;}.simple-modal-input-wrap.show{display:block;}.simple-modal-input{width:100%;box-sizing:border-box;padding:12px 14px;font-size:18px;font-weight:700;border-radius:12px;border:2px solid #9ca3af;color:#111827;}.simple-modal-input:focus{outline:none;border-color:#374151;box-shadow:0 0 0 3px rgba(55,65,81,.12);}.simple-modal-error{min-height:24px;font-size:16px;font-weight:700;color:#b91c1c;text-align:center;margin-bottom:6px;}.simple-modal-actions{display:flex;gap:12px;justify-content:center;margin-top:6px;}.simple-modal-actions button{width:auto;min-width:118px;padding:11px 18px;font-size:17px;font-weight:700;border-radius:12px;border:none;}.simple-modal-cancel{background:#6b7280;color:#fff;}.simple-modal-ok{background:#b91c1c;color:#fff;}@media (max-width:768px){.wrap{padding:12px;}.card{padding:14px;margin-bottom:14px;border-radius:14px;}.topbar{gap:8px;}.switch-row{gap:8px;margin-bottom:12px;}.switch-row a{flex:1 1 160px;}h2{font-size:22px;}.action-bar{gap:8px;}.action-bar button{width:auto;}.pagination{justify-content:flex-start;gap:10px;}.pagination a,.pagination span{font-size:14px;}.simple-modal{padding:16px;}.simple-modal-card{border-radius:16px;padding:18px 16px;}.simple-modal-title{font-size:20px;}.simple-modal-text{font-size:16px;}.simple-modal-input{padding:10px 12px;font-size:16px;border-radius:8px;}.simple-modal-actions button{min-width:108px;padding:10px 16px;font-size:16px;border-radius:10px;}input,select,button{padding:10px;font-size:16px;border-radius:8px;}input[type=text]{width:100%;min-width:0;}.icon-only-btn{width:42px;min-width:42px;height:42px;padding:0;}}
</style>
<script>
const ALL_FILTERED_ITEMS = {{ all_selection_meta|tojson|safe }};
function getVisibleSelectedCheckboxes(){ return Array.from(document.querySelectorAll('#cable-batch-form input[name="selected_items"][type="checkbox"]')); }
function rebuildAllFilteredHiddenSelections(enabled){
    const container = document.getElementById('all-filtered-selected-items');
    if(!container){ return; }
    container.innerHTML = '';
    if(!enabled){ return; }
    const visibleValues = new Set(getVisibleSelectedCheckboxes().map(item => item.value));
    ALL_FILTERED_ITEMS.forEach(item => {
        if(!item || !item.selected_value || visibleValues.has(item.selected_value)){ return; }
        const input = document.createElement('input');
        input.type = 'hidden';
        input.name = 'selected_items';
        input.value = item.selected_value;
        input.className = 'selected-hidden-item';
        input.dataset.rowType = item.row_type || '';
        container.appendChild(input);
    });
}
function getAllSelectedItemElements(){
    const visibleChecked = Array.from(document.querySelectorAll('#cable-batch-form input[name="selected_items"][type="checkbox"]:checked'));
    const hiddenSelected = Array.from(document.querySelectorAll('#cable-batch-form .selected-hidden-item'));
    return visibleChecked.concat(hiddenSelected);
}
function toggleSelectFilteredResults(source){
    const items = getVisibleSelectedCheckboxes();
    items.forEach(item => { item.checked = source.checked; });
    rebuildAllFilteredHiddenSelections(!!source.checked);
}

const deleteModalState = { onOk: null, onCancel: null };
function closeDeleteModal(){
    const modal = document.getElementById('delete-modal');
    if(modal){ modal.classList.remove('show'); }
    const inputWrap = document.getElementById('delete-modal-input-wrap');
    const input = document.getElementById('delete-modal-input');
    const error = document.getElementById('delete-modal-error');
    const cancelBtn = document.getElementById('delete-modal-cancel');
    const okBtn = document.getElementById('delete-modal-ok');
    if(inputWrap){ inputWrap.classList.remove('show'); }
    if(input){ input.value = ''; }
    if(error){ error.textContent = ''; }
    if(cancelBtn){ cancelBtn.style.display = 'inline-block'; }
    if(okBtn){ okBtn.textContent = '确定'; }
    deleteModalState.onOk = null;
    deleteModalState.onCancel = null;
}
function openDeleteModal(config){
    const modal = document.getElementById('delete-modal');
    const title = document.getElementById('delete-modal-title');
    const text = document.getElementById('delete-modal-text');
    const inputWrap = document.getElementById('delete-modal-input-wrap');
    const input = document.getElementById('delete-modal-input');
    const error = document.getElementById('delete-modal-error');
    const cancelBtn = document.getElementById('delete-modal-cancel');
    const okBtn = document.getElementById('delete-modal-ok');
    if(!modal || !title || !text || !inputWrap || !input || !error || !cancelBtn || !okBtn){ return; }
    title.textContent = config.title || '提示';
    text.textContent = config.text || '';
    error.textContent = '';
    input.value = '';
    if(config.showPin){
        inputWrap.classList.add('show');
        window.setTimeout(() => { try { input.focus(); } catch(e){} }, 30);
    }else{
        inputWrap.classList.remove('show');
    }
    cancelBtn.style.display = config.hideCancel ? 'none' : 'inline-block';
    okBtn.textContent = config.okText || '确定';
    deleteModalState.onOk = typeof config.onOk === 'function' ? config.onOk : null;
    deleteModalState.onCancel = typeof config.onCancel === 'function' ? config.onCancel : null;
    modal.classList.add('show');
}
function handleDeleteModalCancel(){
    const fn = deleteModalState.onCancel;
    closeDeleteModal();
    if(fn){ fn(); }
}
function handleDeleteModalOk(){
    if(!deleteModalState.onOk){ closeDeleteModal(); return; }
    const result = deleteModalState.onOk();
    if(result !== false){ closeDeleteModal(); }
}
function showDeleteNotice(message){
    openDeleteModal({ title:'提示', text: message || '请确认操作', hideCancel: true, onOk: function(){ return true; } });
}
function showDeletePinDialog(onSubmit){
    openDeleteModal({
        title:'Pin码确认',
        text:'请输入4位Pin码',
        showPin:true,
        onOk:function(){
            const input = document.getElementById('delete-modal-input');
            const error = document.getElementById('delete-modal-error');
            const pin = (input && input.value ? input.value : '').trim();
            if(!pin){ if(error){ error.textContent = '请输入4位Pin码'; } return false; }
            if(pin !== '0819'){ if(error){ error.textContent = 'Pin码错误'; } return false; }
            if(typeof onSubmit === 'function'){ onSubmit(pin); }
            return true;
        }
    });
}
function openDeleteFlow(message, onPinConfirmed){
    openDeleteModal({
        title:'删除确认',
        text: message || '确认删除吗？',
        onOk:function(){
            showDeletePinDialog(onPinConfirmed);
            return false;
        }
    });
    return false;
}

function confirmDeleteSelected(){
    const checkedItems = getAllSelectedItemElements();
    if(checkedItems.length === 0){ showDeleteNotice('请先勾选要删除的电缆或空货架'); return false; }
    const cableCount = checkedItems.filter(item => item.dataset.rowType === 'cable').length;
    const shelfCount = checkedItems.filter(item => item.dataset.rowType === 'shelf').length;
    return openDeleteFlow(`将删除 ${cableCount} 条电缆、${shelfCount} 个空货架`, function(pin){
        document.getElementById('delete_pin').value = pin;
        const form = document.getElementById('cable-batch-form');
        if(form){
            form.action = '/cable/delete_selected';
            form.method = 'post';
            form.submit();
        }
    });
}

</script>
</head>
<body>
<div class="wrap">
    <div class="card"><div class="topbar"><div><strong>当前用户：</strong>{{ user_display }}</div><div>{% if current_user.is_authenticated %}<a class="power-link" href="/logout" title="退出登录"><button type="button" class="btn-power" aria-label="退出登录"><svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 3v8"></path><path d="M7.05 5.05a9 9 0 1 0 9.9 0"></path></svg></button></a>{% endif %}</div></div></div>
    <div class="card">
        <div class="switch-row"><a href="/"><button type="button" class="tab-btn">资产查询</button></a><a href="/cable"><button type="button" class="tab-btn tab-active-cable">电缆查询</button></a></div>
        <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-bottom:10px;">
            <h2 style="margin:0;">电缆查询</h2>
            <a href="/scan_label?auto=1" title="条形码扫描"><button type="button" class="btn-gray icon-only-btn" aria-label="条形码扫描">
                <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M8 4H6a2 2 0 0 0-2 2v2"></path><path d="M16 4h2a2 2 0 0 1 2 2v2"></path><path d="M8 20H6a2 2 0 0 1-2-2v-2"></path><path d="M16 20h2a2 2 0 0 0 2-2v-2"></path><path d="M5 12h14"></path></svg>
            </button></a>
        </div>
        {% if scan_code %}<div style="color:#666;margin-bottom:10px;word-break:break-all;">当前来自扫码访问，已自动带入电缆编号：<strong>{{ scan_code }}</strong></div>{% endif %}
        <form method="get" action="/cable">
            <input type="hidden" name="searched" value="1">
            <div class="action-bar">
                <input type="text" name="keyword" value="{{ keyword }}" placeholder="编号、责任人、位置、备注">
                <select name="spec_filter"><option value="">规格</option>{% for s in specs %}<option value="{{ s }}" {% if spec_filter == s %}selected{% endif %}>{{ s }}</option>{% endfor %}</select>
                <select name="status_filter"><option value="">状态</option>{% for s in statuses %}<option value="{{ s }}" {% if status_filter == s %}selected{% endif %}>{{ s }}</option>{% endfor %}</select>
                <button type="submit" class="btn-search-home">搜索</button>
                <a href="/cable"><button type="button" class="btn-reset-home">重置</button></a>
                {% if can_manage %}<a href="/cable/new{% if prefill_cable_no %}?cable_no={{ prefill_cable_no|urlencode }}{% endif %}"><button type="button" class="btn-add-home">+</button></a>{% else %}<button type="button" class="btn-disabled" disabled>+</button>{% endif %}
            </div>
        </form>
        <form method="post" action="/cable/import" enctype="multipart/form-data" style="margin-top:12px;"><div class="action-bar"><input type="file" name="excel_file" accept=".xlsx,.xlsm,.xltx,.xltm" {% if not can_manage %}disabled{% endif %}><button type="submit" class="btn-orange {% if not can_manage %}btn-disabled{% endif %}" {% if not can_manage %}disabled{% endif %}>批量上传电缆</button></div><div class="muted" style="margin-top:8px;">Excel导入格式与导出一致</div></form>
        {% if error %}<div class="err">{{ error }}</div>{% endif %}
    </div>
    {% if searched %}
    <div class="card">
        <form id="cable-batch-form" method="post" action="/cable/export">
            <input type="hidden" name="keyword" value="{{ keyword }}"><input type="hidden" name="spec_filter" value="{{ spec_filter }}"><input type="hidden" name="status_filter" value="{{ status_filter }}"><input type="hidden" name="delete_pin" id="delete_pin" value=""><div id="all-filtered-selected-items"></div>
            <div class="action-bar" style="margin-bottom:10px;"><button type="submit" class="btn-green">导出选中</button><button type="submit" class="btn-red {% if not can_manage %}btn-disabled{% endif %}" formaction="/cable/delete_selected" formmethod="post" onclick="return {% if can_manage %}confirmDeleteSelected(){% else %}false{% endif %};" {% if not can_manage %}disabled{% endif %}>删除选中</button><div class="summary-badge">当前总数：{{ total }}</div></div>
            <div class="table-wrap"><table class="cable-search-table"><thead><tr><th><input type="checkbox" onclick="toggleSelectFilteredResults(this)" title="选取当前筛选结果"></th><th>编号</th><th>名称</th><th>位置</th><th>规格</th><th>责任</th><th>状态</th><th>备注</th></tr></thead><tbody>{% for item in rows %}<tr class="result-row result-row-{{ item.row_type }}"><td><input type="checkbox" name="selected_items" value="{{ item.selected_value }}" data-row-type="{{ item.row_type }}"></td><td>{% if item.row_type == 'shelf' %}<a class="num-link" href="{{ item.detail_url }}">（空）</a>{% else %}<a class="num-link" href="{{ item.detail_url }}">{{ item.cable_no }}</a>{% endif %}</td><td>{{ item.name }}</td><td>{% if item.location_detail_url %}<a class="num-link" href="{{ item.location_detail_url }}">{{ item.location }}</a>{% else %}{{ item.location }}{% endif %}</td><td>{{ item.spec }}</td><td>{{ item.owner }}</td><td>{{ item.status }}</td><td>{{ item.remark }}</td></tr>{% endfor %}{% if not rows %}<tr class="empty-row"><td colspan="8">暂无数据</td></tr>{% endif %}</tbody></table></div>
        </form>
        <div class="pagination">{% if page > 1 %}<a href="/cable?searched=1&keyword={{ keyword }}&spec_filter={{ spec_filter }}&status_filter={{ status_filter }}&page={{ page - 1 }}">上一页</a>{% endif %}<span>第 {{ page }} / {{ total_pages }} 页</span>{% if page < total_pages %}<a href="/cable?searched=1&keyword={{ keyword }}&spec_filter={{ spec_filter }}&status_filter={{ status_filter }}&page={{ page + 1 }}">下一页</a>{% endif %}</div>
    </div>
    {% endif %}

<div id="delete-modal" class="simple-modal" onclick="if(event.target===this){closeDeleteModal();}">
    <div class="simple-modal-card">
        <div id="delete-modal-title" class="simple-modal-title">删除确认</div>
        <div id="delete-modal-text" class="simple-modal-text">确认删除吗？</div>
        <div id="delete-modal-input-wrap" class="simple-modal-input-wrap">
            <input id="delete-modal-input" class="simple-modal-input" type="password" inputmode="numeric" maxlength="4" placeholder="请输入4位Pin码">
        </div>
        <div id="delete-modal-error" class="simple-modal-error"></div>
        <div class="simple-modal-actions">
            <button type="button" id="delete-modal-cancel" class="simple-modal-cancel" onclick="handleDeleteModalCancel();">取消</button>
            <button type="button" id="delete-modal-ok" class="simple-modal-ok" onclick="handleDeleteModalOk();">确定</button>
        </div>
    </div>
</div>

</div>
</body>
</html>
"""


CABLE_NEW_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>新增电缆</title>
<style>
:root{--bg:#edf1f5;--card:#ffffff;--line:#d9e1ea;--text:#233243;--muted:#677383;--primary:#5f6f82;--primary-soft:#eef2f6;--green:#2f7d67;--orange:#b88347;--danger:#dc2626;}
*{box-sizing:border-box;}
body{font-family:Arial,sans-serif;margin:0;color:var(--text);background:linear-gradient(180deg,#edf1f5 0%,#f7f8fa 34%,#f1f4f8 100%);}
.wrap{max-width:960px;margin:auto;padding:16px 14px 26px;}
.card{background:rgba(255,255,255,.94);border:1px solid rgba(210,218,227,.98);border-radius:22px;padding:18px 18px 16px;margin-bottom:16px;box-shadow:0 16px 40px rgba(27,39,53,.07);backdrop-filter:blur(10px);}
h2{margin:0 0 14px;font-size:26px;color:#263646;letter-spacing:.4px;}
a{color:#5a6675;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:600;color:#5e6977;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px 12px;font-size:16px;border-radius:12px;border:1px solid #cfd6df;transition:all .2s ease;}
input[type=text],input[type=date],input[type=password],select,textarea{background:#fbfdff;}
input[type=text]:focus,input[type=date]:focus,input[type=password]:focus,select:focus,textarea:focus{outline:none;border-color:#9ba8b8;box-shadow:0 0 0 4px rgba(95,111,130,.12);background:#fff;}
input[type=file]{background:#fff;}
input[type=checkbox]{accent-color:#5f6f82;width:16px;height:16px;}
textarea{min-height:90px;resize:vertical;}
button{background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;border:none;cursor:pointer;font-weight:700;box-shadow:0 8px 18px rgba(46,58,74,.14);}
button:hover{transform:translateY(-1px);box-shadow:0 12px 22px rgba(46,58,74,.16);}
button:disabled{cursor:not-allowed;transform:none;box-shadow:none;}
.btn-green{background:linear-gradient(135deg,#4f8d78 0%,#2f6b58 100%);}
.btn-gray{background:linear-gradient(135deg,#8b95a1 0%,#697380 100%);}
.btn-orange{background:linear-gradient(135deg,#c39a68 0%,#a9753a 100%);}
.btn-red{background:linear-gradient(135deg,#d06f72 0%,#b34f55 100%);}
.btn-back{background:linear-gradient(135deg,#8b2f45 0%,#6b1f33 100%);box-shadow:0 10px 18px rgba(139,47,69,.22);}
.btn-back:hover{background:linear-gradient(135deg,#9c3850 0%,#77243a 100%);}
.btn-disabled{background:#adb5bd !important;color:#fff !important;cursor:not-allowed;pointer-events:none;box-shadow:none !important;}
.readonly{background:#eef2f6;color:#677383;border-color:#d8e0e8;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.table-wrap{overflow-x:auto;border:1px solid #d7dee7;border-radius:16px;background:#fff;box-shadow:inset 0 1px 0 rgba(255,255,255,.8),0 8px 22px rgba(27,39,53,.04);}
table{width:100%;border-collapse:separate;border-spacing:0;min-width:780px;}
th,td{padding:11px 10px;text-align:left;border:none;border-right:1px solid #e7eef7;border-bottom:1px solid #e7eef7;vertical-align:middle;font-size:15px;font-weight:400;color:#6f7b88;}th:last-child,td:last-child{border-right:none;}
thead th{background:linear-gradient(180deg,#fafbfd 0%,#eff2f6 100%);color:#5e6977;font-weight:600;white-space:nowrap;}
tbody tr:last-child td{border-bottom:none;}
tbody tr:nth-child(odd) td{background:#fbfcfd;}
tbody tr:nth-child(even) td{background:#f8fafb;}
tbody tr:hover td{background:#eff3f7;}
.shelf-cable-table{table-layout:fixed;min-width:920px;}
.shelf-cable-table th:nth-child(1),.shelf-cable-table td:nth-child(1){width:17%;min-width:156px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.shelf-cable-table th:nth-child(2),.shelf-cable-table td:nth-child(2){width:16%;min-width:147px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.shelf-cable-table th:nth-child(3),.shelf-cable-table td:nth-child(3){width:8%;min-width:74px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.shelf-cable-table th:nth-child(4),.shelf-cable-table td:nth-child(4){width:9%;min-width:83px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.shelf-cable-table th:nth-child(5),.shelf-cable-table td:nth-child(5){width:8%;min-width:74px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.msg{color:#15803d;margin-bottom:12px;font-weight:700;background:#eef8f3;border:1px solid #cfe4d9;border-radius:12px;padding:10px 12px;}
.err{color:#dc2626;margin-bottom:12px;font-weight:700;background:#fbf1f2;border:1px solid #e7cfd3;border-radius:12px;padding:10px 12px;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:10px;padding:10px 12px;background:#fafbfd;border:1px solid #dde4ec;border-radius:12px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#66788a;font-size:14px;word-break:break-all;line-height:1.6;}
.selected-file-list{display:flex;flex-direction:column;gap:6px;margin-top:8px;}
.selected-file-item{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 10px;background:#fff;border:1px solid #e0e7ef;border-radius:10px;color:#506172;line-height:1.35;}
.selected-file-name{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.selected-file-remove{width:28px;min-width:28px;height:28px;padding:0;border:none;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:#dc2626;color:#fff;font-size:20px;font-weight:800;line-height:1;box-shadow:0 4px 10px rgba(220,38,38,.18);}
.selected-file-remove:hover{background:#b91c1c;transform:none;}
.upload-dialog{position:fixed;left:0;top:0;width:100vw;height:100dvh;background:transparent;display:none;z-index:9999;overflow:hidden;} .upload-dialog.show{display:block;} .upload-dialog-card{position:fixed;left:50%;top:56%;transform:translate(-50%,-50%);width:min(320px,calc(100vw - 24px));max-width:320px;background:#fff;border-radius:16px;padding:18px;box-shadow:0 20px 40px rgba(15,23,42,.22);z-index:10000;}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;color:#163047;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:12px;border:1px solid #cfd9e6;background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;text-align:center;overflow:hidden;box-shadow:0 8px 18px rgba(46,58,74,.14);}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-choice-file.upload-choice-camera{background:linear-gradient(135deg,#67c9ff 0%,#3388ff 100%);box-shadow:0 10px 18px rgba(54,132,255,.2);}
.upload-choice-file.upload-choice-camera:hover{background:linear-gradient(135deg,#7ad1ff 0%,#247dff 100%);}
.upload-choice-file.upload-choice-local{background:linear-gradient(135deg,#54ddb1 0%,#159f76 100%);box-shadow:0 10px 18px rgba(21,159,118,.2);}
.upload-choice-file.upload-choice-local:hover{background:linear-gradient(135deg,#64e4ba 0%,#0f956d 100%);}
.upload-dialog-actions .btn-cancel{background:linear-gradient(135deg,#7b8794 0%,#5f6b77 100%);}
.upload-dialog-actions .btn-cancel:hover{background:linear-gradient(135deg,#8b97a4 0%,#697582 100%);}
.title-row{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:nowrap;margin-bottom:14px;}
.title-row h2{margin:0;}
.title-row .btn-back{width:auto;min-width:108px;flex:0 0 auto;}
.detail-actions{display:flex;gap:82px;flex-wrap:wrap;align-items:center;margin-top:12px;}
.detail-actions button{width:auto;min-width:120px;}
.link-row{margin-top:-2px;margin-bottom:12px;color:#607284;font-size:14px;}
.muted{color:var(--muted);font-size:14px;line-height:1.6;}.num-link{color:#4e5865;text-decoration:underline;text-decoration-thickness:1px;text-underline-offset:2px;font-weight:400;font-size:18px;}.num-link:hover{color:#3f4853;text-decoration:underline;}

.switch-row{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.simple-modal{position:fixed;inset:0;background:rgba(15,23,42,.52);display:none;align-items:center;justify-content:center;z-index:10000;padding:18px;}
.simple-modal.show{display:flex;}
.simple-modal-card{width:100%;max-width:380px;background:#fff;border-radius:18px;padding:22px 20px;box-shadow:0 18px 50px rgba(15,23,42,.28);}
.simple-modal-title{font-size:22px;font-weight:800;color:#111827;text-align:center;margin-bottom:10px;}
.simple-modal-text{font-size:18px;font-weight:600;line-height:1.7;color:#1f2937;text-align:center;margin-bottom:14px;}
.simple-modal-input-wrap{display:none;margin-bottom:12px;}
.simple-modal-input-wrap.show{display:block;}
.simple-modal-input{width:100%;box-sizing:border-box;padding:12px 14px;font-size:18px;font-weight:700;border-radius:12px;border:2px solid #9ca3af;color:#111827;}
.simple-modal-input:focus{outline:none;border-color:#374151;box-shadow:0 0 0 3px rgba(55,65,81,.12);}
.simple-modal-error{min-height:24px;font-size:16px;font-weight:700;color:#b91c1c;text-align:center;margin-bottom:6px;}
.simple-modal-actions{display:flex;gap:12px;justify-content:center;margin-top:6px;}
.simple-modal-actions button{width:auto;min-width:118px;padding:11px 18px;font-size:17px;font-weight:700;border-radius:12px;border:none;}
.simple-modal-cancel{background:#6b7280;color:#fff;}
.simple-modal-ok{background:#b91c1c;color:#fff;}
@media (max-width:768px){.wrap{padding:12px;}.card{padding:14px;margin-bottom:14px;border-radius:14px;}.grid{grid-template-columns:1fr;}.title-row{margin-bottom:12px;}.title-row .btn-back{min-width:96px;padding:10px 14px;}h2{font-size:22px;}.switch-row{gap:8px;margin-bottom:12px;}.switch-row a{flex:1 1 160px;}input,select,textarea,button{padding:10px;font-size:16px;border-radius:8px;}.upload-choice-file{padding:10px;font-size:16px;border-radius:8px;}.upload-dialog-card{border-radius:14px;padding:16px;width:min(320px,calc(100vw - 20px));top:58%;}.simple-modal{padding:16px;}.simple-modal-card{border-radius:16px;padding:18px 16px;}.simple-modal-title{font-size:20px;}.simple-modal-text{font-size:16px;}.simple-modal-input{padding:10px 12px;font-size:16px;border-radius:8px;}.simple-modal-actions button{min-width:108px;padding:10px 16px;font-size:16px;border-radius:10px;}}
</style>
<script>
function openUploadChooser(dialogId, triggerEl){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.add('show'); } }
function closeUploadChooser(dialogId){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.remove('show'); } }

const uploadFileQueues = {};
function makeShortLocalImageName(file){
    const now = new Date();
    const pad = value => String(value).padStart(2, '0');
    const timePart = `${now.getFullYear()}${pad(now.getMonth() + 1)}${pad(now.getDate())}${pad(now.getHours())}${pad(now.getMinutes())}${pad(now.getSeconds())}`;
    const randomPart = Math.random().toString(36).slice(2, 8).padEnd(6, '0').slice(0, 6);
    const originalName = file && file.name ? file.name : '';
    const extMatch = originalName.match(/\\.([A-Za-z0-9]+)$/);
    const ext = extMatch ? extMatch[1].toLowerCase() : ((file.type || '').split('/')[1] || 'jpg').toLowerCase();
    try{
        return new File([file], `${timePart}_${randomPart}.${ext}`, {type: file.type || 'image/jpeg', lastModified: file.lastModified || Date.now()});
    }catch(e){
        return file;
    }
}
function syncQueuedFilesToInput(textEl, inputIds){
    const dt = new DataTransfer();
    (uploadFileQueues[textEl.id] || []).forEach(file => dt.items.add(file));
    inputIds.forEach((id, index) => {
        const input = document.getElementById(id);
        if(!input){ return; }
        input.files = index === 0 ? dt.files : new DataTransfer().files;
        input.value = '';
    });
}
function renderQueuedFiles(textEl, inputIds){
    const queue = uploadFileQueues[textEl.id] || [];
    if(queue.length === 0){
        textEl.textContent = '未选择图片';
        return;
    }
    textEl.innerHTML = `<div>已选择 ${queue.length} 张（最多5张）：</div><div class="selected-file-list"></div>`;
    const list = textEl.querySelector('.selected-file-list');
    queue.forEach((file, index) => {
        const row = document.createElement('div');
        row.className = 'selected-file-item';
        const name = document.createElement('span');
        name.className = 'selected-file-name';
        name.textContent = file.name || `图片${index + 1}`;
        const remove = document.createElement('button');
        remove.type = 'button';
        remove.className = 'selected-file-remove';
        remove.textContent = '×';
        remove.setAttribute('aria-label', '删除该图片');
        remove.onclick = function(){
            uploadFileQueues[textEl.id].splice(index, 1);
            syncQueuedFilesToInput(textEl, inputIds);
            renderQueuedFiles(textEl, inputIds);
        };
        row.appendChild(name);
        row.appendChild(remove);
        list.appendChild(row);
    });
}
function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    if(!uploadFileQueues[textId]){ uploadFileQueues[textId] = []; }
    const input = document.getElementById(inputId);
    const incoming = input && input.files ? Array.from(input.files) : [];
    incoming.forEach(file => {
        if(uploadFileQueues[textId].length < 5){
            uploadFileQueues[textId].push(makeShortLocalImageName(file));
        }
    });
    if(incoming.length > 0 && uploadFileQueues[textId].length >= 5 && incoming.length > 5){
        alert('本次最多上传5张图片');
    }
    syncQueuedFilesToInput(textEl, inputIds);
    renderQueuedFiles(textEl, inputIds);
    if(dialogId){ closeUploadChooser(dialogId); }
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="title-row"><h2>新增电缆</h2><a href="/cable"><button type="button" class="btn-back">返回</button></a></div>
        <div id="page-note" class="muted" style="margin-bottom:10px;">只保留新增电缆。货架无需单独新增，直接在“位置”中输入即可；相同位置会自动归集到同一个货架。</div>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row"><label>电缆编号</label><input type="text" name="cable_no" value="{{ form_data.cable_no }}"></div>
                <div class="row"><label>电缆名称</label><input type="text" name="name" value="{{ form_data.name }}"></div>

                <div class="row">
                    <label>规格</label>
                    <select name="spec">
                        <option value="">请选择</option>
                        {% for s in specs %}
                        <option value="{{ s }}" {% if form_data.spec == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>

                <div class="row"><label>责任人</label><input type="text" name="owner" value="{{ form_data.owner }}"></div>

                <div class="row">
                    <label>位置</label>
                    <input type="text" name="location" value="{{ form_data.location }}" list="cable-location-options">
                    <datalist id="cable-location-options">
                        {% for item in shelf_locations %}
                        <option value="{{ item }}"></option>
                        {% endfor %}
                    </datalist>
                </div>

                <div class="row">
                    <label>状态</label>
                    <select name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s }}" {% if form_data.status == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>

                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" onclick=\"openUploadChooser('new-upload-choice-dialog', this)\">上传图片</button>
                    </div>
                    <div id="new-image-files-text" class="file-list" data-inputs="new-camera-files,new-file-files">未选择图片</div>
                    <div id="new-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('new-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file upload-choice-camera">拍照
                                    <input type="file" id="new-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('new-camera-files', 'new-image-files-text', 'new-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file upload-choice-local">本地上传
                                    <input type="file" id="new-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('new-file-files', 'new-image-files-text', 'new-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('new-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <div class="row"><label>备注</label><textarea name="remark">{{ form_data.remark }}</textarea></div>
            <button type="submit">保存</button>
        </form>
    </div>
</div>
</body>
</html>
"""


CABLE_FORM_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ page_title }}</title>
<style>
:root{--bg:#edf1f5;--card:#ffffff;--line:#d9e1ea;--text:#233243;--muted:#677383;--primary:#5f6f82;--primary-soft:#eef2f6;--green:#2f7d67;--orange:#b88347;--danger:#dc2626;}
*{box-sizing:border-box;}
body{font-family:Arial,sans-serif;margin:0;color:var(--text);background:linear-gradient(180deg,#edf1f5 0%,#f7f8fa 34%,#f1f4f8 100%);}
.wrap{max-width:950px;margin:auto;padding:16px 14px 26px;}
.card{background:rgba(255,255,255,.94);border:1px solid rgba(210,218,227,.98);border-radius:22px;padding:18px 18px 16px;margin-bottom:16px;box-shadow:0 16px 40px rgba(27,39,53,.07);backdrop-filter:blur(10px);}
h2{margin:0 0 14px;font-size:26px;color:#263646;letter-spacing:.4px;}
a{color:#5a6675;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:600;color:#5e6977;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px 12px;font-size:16px;border-radius:12px;border:1px solid #cfd6df;transition:all .2s ease;}
input[type=text],input[type=date],input[type=password],select,textarea{background:#fbfdff;}
input[type=text]:focus,input[type=date]:focus,input[type=password]:focus,select:focus,textarea:focus{outline:none;border-color:#9ba8b8;box-shadow:0 0 0 4px rgba(95,111,130,.12);background:#fff;}
input[type=file]{background:#fff;}
input[type=checkbox]{accent-color:#5f6f82;width:16px;height:16px;}
textarea{min-height:90px;resize:vertical;}
button{background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;border:none;cursor:pointer;font-weight:700;box-shadow:0 8px 18px rgba(46,58,74,.14);}
button:hover{transform:translateY(-1px);box-shadow:0 12px 22px rgba(46,58,74,.16);}
button:disabled{cursor:not-allowed;transform:none;box-shadow:none;}
.btn-green{background:linear-gradient(135deg,#4f8d78 0%,#2f6b58 100%);}
.btn-gray{background:linear-gradient(135deg,#8b95a1 0%,#697380 100%);}
.btn-orange{background:linear-gradient(135deg,#c39a68 0%,#a9753a 100%);}
.btn-red{background:linear-gradient(135deg,#d06f72 0%,#b34f55 100%);}
.btn-back{background:linear-gradient(135deg,#8b2f45 0%,#6b1f33 100%);box-shadow:0 10px 18px rgba(139,47,69,.22);}
.btn-back:hover{background:linear-gradient(135deg,#9c3850 0%,#77243a 100%);}
.btn-disabled{background:#adb5bd !important;color:#fff !important;cursor:not-allowed;pointer-events:none;box-shadow:none !important;}
.readonly{background:#eef2f6;color:#677383;border-color:#d8e0e8;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.table-wrap{overflow-x:auto;border:1px solid #d7dee7;border-radius:16px;background:#fff;box-shadow:inset 0 1px 0 rgba(255,255,255,.8),0 8px 22px rgba(27,39,53,.04);}
table{width:100%;border-collapse:separate;border-spacing:0;min-width:780px;}
th,td{padding:11px 10px;text-align:left;border:none;border-right:1px solid #e7eef7;border-bottom:1px solid #e7eef7;vertical-align:middle;font-size:15px;font-weight:400;color:#6f7b88;}th:last-child,td:last-child{border-right:none;}
thead th{background:linear-gradient(180deg,#fafbfd 0%,#eff2f6 100%);color:#5e6977;font-weight:600;white-space:nowrap;}
tbody tr:last-child td{border-bottom:none;}
tbody tr:nth-child(odd) td{background:#fbfcfd;}
tbody tr:nth-child(even) td{background:#f8fafb;}
tbody tr:hover td{background:#eff3f7;}
.msg{color:#15803d;margin-bottom:12px;font-weight:700;background:#eef8f3;border:1px solid #cfe4d9;border-radius:12px;padding:10px 12px;}
.err{color:#dc2626;margin-bottom:12px;font-weight:700;background:#fbf1f2;border:1px solid #e7cfd3;border-radius:12px;padding:10px 12px;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:10px;padding:10px 12px;background:#fafbfd;border:1px solid #dde4ec;border-radius:12px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#66788a;font-size:14px;word-break:break-all;line-height:1.6;}
.selected-file-list{display:flex;flex-direction:column;gap:6px;margin-top:8px;}
.selected-file-item{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 10px;background:#fff;border:1px solid #e0e7ef;border-radius:10px;color:#506172;line-height:1.35;}
.selected-file-name{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.selected-file-remove{width:28px;min-width:28px;height:28px;padding:0;border:none;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:#dc2626;color:#fff;font-size:20px;font-weight:800;line-height:1;box-shadow:0 4px 10px rgba(220,38,38,.18);}
.selected-file-remove:hover{background:#b91c1c;transform:none;}
.upload-dialog{position:fixed;left:0;top:0;width:100vw;height:100dvh;background:transparent;display:none;z-index:9999;overflow:hidden;} .upload-dialog.show{display:block;} .upload-dialog-card{position:fixed;left:50%;top:56%;transform:translate(-50%,-50%);width:min(320px,calc(100vw - 24px));max-width:320px;background:#fff;border-radius:16px;padding:18px;box-shadow:0 20px 40px rgba(15,23,42,.22);z-index:10000;}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;color:#163047;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:12px;border:1px solid #cfd9e6;background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;text-align:center;overflow:hidden;box-shadow:0 8px 18px rgba(46,58,74,.14);}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-choice-file.upload-choice-camera{background:linear-gradient(135deg,#67c9ff 0%,#3388ff 100%);box-shadow:0 10px 18px rgba(54,132,255,.2);}
.upload-choice-file.upload-choice-camera:hover{background:linear-gradient(135deg,#7ad1ff 0%,#247dff 100%);}
.upload-choice-file.upload-choice-local{background:linear-gradient(135deg,#54ddb1 0%,#159f76 100%);box-shadow:0 10px 18px rgba(21,159,118,.2);}
.upload-choice-file.upload-choice-local:hover{background:linear-gradient(135deg,#64e4ba 0%,#0f956d 100%);}
.upload-dialog-actions .btn-cancel{background:linear-gradient(135deg,#7b8794 0%,#5f6b77 100%);}
.upload-dialog-actions .btn-cancel:hover{background:linear-gradient(135deg,#8b97a4 0%,#697582 100%);}
.title-row{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:nowrap;margin-bottom:14px;}
.title-row h2{margin:0;}
.title-row .btn-back{width:auto;min-width:108px;flex:0 0 auto;}
.detail-actions{display:flex;gap:82px;flex-wrap:wrap;align-items:center;margin-top:12px;}
.detail-actions button{width:auto;min-width:120px;}
.link-row{margin-top:-2px;margin-bottom:12px;color:#607284;font-size:14px;}
.muted{color:var(--muted);font-size:14px;line-height:1.6;}.num-link{color:#4e5865;text-decoration:underline;text-decoration-thickness:1px;text-underline-offset:2px;font-weight:400;font-size:18px;}.num-link:hover{color:#3f4853;text-decoration:underline;}

.switch-row{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.simple-modal{position:fixed;inset:0;background:rgba(15,23,42,.52);display:none;align-items:center;justify-content:center;z-index:10000;padding:18px;}
.simple-modal.show{display:flex;}
.simple-modal-card{width:100%;max-width:380px;background:#fff;border-radius:18px;padding:22px 20px;box-shadow:0 18px 50px rgba(15,23,42,.28);}
.simple-modal-title{font-size:22px;font-weight:800;color:#111827;text-align:center;margin-bottom:10px;}
.simple-modal-text{font-size:18px;font-weight:600;line-height:1.7;color:#1f2937;text-align:center;margin-bottom:14px;}
.simple-modal-input-wrap{display:none;margin-bottom:12px;}
.simple-modal-input-wrap.show{display:block;}
.simple-modal-input{width:100%;box-sizing:border-box;padding:12px 14px;font-size:18px;font-weight:700;border-radius:12px;border:2px solid #9ca3af;color:#111827;}
.simple-modal-input:focus{outline:none;border-color:#374151;box-shadow:0 0 0 3px rgba(55,65,81,.12);}
.simple-modal-error{min-height:24px;font-size:16px;font-weight:700;color:#b91c1c;text-align:center;margin-bottom:6px;}
.simple-modal-actions{display:flex;gap:12px;justify-content:center;margin-top:6px;}
.simple-modal-actions button{width:auto;min-width:118px;padding:11px 18px;font-size:17px;font-weight:700;border-radius:12px;border:none;}
.simple-modal-cancel{background:#6b7280;color:#fff;}
.simple-modal-ok{background:#b91c1c;color:#fff;}
@media (max-width:768px){.wrap{padding:14px 12px 20px;}.card{padding:16px;margin-bottom:14px;border-radius:14px;}.grid{grid-template-columns:1fr;}.title-row{margin-bottom:12px;}.title-row .btn-back{min-width:96px;padding:10px 14px;}h2{font-size:22px;}.switch-row{gap:8px;}.switch-row a{flex:1 1 160px;}input,select,textarea,button{padding:10px;font-size:16px;border-radius:8px;}.upload-choice-file{padding:10px;font-size:16px;border-radius:8px;}.upload-dialog-card{border-radius:14px;padding:16px;width:min(320px,calc(100vw - 20px));top:58%;}.simple-modal{padding:16px;}.simple-modal-card{border-radius:16px;padding:18px 16px;}.simple-modal-title{font-size:20px;}.simple-modal-text{font-size:16px;}.simple-modal-input{padding:10px 12px;font-size:16px;border-radius:8px;}.simple-modal-actions button{min-width:108px;padding:10px 16px;font-size:16px;border-radius:10px;}}
</style>
<script>
function openUploadChooser(dialogId, triggerEl){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.add('show'); } }
function closeUploadChooser(dialogId){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.remove('show'); } }

const uploadFileQueues = {};
function makeShortLocalImageName(file){
    const now = new Date();
    const pad = value => String(value).padStart(2, '0');
    const timePart = `${now.getFullYear()}${pad(now.getMonth() + 1)}${pad(now.getDate())}${pad(now.getHours())}${pad(now.getMinutes())}${pad(now.getSeconds())}`;
    const randomPart = Math.random().toString(36).slice(2, 8).padEnd(6, '0').slice(0, 6);
    const originalName = file && file.name ? file.name : '';
    const extMatch = originalName.match(/\\.([A-Za-z0-9]+)$/);
    const ext = extMatch ? extMatch[1].toLowerCase() : ((file.type || '').split('/')[1] || 'jpg').toLowerCase();
    try{
        return new File([file], `${timePart}_${randomPart}.${ext}`, {type: file.type || 'image/jpeg', lastModified: file.lastModified || Date.now()});
    }catch(e){
        return file;
    }
}
function syncQueuedFilesToInput(textEl, inputIds){
    const dt = new DataTransfer();
    (uploadFileQueues[textEl.id] || []).forEach(file => dt.items.add(file));
    inputIds.forEach((id, index) => {
        const input = document.getElementById(id);
        if(!input){ return; }
        input.files = index === 0 ? dt.files : new DataTransfer().files;
        input.value = '';
    });
}
function renderQueuedFiles(textEl, inputIds){
    const queue = uploadFileQueues[textEl.id] || [];
    if(queue.length === 0){
        textEl.textContent = '未选择图片';
        return;
    }
    textEl.innerHTML = `<div>已选择 ${queue.length} 张（最多5张）：</div><div class="selected-file-list"></div>`;
    const list = textEl.querySelector('.selected-file-list');
    queue.forEach((file, index) => {
        const row = document.createElement('div');
        row.className = 'selected-file-item';
        const name = document.createElement('span');
        name.className = 'selected-file-name';
        name.textContent = file.name || `图片${index + 1}`;
        const remove = document.createElement('button');
        remove.type = 'button';
        remove.className = 'selected-file-remove';
        remove.textContent = '×';
        remove.setAttribute('aria-label', '删除该图片');
        remove.onclick = function(){
            uploadFileQueues[textEl.id].splice(index, 1);
            syncQueuedFilesToInput(textEl, inputIds);
            renderQueuedFiles(textEl, inputIds);
        };
        row.appendChild(name);
        row.appendChild(remove);
        list.appendChild(row);
    });
}
function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    if(!uploadFileQueues[textId]){ uploadFileQueues[textId] = []; }
    const input = document.getElementById(inputId);
    const incoming = input && input.files ? Array.from(input.files) : [];
    incoming.forEach(file => {
        if(uploadFileQueues[textId].length < 5){
            uploadFileQueues[textId].push(makeShortLocalImageName(file));
        }
    });
    if(incoming.length > 0 && uploadFileQueues[textId].length >= 5 && incoming.length > 5){
        alert('本次最多上传5张图片');
    }
    syncQueuedFilesToInput(textEl, inputIds);
    renderQueuedFiles(textEl, inputIds);
    if(dialogId){ closeUploadChooser(dialogId); }
}
function beginEdit(formId, buttonId){ const form = document.getElementById(formId); if(!form){ return false; } const fields = form.querySelectorAll('.edit-field'); fields.forEach(el => { el.disabled = false; el.classList.remove('readonly'); }); const button = document.getElementById(buttonId); if(button){ button.textContent = '确认'; button.setAttribute('data-mode', 'save'); } return false; }
function handleEditOrSave(formId, buttonId){ const button = document.getElementById(buttonId); if(!button){ return false; } const mode = button.getAttribute('data-mode') || 'edit'; if(mode === 'save'){ const form = document.getElementById(formId); if(form){ if(form.requestSubmit){ form.requestSubmit(); } else { form.submit(); } } return false; } return beginEdit(formId, buttonId); }

const deleteModalState = { onOk: null, onCancel: null };
function closeDeleteModal(){
    const modal = document.getElementById('delete-modal');
    if(modal){ modal.classList.remove('show'); }
    const inputWrap = document.getElementById('delete-modal-input-wrap');
    const input = document.getElementById('delete-modal-input');
    const error = document.getElementById('delete-modal-error');
    const cancelBtn = document.getElementById('delete-modal-cancel');
    const okBtn = document.getElementById('delete-modal-ok');
    if(inputWrap){ inputWrap.classList.remove('show'); }
    if(input){ input.value = ''; }
    if(error){ error.textContent = ''; }
    if(cancelBtn){ cancelBtn.style.display = 'inline-block'; }
    if(okBtn){ okBtn.textContent = '确定'; }
    deleteModalState.onOk = null;
    deleteModalState.onCancel = null;
}
function openDeleteModal(config){
    const modal = document.getElementById('delete-modal');
    const title = document.getElementById('delete-modal-title');
    const text = document.getElementById('delete-modal-text');
    const inputWrap = document.getElementById('delete-modal-input-wrap');
    const input = document.getElementById('delete-modal-input');
    const error = document.getElementById('delete-modal-error');
    const cancelBtn = document.getElementById('delete-modal-cancel');
    const okBtn = document.getElementById('delete-modal-ok');
    if(!modal || !title || !text || !inputWrap || !input || !error || !cancelBtn || !okBtn){ return; }
    title.textContent = config.title || '提示';
    text.textContent = config.text || '';
    error.textContent = '';
    input.value = '';
    if(config.showPin){
        inputWrap.classList.add('show');
        window.setTimeout(() => { try { input.focus(); } catch(e){} }, 30);
    }else{
        inputWrap.classList.remove('show');
    }
    cancelBtn.style.display = config.hideCancel ? 'none' : 'inline-block';
    okBtn.textContent = config.okText || '确定';
    deleteModalState.onOk = typeof config.onOk === 'function' ? config.onOk : null;
    deleteModalState.onCancel = typeof config.onCancel === 'function' ? config.onCancel : null;
    modal.classList.add('show');
}
function handleDeleteModalCancel(){
    const fn = deleteModalState.onCancel;
    closeDeleteModal();
    if(fn){ fn(); }
}
function handleDeleteModalOk(){
    if(!deleteModalState.onOk){ closeDeleteModal(); return; }
    const result = deleteModalState.onOk();
    if(result !== false){ closeDeleteModal(); }
}
function showDeleteNotice(message){
    openDeleteModal({ title:'提示', text: message || '请确认操作', hideCancel: true, onOk: function(){ return true; } });
}
function showDeletePinDialog(onSubmit){
    openDeleteModal({
        title:'Pin码确认',
        text:'请输入4位Pin码',
        showPin:true,
        onOk:function(){
            const input = document.getElementById('delete-modal-input');
            const error = document.getElementById('delete-modal-error');
            const pin = (input && input.value ? input.value : '').trim();
            if(!pin){ if(error){ error.textContent = '请输入4位Pin码'; } return false; }
            if(pin !== '0819'){ if(error){ error.textContent = 'Pin码错误'; } return false; }
            if(typeof onSubmit === 'function'){ onSubmit(pin); }
            return true;
        }
    });
}
function openDeleteFlow(message, onPinConfirmed){
    openDeleteModal({
        title:'删除确认',
        text: message || '确认删除吗？',
        onOk:function(){
            showDeletePinDialog(onPinConfirmed);
            return false;
        }
    });
    return false;
}

function requestDeleteWithPin(formId, message){
    return openDeleteFlow(message || '确认删除吗？', function(pin){
        const form = document.getElementById(formId);
        if(!form){ return; }
        const pinInput = form.querySelector('input[name="delete_pin"]');
        if(pinInput){ pinInput.value = pin; }
        form.submit();
    });
}

</script>
</head>
<body>
<div class="wrap">
    {% if page_title != "电缆详情" %}<div class="card"><div class="switch-row"><a href="/"><button type="button" class="tab-btn">资产查询</button></a><a href="/cable"><button type="button" class="tab-btn tab-active-cable">电缆查询</button></a></div></div>{% endif %}
    <div class="card"><div><strong>当前用户：</strong>{{ user_display }}</div></div>
    <div class="card">
        <div class="title-row"><h2>{{ page_title }}</h2><a href="/cable"><button type="button" class="btn-back">返回</button></a></div>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}
        <form id="cable-form" method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row"><label>电缆编号（可空）</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="cable_no" value="{{ form_data.cable_no }}"></div>
                <div class="row"><label>电缆名称</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="name" value="{{ form_data.name }}"></div>
                <div class="row"><label>规格</label><select class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="spec"><option value="">请选择</option>{% for s in specs %}<option value="{{ s }}" {% if form_data.spec == s %}selected{% endif %}>{{ s }}</option>{% endfor %}</select></div>
                <div class="row"><label>责任人</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="owner" value="{{ form_data.owner }}"></div>
                <div class="row"><label>位置</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="location" value="{{ form_data.location }}" list="detail-location-options"><datalist id="detail-location-options">{% for item in shelf_locations %}<option value="{{ item }}"></option>{% endfor %}</datalist></div>
                <div class="row"><label>状态</label><select class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="status"><option value="">请选择</option>{% for s in statuses %}<option value="{{ s }}" {% if form_data.status == s %}selected{% endif %}>{{ s }}</option>{% endfor %}</select></div>
                <div class="row"><label>上传图片（最多5张）</label><div class="upload-actions"><button type="button" class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} onclick=\"openUploadChooser('cable-upload-choice-dialog', this)\">上传图片</button></div><div id="cable-image-files-text" class="file-list" data-inputs="cable-camera-files,cable-file-files">未选择图片</div><div id="cable-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('cable-upload-choice-dialog');}"><div class="upload-dialog-card"><div class="upload-dialog-title">请选择上传方式</div><div class="upload-dialog-actions"><label class="upload-choice-file upload-choice-camera">拍照<input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="cable-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('cable-camera-files', 'cable-image-files-text', 'cable-upload-choice-dialog')"></label><label class="upload-choice-file upload-choice-local">本地上传<input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="cable-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('cable-file-files', 'cable-image-files-text', 'cable-upload-choice-dialog')"></label><button type="button" class="btn-cancel" onclick="closeUploadChooser('cable-upload-choice-dialog')">取消</button></div></div></div></div>
            </div>
            {% if location_detail_url %}<div class="link-row">当前位置对应货架：<a href="{{ location_detail_url }}">{{ form_data.location }}</a></div>{% endif %}
            {% if images is not none %}<div class="row"><label>当前图片</label>{% if images %}{% for img in images %}<div class="image-row"><a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a><label style="display:flex;align-items:center;gap:6px;font-weight:normal;"><input class="edit-field readonly" disabled type="checkbox" name="delete_cable_image_ids" value="{{ img.id }}"> 删除</label></div>{% endfor %}<div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“确认”。</div>{% else %}<div>暂无图片</div>{% endif %}</div>{% endif %}
            <div class="row"><label>备注</label><textarea class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="remark">{{ form_data.remark }}</textarea></div>
            {% if readonly %}
            <div class="detail-actions">
                {% if can_manage %}
                <button type="button" id="cable-form-edit" data-mode="edit" onclick="handleEditOrSave('cable-form', 'cable-form-edit')">修改</button>
                <button type="button" class="btn-red" onclick="requestDeleteWithPin('cable-delete-form', '确认删除该电缆？')">删除</button>
                {% else %}
                <button type="button" class="btn-disabled" disabled>修改</button>
                <button type="button" class="btn-disabled" disabled>删除</button>
                {% endif %}
            </div>
            {% else %}
            <div class="detail-actions"><button type="submit" class="btn-green">保存</button><a href="/cable"><button type="button" class="btn-back">返回</button></a></div>
            {% endif %}
        </form>
        {% if readonly %}<form id="cable-delete-form" method="post" action="/cable/{{ cable_id }}/delete" style="display:none;"><input type="hidden" name="delete_pin" value=""></form>{% endif %}
    </div>

<div id="delete-modal" class="simple-modal" onclick="if(event.target===this){closeDeleteModal();}">
    <div class="simple-modal-card">
        <div id="delete-modal-title" class="simple-modal-title">删除确认</div>
        <div id="delete-modal-text" class="simple-modal-text">确认删除吗？</div>
        <div id="delete-modal-input-wrap" class="simple-modal-input-wrap">
            <input id="delete-modal-input" class="simple-modal-input" type="password" inputmode="numeric" maxlength="4" placeholder="请输入4位Pin码">
        </div>
        <div id="delete-modal-error" class="simple-modal-error"></div>
        <div class="simple-modal-actions">
            <button type="button" id="delete-modal-cancel" class="simple-modal-cancel" onclick="handleDeleteModalCancel();">取消</button>
            <button type="button" id="delete-modal-ok" class="simple-modal-ok" onclick="handleDeleteModalOk();">确定</button>
        </div>
    </div>
</div>

</div>
</body>
</html>
"""


SHELF_DETAIL_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>货架详情</title>
<style>
:root{--bg:#edf1f5;--card:#ffffff;--line:#d9e1ea;--text:#233243;--muted:#677383;--primary:#5f6f82;--primary-soft:#eef2f6;--green:#2f7d67;--orange:#b88347;--danger:#dc2626;}
*{box-sizing:border-box;}
body{font-family:Arial,sans-serif;margin:0;color:var(--text);background:linear-gradient(180deg,#edf1f5 0%,#f7f8fa 34%,#f1f4f8 100%);}
.wrap{max-width:1100px;margin:auto;padding:16px 14px 26px;}
.card{background:rgba(255,255,255,.94);border:1px solid rgba(210,218,227,.98);border-radius:22px;padding:18px 18px 16px;margin-bottom:16px;box-shadow:0 16px 40px rgba(27,39,53,.07);backdrop-filter:blur(10px);}
h2{margin:0 0 14px;font-size:26px;color:#263646;letter-spacing:.4px;}
a{color:#5a6675;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:600;color:#5e6977;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px 12px;font-size:16px;border-radius:12px;border:1px solid #cfd6df;transition:all .2s ease;}
input[type=text],input[type=date],input[type=password],select,textarea{background:#fbfdff;}
input[type=text]:focus,input[type=date]:focus,input[type=password]:focus,select:focus,textarea:focus{outline:none;border-color:#9ba8b8;box-shadow:0 0 0 4px rgba(95,111,130,.12);background:#fff;}
input[type=file]{background:#fff;}
input[type=checkbox]{accent-color:#5f6f82;width:16px;height:16px;}
textarea{min-height:90px;resize:vertical;}
button{background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;border:none;cursor:pointer;font-weight:700;box-shadow:0 8px 18px rgba(46,58,74,.14);}
button:hover{transform:translateY(-1px);box-shadow:0 12px 22px rgba(46,58,74,.16);}
button:disabled{cursor:not-allowed;transform:none;box-shadow:none;}
.btn-green{background:linear-gradient(135deg,#4f8d78 0%,#2f6b58 100%);}
.btn-gray{background:linear-gradient(135deg,#8b95a1 0%,#697380 100%);}
.btn-orange{background:linear-gradient(135deg,#c39a68 0%,#a9753a 100%);}
.btn-red{background:linear-gradient(135deg,#d06f72 0%,#b34f55 100%);}
.btn-back{background:linear-gradient(135deg,#8b2f45 0%,#6b1f33 100%);box-shadow:0 10px 18px rgba(139,47,69,.22);}
.btn-back:hover{background:linear-gradient(135deg,#9c3850 0%,#77243a 100%);}
.btn-disabled{background:#adb5bd !important;color:#fff !important;cursor:not-allowed;pointer-events:none;box-shadow:none !important;}
.readonly{background:#eef2f6;color:#677383;border-color:#d8e0e8;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.table-wrap{overflow-x:auto;border:1px solid #d7dee7;border-radius:16px;background:#fff;box-shadow:inset 0 1px 0 rgba(255,255,255,.8),0 8px 22px rgba(27,39,53,.04);}
table{width:100%;border-collapse:separate;border-spacing:0;min-width:780px;}th,td{padding:11px 10px;text-align:left;border:none;border-right:1px solid #e7eef7;border-bottom:1px solid #e7eef7;vertical-align:middle;font-size:15px;font-weight:400;color:#6f7b88;}th:last-child,td:last-child{border-right:none;}thead th{background:linear-gradient(180deg,#fafbfd 0%,#eff2f6 100%);color:#5e6977;font-weight:600;white-space:nowrap;}tbody tr:last-child td{border-bottom:none;}tbody tr:nth-child(odd) td{background:#fbfcfd;}tbody tr:nth-child(even) td{background:#f8fafb;}tbody tr:hover td{background:#eff3f7;}.msg{color:#15803d;margin-bottom:12px;font-weight:700;background:#eef8f3;border:1px solid #cfe4d9;border-radius:12px;padding:10px 12px;}.err{color:#dc2626;margin-bottom:12px;font-weight:700;background:#fbf1f2;border:1px solid #e7cfd3;border-radius:12px;padding:10px 12px;}.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:10px;padding:10px 12px;background:#fafbfd;border:1px solid #dde4ec;border-radius:12px;}.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}.upload-actions button{width:auto;min-width:120px;}.file-list{margin-top:8px;color:#66788a;font-size:14px;word-break:break-all;line-height:1.6;}
.selected-file-list{display:flex;flex-direction:column;gap:6px;margin-top:8px;}
.selected-file-item{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 10px;background:#fff;border:1px solid #e0e7ef;border-radius:10px;color:#506172;line-height:1.35;}
.selected-file-name{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.selected-file-remove{width:28px;min-width:28px;height:28px;padding:0;border:none;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:#dc2626;color:#fff;font-size:20px;font-weight:800;line-height:1;box-shadow:0 4px 10px rgba(220,38,38,.18);}
.selected-file-remove:hover{background:#b91c1c;transform:none;}.upload-dialog{position:fixed;left:0;top:0;width:100vw;height:100dvh;background:transparent;display:none;z-index:9999;overflow:hidden;} .upload-dialog.show{display:block;} .upload-dialog-card{position:fixed;left:50%;top:56%;transform:translate(-50%,-50%);width:min(320px,calc(100vw - 24px));max-width:320px;background:#fff;border-radius:16px;padding:18px;box-shadow:0 20px 40px rgba(15,23,42,.22);z-index:10000;}.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;color:#163047;}.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}.upload-dialog-actions button{width:100%;}.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:12px;border:1px solid #cfd9e6;background:linear-gradient(135deg,#718194 0%,#556274 100%);color:#fff;text-align:center;overflow:hidden;box-shadow:0 8px 18px rgba(46,58,74,.14);}.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}.upload-dialog-actions .btn-cancel{background:linear-gradient(135deg,#7b8794 0%,#5f6b77 100%);}.shelf-cable-table{table-layout:fixed;min-width:920px;}.shelf-cable-table th:nth-child(1),.shelf-cable-table td:nth-child(1){width:15%;min-width:138px;}.shelf-cable-table th:nth-child(2),.shelf-cable-table td:nth-child(2){width:16%;}.shelf-cable-table th:nth-child(3),.shelf-cable-table td:nth-child(3){width:10%;}.shelf-cable-table th:nth-child(4),.shelf-cable-table td:nth-child(4){width:10%;}.shelf-cable-table th:nth-child(5),.shelf-cable-table td:nth-child(5){width:8%;}.shelf-cable-table th:nth-child(1),.shelf-cable-table td:nth-child(1),.shelf-cable-table th:nth-child(2),.shelf-cable-table td:nth-child(2){white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}.title-row{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:nowrap;margin-bottom:14px;}
.title-row h2{margin:0;}
.title-row .btn-back{width:auto;min-width:108px;flex:0 0 auto;}
.detail-actions{display:flex;gap:82px;flex-wrap:wrap;align-items:center;margin-top:12px;}
.detail-actions button{width:auto;min-width:120px;}
.link-row{margin-top:-2px;margin-bottom:12px;color:#607284;font-size:14px;}
.muted{color:var(--muted);font-size:14px;line-height:1.6;}.num-link{color:#4e5865;text-decoration:underline;text-decoration-thickness:1px;text-underline-offset:2px;font-weight:400;font-size:18px;}.num-link:hover{color:#3f4853;text-decoration:underline;}
.section-head{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;}

.switch-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.switch-row .tab-btn{background:linear-gradient(180deg,#f9fbfd 0%,#eef3f7 100%);color:#677383;border:1px solid #d8e0e8;box-shadow:inset 0 1px 0 rgba(255,255,255,.86),0 5px 12px rgba(63,78,96,.06);font-weight:800;letter-spacing:.2px;}
.switch-row .tab-btn:hover{background:linear-gradient(180deg,#ffffff 0%,#edf3f8 100%);color:#4d5b6b;border-color:#c8d4df;box-shadow:inset 0 1px 0 rgba(255,255,255,.92),0 10px 18px rgba(63,78,96,.1);}
.switch-row .tab-btn.tab-active-cable{background:#ead8aa;color:#6b5630;border-color:#d8c18a;box-shadow:inset 0 1px 0 rgba(255,255,255,.62),0 10px 18px rgba(161,138,84,.16);text-shadow:none;}
.switch-row .tab-btn.tab-active-cable:hover{background:#e4cf9a;color:#5f4c2b;border-color:#ccb277;box-shadow:inset 0 1px 0 rgba(255,255,255,.62),0 12px 20px rgba(161,138,84,.2);transform:translateY(-1px);}
.simple-modal{position:fixed;inset:0;background:rgba(15,23,42,.52);display:none;align-items:center;justify-content:center;z-index:10000;padding:18px;}
.simple-modal.show{display:flex;}
.simple-modal-card{width:100%;max-width:380px;background:#fff;border-radius:18px;padding:22px 20px;box-shadow:0 18px 50px rgba(15,23,42,.28);}
.simple-modal-title{font-size:22px;font-weight:800;color:#111827;text-align:center;margin-bottom:10px;}
.simple-modal-text{font-size:18px;font-weight:600;line-height:1.7;color:#1f2937;text-align:center;margin-bottom:14px;}
.simple-modal-input-wrap{display:none;margin-bottom:12px;}
.simple-modal-input-wrap.show{display:block;}
.simple-modal-input{width:100%;box-sizing:border-box;padding:12px 14px;font-size:18px;font-weight:700;border-radius:12px;border:2px solid #9ca3af;color:#111827;}
.simple-modal-input:focus{outline:none;border-color:#374151;box-shadow:0 0 0 3px rgba(55,65,81,.12);}
.simple-modal-error{min-height:24px;font-size:16px;font-weight:700;color:#b91c1c;text-align:center;margin-bottom:6px;}
.simple-modal-actions{display:flex;gap:12px;justify-content:center;margin-top:6px;}
.simple-modal-actions button{width:auto;min-width:118px;padding:11px 18px;font-size:17px;font-weight:700;border-radius:12px;border:none;}
.simple-modal-cancel{background:#6b7280;color:#fff;}
.simple-modal-ok{background:#b91c1c;color:#fff;}
@media (max-width:768px){.wrap{padding:14px 12px 20px;}.card{padding:16px;margin-bottom:14px;border-radius:14px;}.grid{grid-template-columns:1fr;}h2{font-size:22px;}.switch-row{gap:8px;}.switch-row a{flex:1 1 160px;}input,select,textarea,button{padding:10px;font-size:16px;border-radius:8px;}.upload-choice-file{padding:10px;font-size:16px;border-radius:8px;}.upload-dialog-card{border-radius:14px;padding:16px;width:min(320px,calc(100vw - 20px));top:58%;}.simple-modal{padding:16px;}.simple-modal-card{border-radius:16px;padding:18px 16px;}.simple-modal-title{font-size:20px;}.simple-modal-text{font-size:16px;}.simple-modal-input{padding:10px 12px;font-size:16px;border-radius:8px;}.simple-modal-actions button{min-width:108px;padding:10px 16px;font-size:16px;border-radius:10px;}}
</style>
<script>
function openUploadChooser(dialogId, triggerEl){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.add('show'); } }
function closeUploadChooser(dialogId){ const dialog = document.getElementById(dialogId); if(dialog){ dialog.classList.remove('show'); } }

const uploadFileQueues = {};
function makeShortLocalImageName(file){
    const now = new Date();
    const pad = value => String(value).padStart(2, '0');
    const timePart = `${now.getFullYear()}${pad(now.getMonth() + 1)}${pad(now.getDate())}${pad(now.getHours())}${pad(now.getMinutes())}${pad(now.getSeconds())}`;
    const randomPart = Math.random().toString(36).slice(2, 8).padEnd(6, '0').slice(0, 6);
    const originalName = file && file.name ? file.name : '';
    const extMatch = originalName.match(/\\.([A-Za-z0-9]+)$/);
    const ext = extMatch ? extMatch[1].toLowerCase() : ((file.type || '').split('/')[1] || 'jpg').toLowerCase();
    try{
        return new File([file], `${timePart}_${randomPart}.${ext}`, {type: file.type || 'image/jpeg', lastModified: file.lastModified || Date.now()});
    }catch(e){
        return file;
    }
}
function syncQueuedFilesToInput(textEl, inputIds){
    const dt = new DataTransfer();
    (uploadFileQueues[textEl.id] || []).forEach(file => dt.items.add(file));
    inputIds.forEach((id, index) => {
        const input = document.getElementById(id);
        if(!input){ return; }
        input.files = index === 0 ? dt.files : new DataTransfer().files;
        input.value = '';
    });
}
function renderQueuedFiles(textEl, inputIds){
    const queue = uploadFileQueues[textEl.id] || [];
    if(queue.length === 0){
        textEl.textContent = '未选择图片';
        return;
    }
    textEl.innerHTML = `<div>已选择 ${queue.length} 张（最多5张）：</div><div class="selected-file-list"></div>`;
    const list = textEl.querySelector('.selected-file-list');
    queue.forEach((file, index) => {
        const row = document.createElement('div');
        row.className = 'selected-file-item';
        const name = document.createElement('span');
        name.className = 'selected-file-name';
        name.textContent = file.name || `图片${index + 1}`;
        const remove = document.createElement('button');
        remove.type = 'button';
        remove.className = 'selected-file-remove';
        remove.textContent = '×';
        remove.setAttribute('aria-label', '删除该图片');
        remove.onclick = function(){
            uploadFileQueues[textEl.id].splice(index, 1);
            syncQueuedFilesToInput(textEl, inputIds);
            renderQueuedFiles(textEl, inputIds);
        };
        row.appendChild(name);
        row.appendChild(remove);
        list.appendChild(row);
    });
}
function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    if(!uploadFileQueues[textId]){ uploadFileQueues[textId] = []; }
    const input = document.getElementById(inputId);
    const incoming = input && input.files ? Array.from(input.files) : [];
    incoming.forEach(file => {
        if(uploadFileQueues[textId].length < 5){
            uploadFileQueues[textId].push(makeShortLocalImageName(file));
        }
    });
    if(incoming.length > 0 && uploadFileQueues[textId].length >= 5 && incoming.length > 5){
        alert('本次最多上传5张图片');
    }
    syncQueuedFilesToInput(textEl, inputIds);
    renderQueuedFiles(textEl, inputIds);
    if(dialogId){ closeUploadChooser(dialogId); }
}
function enableEdit(formId){ const form = document.getElementById(formId); const fields = form.querySelectorAll('.edit-field'); fields.forEach(el => { el.disabled = false; el.classList.remove('readonly'); }); document.getElementById(formId + '-save').style.display = 'inline-block'; document.getElementById(formId + '-edit').style.display = 'none'; }
</script>
</head>
<body>
<div class="wrap">
    <div class="card"><div><strong>当前用户：</strong>{{ user_display }}</div></div>
    <div class="card">
        <div class="title-row"><h2>货架详情</h2><a href="/cable"><button type="button" class="btn-back">返回</button></a></div>
        <div class="muted" style="margin-bottom:10px;">货架就是位置，修改会同步更新所有相关电缆。</div>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}
        <form id="shelf-form" method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row"><label>货架位置</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="location" value="{{ form_data.location }}"></div>
                <div class="row"><label>上传货架图片（最多5张）</label><div class="upload-actions"><button type="button" class="{{ 'edit-field readonly' if readonly else '' }} {% if not can_manage %}btn-disabled{% endif %}" {% if readonly or not can_manage %}disabled{% endif %} onclick=\"openUploadChooser('shelf-upload-choice-dialog', this)\">上传图片</button></div><div id="shelf-image-files-text" class="file-list" data-inputs="shelf-camera-files,shelf-file-files">未选择图片</div><div id="shelf-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('shelf-upload-choice-dialog');}"><div class="upload-dialog-card"><div class="upload-dialog-title">请选择上传方式</div><div class="upload-dialog-actions"><label class="upload-choice-file upload-choice-camera">拍照<input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly or not can_manage %}disabled{% endif %} type="file" id="shelf-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('shelf-camera-files', 'shelf-image-files-text', 'shelf-upload-choice-dialog')"></label><label class="upload-choice-file upload-choice-local">本地上传<input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly or not can_manage %}disabled{% endif %} type="file" id="shelf-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('shelf-file-files', 'shelf-image-files-text', 'shelf-upload-choice-dialog')"></label><button type="button" class="btn-cancel" onclick="closeUploadChooser('shelf-upload-choice-dialog')">取消</button></div></div></div></div>
            </div>
            <div class="row"><label>当前货架图片</label>{% if images %}{% for img in images %}<div class="image-row"><a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a><label style="display:flex;align-items:center;gap:6px;font-weight:normal;"><input class="edit-field readonly" disabled type="checkbox" name="delete_shelf_image_ids" value="{{ img.id }}"> 删除</label></div>{% endfor %}<div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“确认”。</div>{% else %}<div>暂无图片</div>{% endif %}</div>
            <div class="row"><label>备注</label><textarea class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="remark">{{ form_data.remark }}</textarea></div>
            <div class="detail-actions">{% if can_manage %}<button type="button" id="shelf-form-edit" onclick="enableEdit('shelf-form')">修改</button><button type="submit" id="shelf-form-save" style="display:none;">确认</button>{% else %}<button type="button" class="btn-disabled" disabled>修改</button>{% endif %}</div>
        </form>
    </div>
    <div class="card"><div class="section-head"><h2 style="margin:0;">该位置下的电缆</h2>{% if can_manage %}<a href="/cable/new?add_type=电缆&location={{ form_data.location|urlencode }}"><button type="button" class="btn-green" style="width:auto;">+</button></a>{% else %}<button type="button" class="btn-disabled" style="width:auto;" disabled>+</button>{% endif %}</div><div class="table-wrap" style="margin-top:12px;"><table class="shelf-cable-table"><thead><tr><th>电缆编号</th><th>电缆名称</th><th>规格</th><th>责任人</th><th>状态</th><th>备注</th></tr></thead><tbody>{% for item in cable_rows %}<tr class="result-row result-row-cable"><td><a class="num-link" href="{{ item.detail_url }}">{{ item.cable_no }}</a></td><td>{{ item.name }}</td><td>{{ item.spec }}</td><td>{{ item.owner }}</td><td>{{ item.status }}</td><td>{{ item.remark }}</td></tr>{% endfor %}{% if not cable_rows %}<tr class="empty-row"><td colspan="6">该位置下暂无电缆</td></tr>{% endif %}</tbody></table></div></div>
</div>
</body>
</html>
"""


def register_cable_routes(app):
    with app.app_context():
        Cable.__table__.create(bind=db.engine, checkfirst=True)
        CableImage.__table__.create(bind=db.engine, checkfirst=True)
        CableShelf.__table__.create(bind=db.engine, checkfirst=True)
        CableShelfImage.__table__.create(bind=db.engine, checkfirst=True)
        backfill_cable_shelves()

    @app.route("/cable", defaults={"scan_code": ""}, methods=["GET"])
    @app.route("/cable<scan_code>", methods=["GET"])
    def search_cables(scan_code=""):
        path_scan_code = normalize_text(scan_code)
        prefill_cable_no = extract_scan_prefill_code(path_scan_code)
        keyword = normalize_text(request.args.get("keyword"))
        if not keyword:
            keyword = prefill_cable_no or extract_scan_keyword(path_scan_code)

        spec_filter = normalize_text(request.args.get("spec_filter"))
        status_filter = normalize_cable_status(request.args.get("status_filter"))
        searched = True if path_scan_code else (normalize_text(request.args.get("searched")) == "1")
        if keyword and not searched:
            searched = True

        per_page = normalize_text(request.args.get("per_page")) or "30"
        page = normalize_text(request.args.get("page")) or "1"

        try:
            per_page = int(per_page)
        except Exception:
            per_page = 30
        if per_page not in [30, 50, 100]:
            per_page = 30

        try:
            page = int(page)
        except Exception:
            page = 1

        all_rows = build_cable_rows(keyword=keyword, searched=searched)
        if searched and spec_filter:
            all_rows = [row for row in all_rows if normalize_text(row.get("spec")) == spec_filter]
        if searched and status_filter:
            all_rows = [row for row in all_rows if normalize_text(row.get("status")) == status_filter]

        all_selection_meta = [
            {"selected_value": row.get("selected_value", ""), "row_type": row.get("row_type", "")}
            for row in all_rows
            if row.get("selected_value")
        ]

        total = len(all_rows)
        start = (page - 1) * per_page
        end = start + per_page
        rows = all_rows[start:end]
        total_pages = (total + per_page - 1) // per_page if total else 1

        error = ""
        if searched and total == 0:
            if keyword and (spec_filter or status_filter):
                error = "未找到符合筛选条件的电缆"
            elif keyword:
                error = "未找到对应电缆"
            elif spec_filter or status_filter:
                error = "未找到符合筛选条件的电缆"

        return render_template_string(
            CABLE_SEARCH_HTML,
            current_user=current_user,
            can_manage=has_manage_access(),
            user_display=get_user_display_name(),
            keyword=keyword,
            spec_filter=spec_filter,
            status_filter=status_filter,
            specs=CABLE_SPECS,
            statuses=CABLE_STATUSES,
            searched=searched,
            per_page=per_page,
            page=page,
            total=total,
            total_pages=total_pages,
            rows=rows,
            error=error,
            scan_code=path_scan_code,
            prefill_cable_no=prefill_cable_no,
            all_selection_meta=all_selection_meta,
        )

    @app.route("/cable/import", methods=["POST"])
    def import_cables():
        guard = ensure_manage_access()
        if guard:
            return guard
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            return redirect(url_for("search_cables"))

        try:
            import_cables_from_excel(excel_file)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量上传失败：{str(e)}"

        return redirect(url_for("search_cables"))

    @app.route("/cable/export", methods=["POST"])
    def export_selected_cables():
        selected_items = request.form.getlist("selected_items")
        wb = Workbook()
        ws = wb.active
        ws.title = "电缆导出"
        ws.append(["电缆编号", "电缆名称", "规格", "责任人", "位置", "状态", "备注"])

        for item in selected_items:
            row_type = "cable"
            row_id_text = normalize_text(item)
            if ":" in row_id_text:
                row_type, row_id_text = row_id_text.split(":", 1)
            if row_type != "cable":
                continue
            try:
                cable_id = int(row_id_text)
            except Exception:
                continue
            obj = Cable.query.get(cable_id)
            if obj:
                ws.append([
                    get_display_cable_no(obj.cable_no),
                    obj.name or "",
                    obj.spec or "",
                    obj.owner or "",
                    obj.location or "",
                    obj.status or "",
                    obj.remark or "",
                ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"cable_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/cable/delete_selected", methods=["POST"])
    def delete_selected_cables():
        guard = ensure_manage_access()
        if guard:
            return guard
        selected_items = request.form.getlist("selected_items")
        delete_pin = normalize_text(request.form.get("delete_pin"))
        keyword = normalize_text(request.form.get("keyword"))
        spec_filter = normalize_text(request.form.get("spec_filter"))
        status_filter = normalize_text(request.form.get("status_filter"))

        if delete_pin != "0819":
            return "批量删除失败：Pin码错误"

        cable_ids = []
        shelf_ids = []
        for item in selected_items:
            row_type = "cable"
            row_id_text = normalize_text(item)
            if ":" in row_id_text:
                row_type, row_id_text = row_id_text.split(":", 1)
            try:
                row_id = int(row_id_text)
            except Exception:
                continue
            if row_type == "shelf":
                shelf_ids.append(row_id)
            else:
                cable_ids.append(row_id)

        try:
            cables = Cable.query.filter(Cable.id.in_(cable_ids)).order_by(Cable.id.asc()).all() if cable_ids else []
            shelves = CableShelf.query.filter(CableShelf.id.in_(shelf_ids)).order_by(CableShelf.id.asc()).all() if shelf_ids else []
            occupied_locations = {
                normalize_location(row[0])
                for row in db.session.query(Cable.location).filter(Cable.location.isnot(None)).distinct().all()
                if normalize_location(row[0])
            }

            for obj in cables:
                delete_cable_with_files(obj)

            blocked_shelves = []
            for shelf in shelves:
                shelf_name = normalize_location(shelf.shelf_name)
                if shelf_name and shelf_name in occupied_locations:
                    blocked_shelves.append(shelf_name)
                    continue
                delete_shelf_with_files(shelf)

            if blocked_shelves:
                db.session.rollback()
                return "批量删除失败：以下货架下仍有电缆，不能删除：" + "，".join(blocked_shelves)

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量删除失败：{str(e)}"

        return redirect(url_for("search_cables", searched=1, keyword=keyword, spec_filter=spec_filter, status_filter=status_filter, page=1))

    @app.route("/cable/new", methods=["GET", "POST"])
    def cable_new():
        guard = ensure_manage_access()
        if guard:
            return guard
        error = ""
        message = ""
        prefill_cable_no = normalize_cable_no(request.args.get("cable_no")) or normalize_cable_no(extract_scan_prefill_code()) or normalize_cable_no(extract_scan_keyword())
        prefill_location = normalize_location(request.args.get("location"))

        form_data = {
            "cable_no": prefill_cable_no,
            "name": "",
            "spec": "高频",
            "owner": "",
            "location": prefill_location,
            "status": "在库",
            "remark": "",
        }

        if request.method == "POST":
            manage_guard = ensure_manage_access()
            if manage_guard:
                return manage_guard
            cable_no = normalize_cable_no(request.form.get("cable_no"))
            name = normalize_text(request.form.get("name"))
            spec = normalize_text(request.form.get("spec"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_location(request.form.get("location"))
            status = normalize_cable_status(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")

            form_data = {
                "cable_no": cable_no,
                "name": name,
                "spec": spec,
                "owner": owner,
                "location": location,
                "status": status,
                "remark": remark,
            }

            if not cable_no:
                error = "电缆编号不能为空"
            elif not is_valid_cable_no(cable_no):
                error = "电缆编号格式不正确，应为 DMDL+6位数字，例如 DMDL123456"
            elif not location:
                error = "位置不能为空"
            elif spec and spec not in CABLE_SPECS:
                error = "规格不合法"
            elif status and status not in CABLE_STATUSES:
                error = "状态不合法"
            elif Cable.query.filter_by(cable_no=cable_no).first():
                error = "电缆编号已存在"
            else:
                try:
                    shelf = ensure_shelf(location, auto_create=True)
                    stored_cable_no = cable_no
                    obj = Cable(
                        cable_no=stored_cable_no,
                        name=name or "",
                        spec=normalize_empty_to_none(spec),
                        owner=normalize_empty_to_none(owner),
                        location=normalize_location(shelf.shelf_name) if shelf else normalize_empty_to_none(location),
                        status=normalize_empty_to_none(status),
                        remark=normalize_empty_to_none(remark),
                    )
                    db.session.add(obj)
                    db.session.flush()

                    image_prefix = build_cable_filename_prefix(stored_cable_no, name, obj.location)
                    for file_storage in image_files[:5]:
                        rel = save_uploaded_image(file_storage, "cable", image_prefix)
                        if rel:
                            db.session.add(CableImage(cable_id=obj.id, image_path=rel))

                    trim_cable_images(obj)
                    db.session.commit()
                    return redirect(url_for("cable_detail", cable_id=obj.id))
                except Exception as e:
                    db.session.rollback()
                    error = f"保存失败：{str(e)}"

        return render_template_string(
            CABLE_NEW_HTML,
            current_user=current_user,
            specs=CABLE_SPECS,
            statuses=CABLE_STATUSES,
            form_data=form_data,
            error=error,
            message=message,
            shelf_locations=get_all_shelf_locations(),
        )

    @app.route("/cable/location/<int:shelf_id>", methods=["GET", "POST"])
    def cable_location_detail(shelf_id):
        shelf = CableShelf.query.get_or_404(shelf_id)
        message = ""
        error = ""

        if request.method == "POST":
            manage_guard = ensure_manage_access()
            if manage_guard:
                return manage_guard
            location = normalize_location(request.form.get("location"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")
            delete_image_ids = request.form.getlist("delete_shelf_image_ids")

            if not location:
                error = "货架位置不能为空"
            else:
                existing = CableShelf.query.filter(CableShelf.shelf_name == location, CableShelf.id != shelf.id).first()
                if existing:
                    error = "该货架位置已存在"
                else:
                    old_location = shelf.shelf_name or ""
                    try:
                        shelf.shelf_name = location
                        shelf.remark = normalize_empty_to_none(remark)

                        if old_location != location:
                            Cable.query.filter_by(location=old_location).update({"location": location}, synchronize_session=False)

                        for image_id in delete_image_ids:
                            try:
                                img_id = int(image_id)
                            except Exception:
                                continue
                            img = CableShelfImage.query.filter_by(id=img_id, shelf_id=shelf.id).first()
                            if img:
                                delete_image_file(img.image_path)
                                db.session.delete(img)

                        for file_storage in image_files[:5]:
                            rel = save_uploaded_image(file_storage, "cable_shelf", shelf.shelf_name)
                            if rel:
                                db.session.add(CableShelfImage(shelf_id=shelf.id, image_path=rel))

                        trim_shelf_images(shelf)
                        db.session.commit()
                        message = "货架更新成功"
                    except Exception as e:
                        db.session.rollback()
                        error = f"更新失败：{str(e)}"

        images = get_shelf_images(shelf.id)
        cable_rows = []
        for item in Cable.query.filter_by(location=shelf.shelf_name).order_by(Cable.cable_no.asc(), Cable.id.asc()).all():
            cable_rows.append({
                "id": item.id,
                "cable_no": get_display_cable_no(item.cable_no),
                "name": item.name or "",
                "spec": item.spec or "",
                "owner": item.owner or "",
                "status": normalize_cable_status(item.status),
                "remark": item.remark or "",
                "detail_url": url_for("cable_detail", cable_id=item.id),
            })

        form_data = OrderedDict([
            ("location", normalize_location(shelf.shelf_name)),
            ("remark", shelf.remark or ""),
        ])

        return render_template_string(
            SHELF_DETAIL_HTML,
            current_user=current_user,
            can_manage=has_manage_access(),
            user_display=get_user_display_name(),
            form_data=form_data,
            error=error,
            message=message,
            readonly=True,
            images=images,
            cable_rows=cable_rows,
        )

    @app.route("/cable/<int:cable_id>", methods=["GET", "POST"])
    def cable_detail(cable_id):
        cable = Cable.query.get_or_404(cable_id)
        message = ""
        error = ""

        if request.method == "POST":
            manage_guard = ensure_manage_access()
            if manage_guard:
                return manage_guard
            cable_no = normalize_cable_no(request.form.get("cable_no"))
            name = normalize_text(request.form.get("name"))
            spec = normalize_text(request.form.get("spec"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_location(request.form.get("location"))
            status = normalize_cable_status(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")
            delete_image_ids = request.form.getlist("delete_cable_image_ids")

            if not cable_no:
                error = "电缆编号不能为空"
            elif not is_valid_cable_no(cable_no):
                error = "电缆编号格式不正确，应为 DMDL+6位数字，例如 DMDL123456"
            elif not location:
                error = "位置不能为空"
            elif spec and spec not in CABLE_SPECS:
                error = "规格不合法"
            elif status and status not in CABLE_STATUSES:
                error = "状态不合法"
            else:
                existing = Cable.query.filter(Cable.cable_no == cable_no, Cable.id != cable.id).first() if cable_no else None
                if existing:
                    error = "电缆编号已存在"
                else:
                    shelf = ensure_shelf(location, auto_create=True)
                    try:
                        cable.cable_no = make_stored_cable_no(cable_no, cable.cable_no)
                        cable.name = name or ""
                        cable.spec = normalize_empty_to_none(spec)
                        cable.owner = normalize_empty_to_none(owner)
                        cable.location = normalize_location(shelf.shelf_name) if shelf else normalize_empty_to_none(location)
                        cable.status = normalize_empty_to_none(status)
                        cable.remark = normalize_empty_to_none(remark)

                        for image_id in delete_image_ids:
                            try:
                                img_id = int(image_id)
                            except Exception:
                                continue
                            img = CableImage.query.filter_by(id=img_id, cable_id=cable.id).first()
                            if img:
                                delete_image_file(img.image_path)
                                db.session.delete(img)

                        image_prefix = build_cable_filename_prefix(cable.cable_no, cable.name, cable.location)
                        for file_storage in image_files[:5]:
                            rel = save_uploaded_image(file_storage, "cable", image_prefix)
                            if rel:
                                db.session.add(CableImage(cable_id=cable.id, image_path=rel))

                        trim_cable_images(cable)
                        db.session.commit()
                        message = "电缆更新成功"
                    except Exception as e:
                        db.session.rollback()
                        error = f"更新失败：{str(e)}"

        images = get_cable_images(cable.id)
        shelf = get_shelf_by_location(normalize_location(cable.location))
        form_data = OrderedDict([
            ("cable_no", get_display_cable_no(cable.cable_no)),
            ("name", cable.name or ""),
            ("spec", cable.spec or ""),
            ("owner", cable.owner or ""),
            ("location", normalize_location(cable.location)),
            ("status", normalize_cable_status(cable.status)),
            ("remark", cable.remark or ""),
        ])

        return render_template_string(
            CABLE_FORM_HTML,
            current_user=current_user,
            can_manage=has_manage_access(),
            user_display=get_user_display_name(),
            page_title="电缆详情",
            specs=CABLE_SPECS,
            statuses=CABLE_STATUSES,
            form_data=form_data,
            error=error,
            message=message,
            readonly=True,
            images=images,
            shelf_locations=get_all_shelf_locations(),
            location_detail_url=url_for("cable_location_detail", shelf_id=shelf.id) if shelf else "",
            cable_id=cable.id,
        )

    @app.route("/cable/<int:cable_id>/delete", methods=["POST"])
    def delete_cable(cable_id):
        guard = ensure_manage_access()
        if guard:
            return guard
        cable = Cable.query.get_or_404(cable_id)
        delete_pin = normalize_text(request.form.get("delete_pin"))
        if delete_pin != "0819":
            return "删除失败：Pin码错误"
        try:
            delete_cable_with_files(cable)
            db.session.commit()
            return redirect(url_for("search_cables"))
        except Exception as e:
            db.session.rollback()
            return f"删除失败：{str(e)}"
