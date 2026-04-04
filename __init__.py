from datetime import datetime, date
from io import BytesIO
import os
import re
import uuid

from flask import request, redirect, url_for, render_template_string, send_file, send_from_directory, abort, jsonify
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash
from sqlalchemy import or_
from openpyxl import Workbook, load_workbook

from app import db
from app.models import (
    User, Asset, Accessory, DictOption,
    AssetImage, AccessoryImage
)
from config import Config


ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

try:
    from .ocr_recognizer import (
        SCAN_TIMEOUT_SECONDS,
        ScanTimeoutError,
        extract_group_no_from_label,
        scan_time_limit,
    )
except ImportError:
    from ocr_recognizer import (
        SCAN_TIMEOUT_SECONDS,
        ScanTimeoutError,
        extract_group_no_from_label,
        scan_time_limit,
    )


def get_statuses():
    return DictOption.query.filter_by(
        dict_type="status",
        is_active=True
    ).order_by(DictOption.sort_order.asc()).all()


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_empty_to_none(value):
    value = normalize_text(value)
    return value if value else None


def make_accessory_prefix(value):
    value = normalize_text(value)
    if not value:
        return ""
    return value if value.endswith("-") else f"{value}-"


def validate_required_number_pair(internal_no, group_no, internal_label="内部编号", group_label="集团编号"):
    if not normalize_text(internal_no) and not normalize_text(group_no):
        return f"{internal_label}和{group_label}至少填写一个"
    return ""


def validate_accessory_no_format(internal_no, group_no, internal_label="附属资产内部编号", group_label="附属资产集团编号"):
    pattern = re.compile(r"^[A-Za-z0-9]+-\d+$")
    for value, label in [(normalize_text(internal_no), internal_label), (normalize_text(group_no), group_label)]:
        if value and not pattern.match(value):
            return f"{label}格式不正确，应类似 651411001001-001"
    return ""


def validate_asset_group_no(group_no, label="集团编号"):
    group_no = normalize_text(group_no)
    if group_no and not re.fullmatch(r"\d{18}", group_no):
        return f"{label}必须为18位数字"
    return ""


def validate_accessory_group_no(group_no, label="附属资产集团编号"):
    group_no = normalize_text(group_no)
    if group_no and not re.fullmatch(r"\d{18}-\d+", group_no):
        return f"{label}必须为18位主编号-序号"
    return ""


def find_existing_asset(internal_no="", group_no="", exclude_id=None):
    conditions = []
    internal_no = normalize_text(internal_no)
    group_no = normalize_text(group_no)

    if internal_no:
        conditions.append(Asset.internal_no == internal_no)
    if group_no:
        conditions.append(Asset.group_no == group_no)

    if not conditions:
        return None

    query = Asset.query.filter(or_(*conditions))
    if exclude_id is not None:
        query = query.filter(Asset.id != exclude_id)
    return query.first()


def find_existing_accessory(internal_no="", group_no="", exclude_id=None):
    conditions = []
    internal_no = normalize_text(internal_no)
    group_no = normalize_text(group_no)

    if internal_no:
        conditions.append(Accessory.sub_internal_no == internal_no)
    if group_no:
        conditions.append(Accessory.sub_group_no == group_no)

    if not conditions:
        return None

    query = Accessory.query.filter(or_(*conditions))
    if exclude_id is not None:
        query = query.filter(Accessory.id != exclude_id)
    return query.first()


def find_parent_asset_by_accessory_numbers(internal_no="", group_no=""):
    internal_no = normalize_text(internal_no)
    group_no = normalize_text(group_no)

    if "-" in internal_no:
        base_internal = normalize_text(internal_no.rsplit("-", 1)[0])
        if base_internal:
            parent = Asset.query.filter_by(internal_no=base_internal).first()
            if parent:
                return parent

    if "-" in group_no:
        base_group = normalize_text(group_no.rsplit("-", 1)[0])
        if base_group:
            parent = Asset.query.filter_by(group_no=base_group).first()
            if parent:
                return parent

    return None


def resolve_parent_asset_id(internal_no="", group_no="", fallback_parent_asset_id=None):
    parent = find_parent_asset_by_accessory_numbers(internal_no=internal_no, group_no=group_no)
    if parent:
        return parent.id

    fallback_parent_asset_id = normalize_text(fallback_parent_asset_id)
    if fallback_parent_asset_id.isdigit():
        fallback_parent = Asset.query.get(int(fallback_parent_asset_id))
        if fallback_parent:
            return fallback_parent.id

    return None


def get_asset_related_accessories(asset):
    if not asset:
        return []

    accessory_map = {}

    for item in Accessory.query.filter_by(parent_asset_id=asset.id).all():
        accessory_map[item.id] = item

    conditions = []
    if asset.internal_no:
        conditions.append(Accessory.sub_internal_no.like(f"{asset.internal_no}-%"))
    if asset.group_no:
        conditions.append(Accessory.sub_group_no.like(f"{asset.group_no}-%"))

    if conditions:
        standalone_items = Accessory.query.filter(
            Accessory.parent_asset_id.is_(None),
            or_(*conditions)
        ).all()
        for item in standalone_items:
            accessory_map[item.id] = item

    return sorted(
        accessory_map.values(),
        key=lambda x: (
            normalize_text(x.sub_internal_no) or "~~~~",
            normalize_text(x.sub_group_no) or "~~~~",
            x.id
        )
    )


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_image(file_storage, subdir):
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        return None

    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"

    folder = os.path.join(Config.UPLOAD_FOLDER, subdir)
    os.makedirs(folder, exist_ok=True)

    abs_path = os.path.join(folder, filename)
    file_storage.save(abs_path)

    return f"{subdir}/{filename}"


def delete_image_file(relative_path):
    if not relative_path:
        return
    abs_path = os.path.join(Config.UPLOAD_FOLDER, relative_path)
    if os.path.exists(abs_path):
        os.remove(abs_path)


def trim_asset_images(asset):
    images = AssetImage.query.filter_by(asset_id=asset.id).order_by(AssetImage.created_at.asc(), AssetImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        delete_image_file(old.image_path)
        db.session.delete(old)


def trim_accessory_images(accessory):
    images = AccessoryImage.query.filter_by(accessory_id=accessory.id).order_by(AccessoryImage.created_at.asc(), AccessoryImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        delete_image_file(old.image_path)
        db.session.delete(old)


def delete_accessory_with_files(accessory):
    if not accessory:
        return
    for img in AccessoryImage.query.filter_by(accessory_id=accessory.id).all():
        delete_image_file(img.image_path)
    db.session.delete(accessory)


def delete_asset_with_files(asset):
    if not asset:
        return
    accessories = get_asset_related_accessories(asset)
    for accessory in accessories:
        delete_accessory_with_files(accessory)
    for img in AssetImage.query.filter_by(asset_id=asset.id).all():
        delete_image_file(img.image_path)
    db.session.delete(asset)



def build_search_rows(keyword="", searched=False):
    if not searched:
        return []

    keyword = normalize_text(keyword)
    rows = []

    if not keyword:
        asset_rows = Asset.query.order_by(Asset.internal_no.asc()).all()
        accessory_rows = Accessory.query.order_by(Accessory.sub_internal_no.asc()).all()

        for item in asset_rows:
            rows.append({
                "row_type": "asset",
                "id": item.id,
                "type_text": "主设备",
                "internal_no": item.internal_no or "",
                "group_no": item.group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": "",
                "accessory_count": len(get_asset_related_accessories(item)),
                "detail_url": url_for("asset_detail", asset_id=item.id)
            })

        for item in accessory_rows:
            rows.append({
                "row_type": "accessory",
                "id": item.id,
                "type_text": "配件",
                "internal_no": item.sub_internal_no or "",
                "group_no": item.sub_group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": item.parent_asset_id or "",
                "accessory_count": 0,
                "detail_url": url_for("accessory_detail", accessory_id=item.id)
            })

        rows.sort(key=lambda x: (x["internal_no"] or x["group_no"] or ""))
        return rows

    suffix_keyword = keyword[-6:] if len(keyword) >= 6 else ""

    exact_assets = Asset.query.filter(
        or_(
            Asset.internal_no == keyword,
            Asset.group_no == keyword
        )
    ).order_by(Asset.internal_no.asc()).all()

    suffix_assets = []
    if suffix_keyword:
        suffix_assets = Asset.query.filter(
            or_(
                Asset.internal_no.like(f"%{suffix_keyword}"),
                Asset.group_no.like(f"%{suffix_keyword}")
            )
        ).order_by(Asset.internal_no.asc()).all()

    asset_ids = []
    for a in exact_assets + suffix_assets:
        if a.id not in asset_ids:
            asset_ids.append(a.id)

    owner_assets = Asset.query.filter(
        Asset.owner.like(f"%{keyword}%")
    ).order_by(Asset.internal_no.asc()).all()

    for a in owner_assets:
        if a.id not in asset_ids:
            asset_ids.append(a.id)

    exact_accessories = Accessory.query.filter(
        or_(
            Accessory.sub_internal_no == keyword,
            Accessory.sub_group_no == keyword
        )
    ).all()

    suffix_accessories = []
    if suffix_keyword:
        suffix_accessories = Accessory.query.filter(
            or_(
                Accessory.sub_internal_no.like(f"%{suffix_keyword}"),
                Accessory.sub_group_no.like(f"%{suffix_keyword}-%")
            )
        ).all()

    owner_accessories = Accessory.query.filter(
        Accessory.owner.like(f"%{keyword}%")
    ).all()

    for acc in exact_accessories + suffix_accessories + owner_accessories:
        if acc.parent_asset_id and acc.parent_asset_id not in asset_ids:
            asset_ids.append(acc.parent_asset_id)

    fuzzy_accessories = Accessory.query.filter(
        or_(
            Accessory.sub_internal_no.like(f"{keyword}-%"),
            Accessory.sub_group_no.like(f"{keyword}-%")
        )
    ).all()

    standalone_accessories = []
    standalone_accessory_ids = set()
    for acc in fuzzy_accessories + exact_accessories + suffix_accessories + owner_accessories:
        resolved_parent_id = acc.parent_asset_id or resolve_parent_asset_id(
            internal_no=acc.sub_internal_no,
            group_no=acc.sub_group_no
        )
        if resolved_parent_id:
            if resolved_parent_id not in asset_ids:
                asset_ids.append(resolved_parent_id)
        else:
            if acc.id not in standalone_accessory_ids:
                standalone_accessories.append(acc)
                standalone_accessory_ids.add(acc.id)

    if asset_ids:
        matched_assets = Asset.query.filter(Asset.id.in_(asset_ids)).order_by(Asset.internal_no.asc()).all()

        for asset in matched_assets:
            accessories = get_asset_related_accessories(asset)

            rows.append({
                "row_type": "asset",
                "id": asset.id,
                "type_text": "主设备",
                "internal_no": asset.internal_no or "",
                "group_no": asset.group_no or "",
                "name": asset.name,
                "model": asset.model or "",
                "status": asset.status or "",
                "owner": asset.owner or "",
                "parent_asset_id": "",
                "accessory_count": len(accessories),
                "detail_url": url_for("asset_detail", asset_id=asset.id)
            })

            for item in accessories:
                rows.append({
                    "row_type": "accessory",
                    "id": item.id,
                    "type_text": "配件",
                    "internal_no": item.sub_internal_no or "",
                    "group_no": item.sub_group_no or "",
                    "name": item.name,
                    "model": item.model or "",
                    "status": item.status or "",
                    "owner": item.owner or "",
                    "parent_asset_id": item.parent_asset_id or resolve_parent_asset_id(item.sub_internal_no, item.sub_group_no) or "",
                    "accessory_count": 0,
                    "detail_url": url_for("accessory_detail", accessory_id=item.id)
                })

    existing_row_keys = {(row["row_type"], row["id"]) for row in rows}
    for item in standalone_accessories:
        if ("accessory", item.id) not in existing_row_keys:
            rows.append({
                "row_type": "accessory",
                "id": item.id,
                "type_text": "配件",
                "internal_no": item.sub_internal_no or "",
                "group_no": item.sub_group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": item.parent_asset_id or "",
                "accessory_count": 0,
                "detail_url": url_for("accessory_detail", accessory_id=item.id)
            })

    if not rows:
        standalone_conditions = [
            Accessory.sub_internal_no == keyword,
            Accessory.sub_group_no == keyword,
            Accessory.sub_internal_no.like(f"{keyword}-%"),
            Accessory.sub_group_no.like(f"{keyword}-%"),
            Accessory.owner.like(f"%{keyword}%")
        ]
        if suffix_keyword:
            standalone_conditions.append(Accessory.sub_internal_no.like(f"%{suffix_keyword}"))
            standalone_conditions.append(Accessory.sub_group_no.like(f"%{suffix_keyword}-%"))

        standalone_exact = Accessory.query.filter(
            or_(*standalone_conditions)
        ).order_by(Accessory.sub_internal_no.asc()).all()

        for item in standalone_exact:
            rows.append({
                "row_type": "accessory",
                "id": item.id,
                "type_text": "配件",
                "internal_no": item.sub_internal_no or "",
                "group_no": item.sub_group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": item.parent_asset_id or "",
                "accessory_count": 0,
                "detail_url": url_for("accessory_detail", accessory_id=item.id)
            })

    return rows


    exact_assets = Asset.query.filter(
        or_(
            Asset.internal_no == keyword,
            Asset.group_no == keyword
        )
    ).order_by(Asset.internal_no.asc()).all()

    asset_ids = []
    for a in exact_assets:
        if a.id not in asset_ids:
            asset_ids.append(a.id)

    owner_assets = Asset.query.filter(
        Asset.owner.like(f"%{keyword}%")
    ).order_by(Asset.internal_no.asc()).all()

    for a in owner_assets:
        if a.id not in asset_ids:
            asset_ids.append(a.id)

    exact_accessories = Accessory.query.filter(
        or_(
            Accessory.sub_internal_no == keyword,
            Accessory.sub_group_no == keyword
        )
    ).all()

    owner_accessories = Accessory.query.filter(
        Accessory.owner.like(f"%{keyword}%")
    ).all()

    for acc in exact_accessories + owner_accessories:
        if acc.parent_asset_id and acc.parent_asset_id not in asset_ids:
            asset_ids.append(acc.parent_asset_id)

    fuzzy_accessories = Accessory.query.filter(
        or_(
            Accessory.sub_internal_no.like(f"{keyword}-%"),
            Accessory.sub_group_no.like(f"{keyword}-%")
        )
    ).all()

    standalone_accessories = []
    standalone_accessory_ids = set()
    for acc in fuzzy_accessories + exact_accessories + owner_accessories:
        resolved_parent_id = acc.parent_asset_id or resolve_parent_asset_id(
            internal_no=acc.sub_internal_no,
            group_no=acc.sub_group_no
        )
        if resolved_parent_id:
            if resolved_parent_id not in asset_ids:
                asset_ids.append(resolved_parent_id)
        else:
            if acc.id not in standalone_accessory_ids:
                standalone_accessories.append(acc)
                standalone_accessory_ids.add(acc.id)

    if asset_ids:
        matched_assets = Asset.query.filter(Asset.id.in_(asset_ids)).order_by(Asset.internal_no.asc()).all()

        for asset in matched_assets:
            accessories = get_asset_related_accessories(asset)

            rows.append({
                "row_type": "asset",
                "id": asset.id,
                "type_text": "主设备",
                "internal_no": asset.internal_no or "",
                "group_no": asset.group_no or "",
                "name": asset.name,
                "model": asset.model or "",
                "status": asset.status or "",
                "owner": asset.owner or "",
                "parent_asset_id": "",
                "accessory_count": len(accessories),
                "detail_url": url_for("asset_detail", asset_id=asset.id)
            })

            for item in accessories:
                rows.append({
                    "row_type": "accessory",
                    "id": item.id,
                    "type_text": "配件",
                    "internal_no": item.sub_internal_no or "",
                    "group_no": item.sub_group_no or "",
                    "name": item.name,
                    "model": item.model or "",
                    "status": item.status or "",
                    "owner": item.owner or "",
                    "parent_asset_id": item.parent_asset_id or resolve_parent_asset_id(item.sub_internal_no, item.sub_group_no) or "",
                    "accessory_count": 0,
                    "detail_url": url_for("accessory_detail", accessory_id=item.id)
                })

    existing_row_keys = {(row["row_type"], row["id"]) for row in rows}
    for item in standalone_accessories:
        if ("accessory", item.id) not in existing_row_keys:
            rows.append({
                "row_type": "accessory",
                "id": item.id,
                "type_text": "配件",
                "internal_no": item.sub_internal_no or "",
                "group_no": item.sub_group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": item.parent_asset_id or "",
                "accessory_count": 0,
                "detail_url": url_for("accessory_detail", accessory_id=item.id)
            })

    if not rows:
        standalone_exact = Accessory.query.filter(
            or_(
                Accessory.sub_internal_no == keyword,
                Accessory.sub_group_no == keyword,
                Accessory.sub_internal_no.like(f"{keyword}-%"),
                Accessory.sub_group_no.like(f"{keyword}-%"),
                Accessory.owner.like(f"%{keyword}%")
            )
        ).order_by(Accessory.sub_internal_no.asc()).all()

        for item in standalone_exact:
            rows.append({
                "row_type": "accessory",
                "id": item.id,
                "type_text": "配件",
                "internal_no": item.sub_internal_no or "",
                "group_no": item.sub_group_no or "",
                "name": item.name,
                "model": item.model or "",
                "status": item.status or "",
                "owner": item.owner or "",
                "parent_asset_id": item.parent_asset_id or "",
                "accessory_count": 0,
                "detail_url": url_for("accessory_detail", accessory_id=item.id)
            })

    return rows


def import_devices_from_excel(file_storage):
    wb = load_workbook(file_storage)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        group_no = normalize_text(row[0] if len(row) > 0 else "")
        internal_no = normalize_text(row[1] if len(row) > 1 else "")
        name = normalize_text(row[2] if len(row) > 2 else "")
        model = normalize_text(row[3] if len(row) > 3 else "")
        owner = normalize_text(row[4] if len(row) > 4 else "")
        location = normalize_text(row[5] if len(row) > 5 else "")
        status = normalize_text(row[6] if len(row) > 6 else "")
        remark = normalize_text(row[7] if len(row) > 7 else "")
        device_type = normalize_text(row[8] if len(row) > 8 else "")

        if not internal_no and not group_no:
            continue

        if device_type == "主设备":
            obj = find_existing_asset(internal_no=internal_no, group_no=group_no)

            if obj:
                obj.group_no = normalize_empty_to_none(group_no)
                obj.internal_no = normalize_empty_to_none(internal_no)
                obj.name = name
                obj.model = model
                obj.owner = owner
                obj.location = location
                obj.status = status
                obj.remark = remark
            else:
                obj = Asset(
                    group_no=normalize_empty_to_none(group_no),
                    internal_no=normalize_empty_to_none(internal_no),
                    name=name or internal_no or group_no,
                    model=model,
                    owner=owner,
                    location=location,
                    status=status,
                    remark=remark,
                    asset_date=date.today()
                )
                db.session.add(obj)

        elif device_type == "配件":
            parent_asset_id = resolve_parent_asset_id(internal_no=internal_no, group_no=group_no)

            obj = find_existing_accessory(internal_no=internal_no, group_no=group_no)

            if obj:
                obj.parent_asset_id = parent_asset_id
                obj.sub_group_no = normalize_empty_to_none(group_no)
                obj.sub_internal_no = normalize_empty_to_none(internal_no)
                obj.name = name
                obj.model = model
                obj.owner = owner
                obj.location = location
                obj.status = status
                obj.remark = remark
            else:
                obj = Accessory(
                    parent_asset_id=parent_asset_id,
                    sub_group_no=normalize_empty_to_none(group_no),
                    sub_internal_no=normalize_empty_to_none(internal_no),
                    name=name or internal_no or group_no,
                    model=model,
                    owner=owner,
                    location=location,
                    status=status,
                    remark=remark,
                    asset_date=date.today()
                )
                db.session.add(obj)


def is_group_no_value(value):
    value = normalize_text(value)
    return bool(re.fullmatch(r"\d{18}", value))


def get_recognized_no_label(value):
    return "集团编号" if is_group_no_value(value) else "内部编号"


def resolve_scan_target(recognized_no):
    recognized_no = normalize_text(recognized_no)
    if not recognized_no:
        return None, []

    rows = build_search_rows(keyword=recognized_no, searched=True)
    if not rows:
        return None, []

    asset_row = next(
        (
            row for row in rows
            if row.get("row_type") == "asset"
            and (
                normalize_text(row.get("group_no")) == recognized_no
                or normalize_text(row.get("internal_no")) == recognized_no
            )
        ),
        None,
    )
    if asset_row:
        return asset_row, rows

    accessory_row = next(
        (
            row for row in rows
            if row.get("row_type") == "accessory"
            and (
                normalize_text(row.get("group_no")) == recognized_no
                or normalize_text(row.get("internal_no")) == recognized_no
            )
        ),
        None,
    )
    if accessory_row:
        return accessory_row, rows

    accessory_prefix_row = next(
        (
            row for row in rows
            if row.get("row_type") == "accessory"
            and (
                normalize_text(row.get("group_no")).startswith(f"{recognized_no}-")
                or normalize_text(row.get("internal_no")).startswith(f"{recognized_no}-")
            )
        ),
        None,
    )
    if accessory_prefix_row:
        return accessory_prefix_row, rows

    return rows[0], rows


def register_routes(app):
    @app.route("/uploads/<path:filename>")
    @login_required
    def uploaded_file(filename):
        safe_path = os.path.normpath(filename)
        if safe_path.startswith(".."):
            abort(404)
        return send_from_directory(Config.UPLOAD_FOLDER, safe_path)

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = normalize_text(request.form.get("username"))
            password = normalize_text(request.form.get("password"))

            user = User.query.filter_by(username=username).first()

            if user and check_password_hash(user.password_hash, password):
                login_user(user)
                return redirect(url_for("search_assets"))

            return render_template_string(LOGIN_HTML, error="用户名或密码错误")

        return render_template_string(LOGIN_HTML, error="")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    @app.route("/", methods=["GET"])
    @login_required
    def search_assets():
        keyword = normalize_text(request.args.get("keyword"))
        status_filter = normalize_text(request.args.get("status_filter"))
        device_type_filter = normalize_text(request.args.get("device_type_filter"))
        searched = normalize_text(request.args.get("searched")) == "1"
        per_page = normalize_text(request.args.get("per_page")) or "30"
        page = normalize_text(request.args.get("page")) or "1"
        statuses = get_statuses()

        try:
            per_page = int(per_page)
        except:
            per_page = 30

        if per_page not in [30, 50, 100]:
            per_page = 30

        try:
            page = int(page)
        except:
            page = 1

        all_rows = build_search_rows(keyword=keyword, searched=searched)

        if searched and status_filter:
            all_rows = [
                row for row in all_rows
                if normalize_text(row.get("status")) == status_filter
            ]

        if searched and device_type_filter:
            all_rows = [
                row for row in all_rows
                if normalize_text(row.get("type_text")) == device_type_filter
            ]

        total = len(all_rows)

        start = (page - 1) * per_page
        end = start + per_page
        rows = all_rows[start:end]

        total_pages = (total + per_page - 1) // per_page if total else 1
        error = ""
        if searched and total == 0:
            if keyword and (status_filter or device_type_filter):
                error = "未找到符合筛选条件的资产"
            elif keyword:
                error = "未找到对应资产"
            elif status_filter or device_type_filter:
                error = "未找到符合筛选条件的资产"

        return render_template_string(
            SEARCH_HTML,
            current_user=current_user,
            keyword=keyword,
            status_filter=status_filter,
            device_type_filter=device_type_filter,
            statuses=statuses,
            searched=searched,
            per_page=per_page,
            page=page,
            total=total,
            total_pages=total_pages,
            rows=rows,
            error=error
        )


    @app.route("/scan_label", methods=["GET", "POST"])
    @login_required
    def scan_label():
        error = ""
        recognized_no = ""
        recognized_label = "资产编号"
        debug_texts = []
        target_row = None
        matched_rows = []

        if request.method == "POST":
            scan_file = request.files.get("scan_image")
            if not scan_file or not scan_file.filename:
                error = "请先拍照或上传标签图片"
            else:
                try:
                    with scan_time_limit(SCAN_TIMEOUT_SECONDS):
                        recognized_no, debug_texts = extract_group_no_from_label(scan_file)
                    if not recognized_no:
                        error = "未识别到有效资产编号，请尽量只拍一张标签并让编号区域更清晰"
                    else:
                        recognized_label = get_recognized_no_label(recognized_no)
                        target_row, matched_rows = resolve_scan_target(recognized_no)
                except ScanTimeoutError:
                    error = f"识别超时，已自动终止（超过{SCAN_TIMEOUT_SECONDS}秒）"
                except Exception as e:
                    error = f"识别失败：{str(e)}"

        rendered_html = render_template_string(
            SCAN_LABEL_HTML,
            current_user=current_user,
            recognized_no=recognized_no,
            recognized_label=recognized_label,
            debug_texts=debug_texts,
            target_row=target_row,
            matched_rows=matched_rows,
            error=error,
            timeout_seconds=SCAN_TIMEOUT_SECONDS
        )
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"html": rendered_html})
        return rendered_html

    @app.route("/export", methods=["POST"])
    @login_required
    def export_selected():
        selected_items = request.form.getlist("selected_items")

        wb = Workbook()
        ws = wb.active
        ws.title = "设备导出"

        ws.append([
            "类型", "内部编号", "集团编号", "名称", "型号", "责任人",
            "位置", "时间", "状态", "备注"
        ])

        for item in selected_items:
            try:
                row_type, row_id = item.split(":")
                row_id = int(row_id)
            except:
                continue

            if row_type == "asset":
                obj = Asset.query.get(row_id)
                if obj:
                    ws.append([
                        "主设备", obj.internal_no, obj.group_no, obj.name, obj.model or "",
                        obj.owner or "", obj.location or "", obj.asset_date.isoformat() if obj.asset_date else "",
                        obj.status or "", obj.remark or ""
                    ])
            elif row_type == "accessory":
                obj = Accessory.query.get(row_id)
                if obj:
                    ws.append([
                        "配件", obj.sub_internal_no, obj.sub_group_no, obj.name, obj.model or "",
                        obj.owner or "", obj.location or "", obj.asset_date.isoformat() if obj.asset_date else "",
                        obj.status or "", obj.remark or ""
                    ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"asset_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/delete_selected", methods=["POST"])
    @login_required
    def delete_selected():
        selected_items = request.form.getlist("selected_items")
        delete_pin = normalize_text(request.form.get("delete_pin"))
        keyword = normalize_text(request.form.get("keyword"))
        status_filter = normalize_text(request.form.get("status_filter"))
        device_type_filter = normalize_text(request.form.get("device_type_filter"))
        per_page = normalize_text(request.form.get("per_page")) or "30"

        if delete_pin != "0819":
            return "批量删除失败：Pin码错误"

        asset_ids = set()
        accessory_ids = set()

        for item in selected_items:
            try:
                row_type, row_id = item.split(":")
                row_id = int(row_id)
            except:
                continue

            if row_type == "asset":
                asset_ids.add(row_id)
            elif row_type == "accessory":
                accessory_ids.add(row_id)

        try:
            for asset_id in asset_ids:
                asset = Asset.query.get(asset_id)
                if asset:
                    delete_asset_with_files(asset)

            for accessory_id in accessory_ids:
                accessory = Accessory.query.get(accessory_id)
                resolved_parent_id = None
                if accessory:
                    resolved_parent_id = accessory.parent_asset_id or resolve_parent_asset_id(
                        internal_no=accessory.sub_internal_no,
                        group_no=accessory.sub_group_no
                    )
                if accessory and resolved_parent_id not in asset_ids:
                    delete_accessory_with_files(accessory)

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量删除失败：{str(e)}"

        return redirect(url_for(
            "search_assets",
            searched=1,
            keyword=keyword,
            status_filter=status_filter,
            device_type_filter=device_type_filter,
            per_page=per_page,
            page=1
        ))

    @app.route("/import_devices", methods=["POST"])
    @login_required
    def import_devices():
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            return redirect(url_for("search_assets"))

        try:
            import_devices_from_excel(excel_file)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量上传失败：{str(e)}"

        return redirect(url_for("search_assets"))

    @app.route("/device/new", methods=["GET", "POST"])
    @login_required
    def device_new():
        statuses = get_statuses()
        error = ""

        device_type = normalize_text(request.args.get("type")) or "主设备"
        parent_asset_id = normalize_text(request.args.get("parent_asset_id"))
        recognized_no = normalize_text(request.args.get("recognized_no")) or normalize_text(request.args.get("recognized_group_no"))
        parent_asset = Asset.query.get(int(parent_asset_id)) if parent_asset_id.isdigit() else None

        default_group_no = ""
        default_internal_no = ""
        if device_type == "主设备" and recognized_no:
            if is_group_no_value(recognized_no):
                default_group_no = recognized_no
            else:
                default_internal_no = recognized_no
        if device_type == "配件" and parent_asset:
            default_group_no = make_accessory_prefix(parent_asset.group_no)
            default_internal_no = make_accessory_prefix(parent_asset.internal_no)

        form_data = {
            "device_type": device_type,
            "group_no": default_group_no,
            "internal_no": default_internal_no,
            "name": "",
            "model": "",
            "owner": "",
            "location": "",
            "asset_date": date.today().isoformat(),
            "status": "",
            "remark": "",
            "parent_asset_id": parent_asset_id
        }

        if request.method == "POST":
            device_type = normalize_text(request.form.get("device_type")) or "主设备"
            group_no = normalize_text(request.form.get("group_no"))
            internal_no = normalize_text(request.form.get("internal_no"))
            name = normalize_text(request.form.get("name"))
            model = normalize_text(request.form.get("model"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_text(request.form.get("location"))
            asset_date_str = normalize_text(request.form.get("asset_date"))
            status = normalize_text(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            parent_asset_id = normalize_text(request.form.get("parent_asset_id"))
            image_files = request.files.getlist("image_files")

            form_data = {
                "device_type": device_type,
                "group_no": group_no,
                "internal_no": internal_no,
                "name": name,
                "model": model,
                "owner": owner,
                "location": location,
                "asset_date": asset_date_str or date.today().isoformat(),
                "status": status,
                "remark": remark,
                "parent_asset_id": parent_asset_id
            }

            if device_type == "主设备":
                number_error = validate_required_number_pair(internal_no, group_no, "内部编号", "集团编号")
                if not number_error:
                    number_error = validate_asset_group_no(group_no, "集团编号")
            else:
                number_error = validate_required_number_pair(internal_no, group_no, "附属资产内部编号", "附属资产集团编号")
                if not number_error:
                    number_error = validate_accessory_no_format(internal_no, group_no)
                if not number_error:
                    number_error = validate_accessory_group_no(group_no, "附属资产集团编号")

            if number_error:
                error = number_error
            elif not name:
                error = "名称不能为空"
            else:
                try:
                    if device_type == "主设备":
                        existing_group = Asset.query.filter_by(group_no=group_no).first() if group_no else None
                        existing_internal = Asset.query.filter_by(internal_no=internal_no).first() if internal_no else None

                        if existing_group:
                            error = "集团编号已存在"
                        elif existing_internal:
                            error = "内部编号已存在"
                        else:
                            obj = Asset(
                                group_no=normalize_empty_to_none(group_no),
                                internal_no=normalize_empty_to_none(internal_no),
                                name=name,
                                model=model,
                                owner=owner,
                                location=location,
                                asset_date=datetime.strptime(asset_date_str, "%Y-%m-%d").date() if asset_date_str else date.today(),
                                status=status,
                                remark=remark
                            )
                            db.session.add(obj)
                            db.session.flush()

                            for file_storage in image_files:
                                rel = save_uploaded_image(file_storage, "assets")
                                if rel:
                                    db.session.add(AssetImage(asset_id=obj.id, image_path=rel))

                            trim_asset_images(obj)
                            db.session.commit()
                            return redirect(url_for("asset_detail", asset_id=obj.id))

                    else:
                        existing_group = Accessory.query.filter_by(sub_group_no=group_no).first() if group_no else None
                        existing_internal = Accessory.query.filter_by(sub_internal_no=internal_no).first() if internal_no else None

                        if existing_group:
                            error = "附属资产集团编号已存在"
                        elif existing_internal:
                            error = "附属资产内部编号已存在"
                        else:
                            parent_id = resolve_parent_asset_id(
                                internal_no=internal_no,
                                group_no=group_no,
                                fallback_parent_asset_id=parent_asset_id
                            )
                            obj = Accessory(
                                parent_asset_id=parent_id,
                                sub_group_no=normalize_empty_to_none(group_no),
                                sub_internal_no=normalize_empty_to_none(internal_no),
                                name=name,
                                model=model,
                                owner=owner,
                                location=location,
                                asset_date=datetime.strptime(asset_date_str, "%Y-%m-%d").date() if asset_date_str else date.today(),
                                status=status,
                                remark=remark
                            )
                            db.session.add(obj)
                            db.session.flush()

                            for file_storage in image_files:
                                rel = save_uploaded_image(file_storage, "accessories")
                                if rel:
                                    db.session.add(AccessoryImage(accessory_id=obj.id, image_path=rel))

                            trim_accessory_images(obj)
                            db.session.commit()
                            return redirect(url_for("accessory_detail", accessory_id=obj.id))

                except Exception as e:
                    db.session.rollback()
                    error = f"保存失败：{str(e)}"

        return render_template_string(
            DEVICE_NEW_HTML,
            current_user=current_user,
            statuses=statuses,
            form_data=form_data,
            error=error
        )

    @app.route("/asset/<int:asset_id>", methods=["GET", "POST"])
    @login_required
    def asset_detail(asset_id):
        asset = Asset.query.get_or_404(asset_id)
        statuses = get_statuses()
        message = ""
        error = ""

        if request.method == "POST":
            action = normalize_text(request.form.get("action"))

            if action == "save_asset":
                group_no = normalize_text(request.form.get("group_no"))
                internal_no = normalize_text(request.form.get("internal_no"))
                name = normalize_text(request.form.get("name"))
                model = normalize_text(request.form.get("model"))
                owner = normalize_text(request.form.get("owner"))
                location = normalize_text(request.form.get("location"))
                asset_date_str = normalize_text(request.form.get("asset_date"))
                status = normalize_text(request.form.get("status"))
                remark = normalize_text(request.form.get("remark"))
                image_files = request.files.getlist("image_files")
                delete_image_ids = request.form.getlist("delete_asset_image_ids")

                number_error = validate_required_number_pair(internal_no, group_no, "内部编号", "集团编号")
                if not number_error:
                    number_error = validate_asset_group_no(group_no, "集团编号")

                if number_error:
                    error = number_error
                elif not name:
                    error = "名称不能为空"
                else:
                    try:
                        existing_group = Asset.query.filter(Asset.group_no == group_no, Asset.id != asset.id).first() if group_no else None
                        existing_internal = Asset.query.filter(Asset.internal_no == internal_no, Asset.id != asset.id).first() if internal_no else None

                        if existing_group:
                            error = "集团编号已存在"
                        elif existing_internal:
                            error = "内部编号已存在"
                        else:
                            asset.group_no = normalize_empty_to_none(group_no)
                            asset.internal_no = normalize_empty_to_none(internal_no)
                            asset.name = name
                            asset.model = model
                            asset.owner = owner
                            asset.location = location
                            asset.asset_date = datetime.strptime(asset_date_str, "%Y-%m-%d").date() if asset_date_str else date.today()
                            asset.status = status
                            asset.remark = remark

                            for image_id in delete_image_ids:
                                try:
                                    img_id = int(image_id)
                                except:
                                    continue

                                img = AssetImage.query.filter_by(id=img_id, asset_id=asset.id).first()
                                if img:
                                    delete_image_file(img.image_path)
                                    db.session.delete(img)

                            for file_storage in image_files:
                                rel = save_uploaded_image(file_storage, "assets")
                                if rel:
                                    db.session.add(AssetImage(asset_id=asset.id, image_path=rel))

                            trim_asset_images(asset)
                            db.session.commit()
                            message = "主设备更新成功"
                    except Exception as e:
                        db.session.rollback()
                        error = f"更新失败：{str(e)}"

        accessories = get_asset_related_accessories(asset)
        images = AssetImage.query.filter_by(asset_id=asset.id).order_by(AssetImage.created_at.asc(), AssetImage.id.asc()).all()

        return render_template_string(
            ASSET_DETAIL_HTML,
            current_user=current_user,
            asset=asset,
            accessories=accessories,
            images=images,
            statuses=statuses,
            today=date.today().isoformat(),
            message=message,
            error=error
        )

    @app.route("/asset/<int:asset_id>/delete", methods=["POST"])
    @login_required
    def delete_asset(asset_id):
        asset = Asset.query.get_or_404(asset_id)
        try:
            delete_asset_with_files(asset)
            db.session.commit()
            return redirect(url_for("search_assets"))
        except Exception as e:
            db.session.rollback()
            return f"删除失败：{str(e)}"

    @app.route("/accessory/<int:accessory_id>", methods=["GET", "POST"])
    @login_required
    def accessory_detail(accessory_id):
        accessory = Accessory.query.get_or_404(accessory_id)
        statuses = get_statuses()
        message = ""
        error = ""

        if request.method == "POST":
            sub_group_no = normalize_text(request.form.get("sub_group_no"))
            sub_internal_no = normalize_text(request.form.get("sub_internal_no"))
            name = normalize_text(request.form.get("name"))
            model = normalize_text(request.form.get("model"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_text(request.form.get("location"))
            asset_date_str = normalize_text(request.form.get("asset_date"))
            status = normalize_text(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")
            delete_image_ids = request.form.getlist("delete_accessory_image_ids")

            number_error = validate_required_number_pair(sub_internal_no, sub_group_no, "附属资产内部编号", "附属资产集团编号")
            if not number_error:
                number_error = validate_accessory_no_format(sub_internal_no, sub_group_no)
            if not number_error:
                number_error = validate_accessory_group_no(sub_group_no, "附属资产集团编号")

            if number_error:
                error = number_error
            elif not name:
                error = "名称不能为空"
            else:
                try:
                    existing_group = Accessory.query.filter(Accessory.sub_group_no == sub_group_no, Accessory.id != accessory.id).first() if sub_group_no else None
                    existing_internal = Accessory.query.filter(Accessory.sub_internal_no == sub_internal_no, Accessory.id != accessory.id).first() if sub_internal_no else None

                    if existing_group:
                        error = "附属资产集团编号已存在"
                    elif existing_internal:
                        error = "附属资产内部编号已存在"
                    else:
                        accessory.parent_asset_id = resolve_parent_asset_id(
                            internal_no=sub_internal_no,
                            group_no=sub_group_no,
                            fallback_parent_asset_id=accessory.parent_asset_id
                        )
                        accessory.sub_group_no = normalize_empty_to_none(sub_group_no)
                        accessory.sub_internal_no = normalize_empty_to_none(sub_internal_no)
                        accessory.name = name
                        accessory.model = model
                        accessory.owner = owner
                        accessory.location = location
                        accessory.asset_date = datetime.strptime(asset_date_str, "%Y-%m-%d").date() if asset_date_str else date.today()
                        accessory.status = status
                        accessory.remark = remark

                        for image_id in delete_image_ids:
                            try:
                                img_id = int(image_id)
                            except:
                                continue

                            img = AccessoryImage.query.filter_by(id=img_id, accessory_id=accessory.id).first()
                            if img:
                                delete_image_file(img.image_path)
                                db.session.delete(img)

                        for file_storage in image_files:
                            rel = save_uploaded_image(file_storage, "accessories")
                            if rel:
                                db.session.add(AccessoryImage(accessory_id=accessory.id, image_path=rel))

                        trim_accessory_images(accessory)
                        db.session.commit()
                        message = "配件更新成功"
                except Exception as e:
                    db.session.rollback()
                    error = f"更新失败：{str(e)}"

        images = AccessoryImage.query.filter_by(accessory_id=accessory.id).order_by(AccessoryImage.created_at.asc(), AccessoryImage.id.asc()).all()
        resolved_parent_asset_id = accessory.parent_asset_id or resolve_parent_asset_id(
            internal_no=accessory.sub_internal_no,
            group_no=accessory.sub_group_no
        )
        back_url = url_for("asset_detail", asset_id=resolved_parent_asset_id) if resolved_parent_asset_id else url_for("search_assets")

        return render_template_string(
            ACCESSORY_DETAIL_HTML,
            current_user=current_user,
            accessory=accessory,
            statuses=statuses,
            message=message,
            error=error,
            images=images,
            back_url=back_url
        )

    @app.route("/accessory/<int:accessory_id>/delete", methods=["POST"])
    @login_required
    def delete_accessory(accessory_id):
        accessory = Accessory.query.get_or_404(accessory_id)
        resolved_parent_asset_id = accessory.parent_asset_id or resolve_parent_asset_id(
            internal_no=accessory.sub_internal_no,
            group_no=accessory.sub_group_no
        )
        back_url = url_for("asset_detail", asset_id=resolved_parent_asset_id) if resolved_parent_asset_id else url_for("search_assets")
        try:
            delete_accessory_with_files(accessory)
            db.session.commit()
            return redirect(back_url)
        except Exception as e:
            db.session.rollback()
            return f"删除失败：{str(e)}"


LOGIN_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>设备管理系统</title>
<style>
body{margin:0;font-family:Arial,sans-serif;background:#0b1730;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:16px;}
.card{width:100%;max-width:420px;background:#16213c;border-radius:24px;padding:28px 22px;color:#fff;box-sizing:border-box;}
h1{text-align:center;margin:0 0 24px 0;font-size:32px;}
input{width:100%;box-sizing:border-box;padding:16px;margin-bottom:14px;border:none;border-radius:16px;background:#243454;color:#fff;font-size:18px;}
button{width:100%;padding:16px;border:none;border-radius:18px;background:#00a88f;color:white;font-size:22px;}
.err{color:#ff8080;text-align:center;margin-bottom:12px;}
</style>
</head>
<body>
<div class="card">
    <h1>设备管理系统</h1>
    {% if error %}<div class="err">{{ error }}</div>{% endif %}
    <form method="post">
        <input type="text" name="username" placeholder="用户名">
        <input type="password" name="password" placeholder="密码">
        <button type="submit">登录</button>
    </form>
</div>
</body>
</html>
"""

SEARCH_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>资产查询</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:1200px;margin:auto;padding:12px;}
.card{background:#fff;border-radius:10px;padding:14px;margin-bottom:14px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.topbar{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;}
input,select,button{padding:10px;font-size:16px;border-radius:6px;border:1px solid #ccc;box-sizing:border-box;}
input[type=text]{min-width:260px;}
button{background:#0d6efd;color:#fff;border:none;cursor:pointer;}
.btn-green{background:#198754;}
.btn-gray{background:#6c757d;}
.btn-orange{background:#fd7e14;}
.btn-red{background:#dc3545;}
.table-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;min-width:900px;}
th,td{border:1px solid #ccc;padding:8px;text-align:left;}
th{background:#f0f3f8;}
.err{color:red;margin-top:10px;}
.num-link{color:#0d6efd;text-decoration:underline;}
.action-bar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;}
.pagination{margin-top:12px;display:flex;gap:10px;flex-wrap:wrap;}
.icon-only-btn{width:42px;min-width:42px;height:42px;padding:0;display:inline-flex;align-items:center;justify-content:center;}
.icon-only-btn svg{width:22px;height:22px;stroke:currentColor;fill:none;stroke-width:2.2;stroke-linecap:round;stroke-linejoin:round;}
@media (max-width:768px){
  input[type=text]{width:100%;}
}
</style>
<script>
function toggleSelectCurrentPage(source){
    const items = document.querySelectorAll('input[name="selected_items"]');
    items.forEach(item => {
        item.checked = source.checked;
    });
}

function confirmDeleteSelected(){
    const checkedItems = Array.from(document.querySelectorAll('input[name="selected_items"]:checked'));
    if(checkedItems.length === 0){
        alert('请先勾选要删除的资产');
        return false;
    }

    const selectedAssetIds = new Set();
    let assetCount = 0;
    let accessoryCount = 0;

    checkedItems.forEach(item => {
        if(item.dataset.rowType === 'asset'){
            assetCount += 1;
            selectedAssetIds.add(item.dataset.rowId);
            accessoryCount += parseInt(item.dataset.accessoryCount || '0', 10);
        }
    });

    checkedItems.forEach(item => {
        if(item.dataset.rowType === 'accessory'){
            const parentAssetId = item.dataset.parentAssetId || '';
            if(!parentAssetId || !selectedAssetIds.has(parentAssetId)){
                accessoryCount += 1;
            }
        }
    });

    if(!confirm(`将删除 ${assetCount} 条主设备、${accessoryCount} 条配件，是否继续？`)){
        return false;
    }

    const pin = prompt('请输入4位Pin码确认批量删除');
    if(pin === null){
        return false;
    }
    if(pin !== '0819'){
        alert('Pin码错误，已取消批量删除');
        return false;
    }

    document.getElementById('delete_pin').value = pin;
    return true;
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="topbar">
            <div><strong>当前用户：</strong>{{ current_user.username }}</div>
            <div><a href="/logout">退出登录</a></div>
        </div>
    </div>

    <div class="card">
        <h2>资产查询</h2>
        <form method="get" action="/">
            <input type="hidden" name="searched" value="1">
            <div class="action-bar">
                <input type="text" name="keyword" value="{{ keyword }}" placeholder="编号/后6位、责任人">
                <select name="status_filter">
                    <option value="">全部状态</option>
                    {% for s in statuses %}
                    <option value="{{ s.dict_value }}" {% if status_filter == s.dict_value %}selected{% endif %}>{{ s.dict_value }}</option>
                    {% endfor %}
                </select>
                <select name="device_type_filter">
                    <option value="">全部类型</option>
                    <option value="主设备" {% if device_type_filter == '主设备' %}selected{% endif %}>主设备</option>
                    <option value="配件" {% if device_type_filter == '配件' %}selected{% endif %}>配件</option>
                </select>
                <select name="per_page">
                    <option value="30" {% if per_page == 30 %}selected{% endif %}>每页30条</option>
                    <option value="50" {% if per_page == 50 %}selected{% endif %}>每页50条</option>
                    <option value="100" {% if per_page == 100 %}selected{% endif %}>每页100条</option>
                </select>
                <button type="submit">搜索</button>
                <a href="/"><button type="button" class="btn-gray">重置</button></a>
                <a href="/scan_label" title="扫一扫"><button type="button" class="btn-gray icon-only-btn" aria-label="扫一扫">
                    <svg viewBox="0 0 24 24" aria-hidden="true">
                        <path d="M8 4H6a2 2 0 0 0-2 2v2"></path>
                        <path d="M16 4h2a2 2 0 0 1 2 2v2"></path>
                        <path d="M8 20H6a2 2 0 0 1-2-2v-2"></path>
                        <path d="M16 20h2a2 2 0 0 0 2-2v-2"></path>
                        <path d="M5 12h14"></path>
                    </svg>
                </button></a>
                <a href="/device/new"><button type="button" class="btn-green">+</button></a>
            </div>
        </form>

        <form method="post" action="/import_devices" enctype="multipart/form-data" style="margin-top:12px;">
            <div class="action-bar">
                <input type="file" name="excel_file" accept=".xlsx,.xlsm,.xltx,.xltm">
                <button type="submit" class="btn-orange">批量上传设备</button>
            </div>
        </form>

        {% if error %}<div class="err">{{ error }}</div>{% endif %}
    </div>

    {% if searched %}
    <div class="card">
        <form method="post" action="/export">
            <input type="hidden" name="keyword" value="{{ keyword }}">
            <input type="hidden" name="status_filter" value="{{ status_filter }}">
            <input type="hidden" name="device_type_filter" value="{{ device_type_filter }}">
            <input type="hidden" name="per_page" value="{{ per_page }}">
            <input type="hidden" name="delete_pin" id="delete_pin" value="">

            <div class="action-bar" style="margin-bottom:10px;">
                <button type="submit" class="btn-green">导出选中</button>
                <button type="submit" class="btn-red" formaction="/delete_selected" formmethod="post" onclick="return confirmDeleteSelected();">删除选中</button>
                <div>当前总数：{{ total }}</div>
            </div>

            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th><input type="checkbox" onclick="toggleSelectCurrentPage(this)" title="选取当前页"></th>
                            <th>集团编号</th>
                            <th>内部编号</th>
                            <th>资产名称</th>
                            <th>型号</th>
                            <th>状态</th>
                            <th>责任人</th>
                            <th>类型</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for item in rows %}
                        <tr>
                            <td><input type="checkbox" name="selected_items" value="{{ item.row_type }}:{{ item.id }}" data-row-type="{{ item.row_type }}" data-row-id="{{ item.id }}" data-parent-asset-id="{{ item.parent_asset_id }}" data-accessory-count="{{ item.accessory_count }}"></td>
                            <td><a class="num-link" href="{{ item.detail_url }}">{{ item.group_no }}</a></td>
                            <td><a class="num-link" href="{{ item.detail_url }}">{{ item.internal_no }}</a></td>
                            <td>{{ item.name }}</td>
                            <td>{{ item.model }}</td>
                            <td>{{ item.status }}</td>
                            <td>{{ item.owner }}</td>
                            <td>{{ item.type_text }}</td>
                        </tr>
                        {% endfor %}
                        {% if not rows %}
                        <tr><td colspan="8">暂无数据</td></tr>
                        {% endif %}
                    </tbody>
                </table>
            </div>
        </form>

        <div class="pagination">
            {% if page > 1 %}
                <a href="/?searched=1&keyword={{ keyword }}&status_filter={{ status_filter }}&device_type_filter={{ device_type_filter }}&per_page={{ per_page }}&page={{ page - 1 }}">上一页</a>
            {% endif %}
            <span>第 {{ page }} / {{ total_pages }} 页</span>
            {% if page < total_pages %}
                <a href="/?searched=1&keyword={{ keyword }}&status_filter={{ status_filter }}&device_type_filter={{ device_type_filter }}&per_page={{ per_page }}&page={{ page + 1 }}">下一页</a>
            {% endif %}
        </div>
    </div>
    {% endif %}
</div>
</body>
</html>
"""

SCAN_LABEL_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>扫一扫识别资产编号</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:900px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,button,textarea{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
button{background:#0d6efd;color:#fff;border:none;}
.btn-gray{background:#6c757d;}
.btn-green{background:#198754;}
.err{color:red;margin-bottom:10px;}
.msg{color:green;margin-bottom:10px;}
.result-box{padding:12px;border:1px solid #d0d7e2;border-radius:10px;background:#f8fafc;font-size:24px;font-weight:bold;letter-spacing:1px;word-break:break-all;}
.muted{color:#666;font-size:14px;}
.action-row{display:flex;gap:10px;flex-wrap:wrap;}
.action-row a,.action-row button{width:auto;}
.debug-box{white-space:pre-wrap;word-break:break-all;background:#f8fafc;border:1px solid #d0d7e2;border-radius:10px;padding:12px;font-size:13px;color:#444;max-height:240px;overflow:auto;}
.scan-icon-wrap{display:flex;align-items:center;justify-content:center;margin-bottom:12px;}
.scan-icon{width:64px;height:64px;border-radius:50%;background:#f0f2f5;display:flex;align-items:center;justify-content:center;color:#111827;}
.scan-icon svg{width:34px;height:34px;stroke:currentColor;fill:none;stroke-width:2.2;stroke-linecap:round;stroke-linejoin:round;}
.live-status{display:none;margin-bottom:12px;padding:10px 12px;border-radius:10px;font-size:14px;line-height:1.5;}
.live-status.show{display:block;}
.live-status.error{background:#fff1f0;border:1px solid #f5c2c0;color:#c62828;}
.live-status.info{background:#eef6ff;border:1px solid #c9ddff;color:#114a8b;}
.progress-wrap{display:none;margin-top:12px;}
.progress-wrap.show{display:block;}
.progress-bar-bg{width:100%;height:10px;background:#e9ecef;border-radius:999px;overflow:hidden;}
.progress-bar-fill{width:0;height:100%;background:#0d6efd;transition:width .2s ease;}
.progress-text{margin-top:8px;color:#555;font-size:14px;}
.progress-time{margin-top:4px;color:#777;font-size:13px;}
</style>
<script>
let scanController = null;
let scanProgressTimer = null;
let scanTimeoutTimer = null;
let scanStartedAt = 0;
let isScanning = false;
const SCAN_TIMEOUT_SECONDS = {{ timeout_seconds|default(15) }};

function setScanStatus(message, variant){
    const box = document.getElementById('scan-live-status');
    if(!box){
        return;
    }
    if(!message){
        box.textContent = '';
        box.className = 'live-status';
        return;
    }
    box.textContent = message;
    box.className = 'live-status show ' + (variant || 'info');
}

function updateSelectedImageTip(){
    const fileInput = document.getElementById('scan-image-input');
    const tip = document.getElementById('scan-file-tip');
    if(!tip || !fileInput){
        return;
    }
    if(fileInput.files && fileInput.files.length > 0){
        const fileName = fileInput.files[0].name || '已选择图片';
        tip.textContent = '图片已选择：' + fileName;
        tip.style.display = 'block';
    }else{
        tip.textContent = '';
        tip.style.display = 'none';
    }
}

function setScanButtons(scanning){
    const submitBtn = document.getElementById('scan-submit-btn');
    const backBtn = document.getElementById('scan-back-btn');
    const fileInput = document.getElementById('scan-image-input');
    if(submitBtn){
        submitBtn.disabled = scanning;
        submitBtn.textContent = scanning ? '识别中...' : (submitBtn.dataset.idleText || '识别');
    }
    if(backBtn){
        backBtn.textContent = scanning ? '终止识别' : '返回';
    }
    if(fileInput){
        fileInput.disabled = scanning;
    }
}

function updateScanProgress(){
    const wrap = document.getElementById('scan-progress-wrap');
    const fill = document.getElementById('scan-progress-fill');
    const text = document.getElementById('scan-progress-text');
    const time = document.getElementById('scan-progress-time');
    if(!wrap || !fill || !text || !time || !isScanning){
        return;
    }
    const elapsedMs = Date.now() - scanStartedAt;
    const elapsedSec = elapsedMs / 1000;
    const ratio = Math.min(elapsedSec / SCAN_TIMEOUT_SECONDS, 1);
    const percent = Math.min(95, Math.max(6, Math.round(ratio * 95)));
    fill.style.width = percent + '%';

    if(elapsedSec < 3){
        text.textContent = '正在上传并读取图片...';
    }else if(elapsedSec < 7){
        text.textContent = '正在分析标签区域...';
    }else if(elapsedSec < 11){
        text.textContent = '正在识别资产编号...';
    }else{
        text.textContent = '正在整理识别结果...';
    }
    time.textContent = '已用时 ' + elapsedSec.toFixed(1) + ' 秒 / 最长 ' + SCAN_TIMEOUT_SECONDS + ' 秒';
}

function startScanProgress(){
    const wrap = document.getElementById('scan-progress-wrap');
    const fill = document.getElementById('scan-progress-fill');
    if(wrap){
        wrap.classList.add('show');
    }
    if(fill){
        fill.style.width = '6%';
    }
    scanStartedAt = Date.now();
    updateScanProgress();
    scanProgressTimer = window.setInterval(updateScanProgress, 200);
}

function stopScanProgress(){
    const wrap = document.getElementById('scan-progress-wrap');
    const fill = document.getElementById('scan-progress-fill');
    if(scanProgressTimer){
        window.clearInterval(scanProgressTimer);
        scanProgressTimer = null;
    }
    if(scanTimeoutTimer){
        window.clearTimeout(scanTimeoutTimer);
        scanTimeoutTimer = null;
    }
    if(wrap){
        wrap.classList.remove('show');
    }
    if(fill){
        fill.style.width = '0%';
    }
}

function finishScanUi(){
    isScanning = false;
    scanController = null;
    stopScanProgress();
    setScanButtons(false);
}

function resetScanPage(){
    try {
        const form = document.getElementById('scan-form');
        const fileInput = document.getElementById('scan-image-input');
        const tip = document.getElementById('scan-file-tip');
        const liveStatus = document.getElementById('scan-live-status');
        if(form){
            form.reset();
        }
        if(fileInput){
            fileInput.value = '';
        }
        if(tip){
            tip.textContent = '';
            tip.style.display = 'none';
        }
        if(liveStatus){
            liveStatus.textContent = '';
            liveStatus.className = 'live-status';
        }
        finishScanUi();
    } catch (e) {}
    window.location.replace('/scan_label?_reset=' + Date.now());
}

function triggerSubmitScan(event){
    submitScanForm(event);
    return false;
}

async function submitScanForm(event){
    event.preventDefault();
    if(isScanning){
        return false;
    }

    const submitBtn = document.getElementById('scan-submit-btn');
    if(submitBtn && submitBtn.dataset.mode === 'reset'){
        resetScanPage();
        return false;
    }

    const form = document.getElementById('scan-form');
    const fileInput = document.getElementById('scan-image-input');
    if(!fileInput || !fileInput.files || fileInput.files.length === 0){
        setScanStatus('请先拍照或上传标签图片', 'error');
        return false;
    }

    const formData = new FormData(form);

    isScanning = true;
    setScanStatus('已开始识别，请稍候...', 'info');
    setScanButtons(true);
    startScanProgress();

    scanController = new AbortController();
    scanTimeoutTimer = window.setTimeout(() => {
        if(scanController){
            scanController.abort();
            setScanStatus('15秒内未识别成功，已自动终止本次识别。', 'error');
        }
    }, SCAN_TIMEOUT_SECONDS * 1000);

    try {
        const response = await fetch(form.action || window.location.pathname, {
            method: 'POST',
            body: formData,
            headers: {'X-Requested-With': 'XMLHttpRequest'},
            signal: scanController.signal,
            cache: 'no-store'
        });

        const contentType = response.headers.get('content-type') || '';
        if(contentType.indexOf('application/json') >= 0){
            const data = await response.json();
            if(data && data.html){
                document.open();
                document.write(data.html);
                document.close();
                return false;
            }
        }

        const html = await response.text();
        document.open();
        document.write(html);
        document.close();
        return false;
    } catch (error) {
        if(error && error.name === 'AbortError'){
            if(!document.getElementById('scan-live-status').textContent){
                setScanStatus('已终止本次识别。', 'error');
            }
        }else{
            setScanStatus('识别失败：' + ((error && error.message) ? error.message : '网络异常'), 'error');
        }
    } finally {
        finishScanUi();
    }
    return false;
}

function handleBackOrStop(){
    if(isScanning && scanController){
        scanController.abort();
        setScanStatus('已手动终止本次识别。', 'error');
        return;
    }
    try {
        const form = document.getElementById('scan-form');
        if(form){ form.reset(); }
    } catch (e) {}
    window.location.replace('/?_back=' + Date.now());
}

function initScanPage(){
    updateSelectedImageTip();
    setScanButtons(false);
    const submitBtn = document.getElementById('scan-submit-btn');
    if(submitBtn && submitBtn.dataset.mode === 'reset'){
        setScanStatus('当前为识别结果页，如需继续识别，请先点击“重置”。', 'info');
    }
}

initScanPage();
</script>
</head>
<body>
<div class="wrap">
    <div class="card"><a href="/">返回查询页</a></div>

    <div class="card">
        <div class="scan-icon-wrap">
            <div class="scan-icon" aria-hidden="true">
                <svg viewBox="0 0 24 24">
                    <path d="M8 4H6a2 2 0 0 0-2 2v2"></path>
                    <path d="M16 4h2a2 2 0 0 1 2 2v2"></path>
                    <path d="M8 20H6a2 2 0 0 1-2-2v-2"></path>
                    <path d="M16 20h2a2 2 0 0 0 2-2v-2"></path>
                    <path d="M5 12h14"></path>
                </svg>
            </div>
        </div>
        <h2 style="text-align:center;margin-top:0;">拍照识别资产编号</h2>
        <div id="scan-live-status" class="live-status"></div>
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form id="scan-form" method="post" enctype="multipart/form-data" autocomplete="off" onsubmit="return triggerSubmitScan(event)">
            <div class="row">
                <label>拍照或上传标签图片</label>
                <input id="scan-image-input" type="file" name="scan_image" accept="image/*" capture="environment" required onchange="updateSelectedImageTip(); setScanStatus('', 'info');">
                <div id="scan-file-tip" class="muted" style="display:none;margin-top:6px;"></div>
            </div>
            <div id="scan-progress-wrap" class="progress-wrap">
                <div class="progress-bar-bg"><div id="scan-progress-fill" class="progress-bar-fill"></div></div>
                <div id="scan-progress-text" class="progress-text">准备开始识别...</div>
                <div id="scan-progress-time" class="progress-time"></div>
            </div>
            <div class="action-row">
                <button id="scan-submit-btn" type="{{ 'button' if recognized_no else 'submit' }}" data-mode="{{ 'reset' if recognized_no else 'scan' }}" data-idle-text="{{ '重置' if recognized_no else '识别' }}" {% if recognized_no %}onclick="resetScanPage(); return false;"{% endif %}>{{ '重置' if recognized_no else '识别' }}</button>
                <button id="scan-back-btn" type="button" class="btn-gray" onclick="handleBackOrStop(); return false;">返回</button>
            </div>
        </form>
    </div>

    {% if recognized_no %}
    <div class="card">
        <div class="msg">识别成功</div>
        <div class="row">
            <label>识别结果（{{ recognized_label }}）</label>
            <div class="result-box">{{ recognized_no }}</div>
        </div>

        {% if target_row %}
        <div class="row">
            <label>匹配结果</label>
            <div class="msg">已找到对应设备：{{ target_row.type_text }} / {{ target_row.name }}</div>
            <div class="action-row">
                <a href="{{ target_row.detail_url }}"><button type="button" class="btn-green">打开对应设备</button></a>
            </div>
        </div>
        {% else %}
        <div class="row">
            <label>匹配结果</label>
            <div class="muted">当前系统中未找到该{{ recognized_label }}对应的设备，可直接新建设备并自动带入该编号。</div>
            <div class="action-row">
                <a href="/device/new?type=主设备&recognized_no={{ recognized_no }}"><button type="button" class="btn-green">新增主设备</button></a>
            </div>
        </div>
        {% endif %}

        {% if matched_rows %}
        <div class="row">
            <label>相关匹配列表</label>
            <div class="muted">
                {% for row in matched_rows[:8] %}
                    <div><a href="{{ row.detail_url }}">{{ row.type_text }} - {{ row.name }} - {{ row.group_no or row.internal_no }}</a></div>
                {% endfor %}
            </div>
        </div>
        {% endif %}
    </div>
    {% endif %}

    {% if debug_texts %}
    <div class="card">
        <div class="row">
            <label>识别调试信息</label>
            <div class="debug-box">{{ debug_texts|join('
') }}</div>
        </div>
    </div>
    {% endif %}
</div>
</body>
</html>
"""


DEVICE_NEW_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>新增设备</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:900px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.err{color:red;margin-bottom:10px;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
let activeImageInputId = '';

function validateAssetGroupNoValue(value){
    value = (value || '').trim();
    if(!value){
        return '';
    }
    return /^\\d{18}$/.test(value) ? '' : '集团编号必须为18位数字';
}

function validateAccessoryGroupNoValue(value){
    value = (value || '').trim();
    if(!value){
        return '';
    }
    return /^\\d{18}-\\d+$/.test(value) ? '' : '附属资产集团编号必须为18位主编号-序号';
}

function confirmDeviceSave(form){
    const deviceType = (form.querySelector('[name="device_type"]') || {}).value || '主设备';
    const groupNo = ((form.querySelector('[name="group_no"]') || {}).value || '').trim();
    const message = deviceType === '配件' ? validateAccessoryGroupNoValue(groupNo) : validateAssetGroupNoValue(groupNo);
    if(message){
        alert(message);
        return false;
    }
    return confirm('确认保存吗？');
}

function toggleType(){
    const type = document.getElementById('device_type').value;
    const parentRow = document.getElementById('parent_row');
    if(type === '配件'){
        parentRow.style.display = 'block';
    }else{
        parentRow.style.display = 'none';
    }
}

function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.add('show');
    }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.remove('show');
    }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){
        return;
    }

    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });

    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }

    if(dialogId){
        closeUploadChooser(dialogId);
    }
}
</script>
</head>
<body onload="toggleType()">
<div class="wrap">
    <div class="card"><a href="/">返回查询页</a></div>
    <div class="card">
        <h2>新增设备</h2>
        <div style="color:#666;margin-bottom:10px;">内部编号和集团编号至少填写一个。新增配件时，已自动带出主设备编号前缀。</div>
        {% if error %}<div class="err">{{ error }}</div>{% endif %}
        <form method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row">
                    <label>类型</label>
                    <select name="device_type" id="device_type" onchange="toggleType()">
                        <option value="主设备" {% if form_data.device_type == '主设备' %}selected{% endif %}>主设备</option>
                        <option value="配件" {% if form_data.device_type == '配件' %}selected{% endif %}>配件</option>
                    </select>
                </div>
                <div class="row" id="parent_row">
                    <label>所属主设备ID（可空）</label>
                    <input type="text" name="parent_asset_id" value="{{ form_data.parent_asset_id }}">
                </div>
                <div class="row"><label>集团编号（可空）</label><input type="text" name="group_no" value="{{ form_data.group_no }}"></div>
                <div class="row"><label>内部编号（可空）</label><input type="text" name="internal_no" value="{{ form_data.internal_no }}"></div>
                <div class="row"><label>名称</label><input type="text" name="name" value="{{ form_data.name }}"></div>
                <div class="row"><label>型号</label><input type="text" name="model" value="{{ form_data.model }}"></div>
                <div class="row"><label>责任人</label><input type="text" name="owner" value="{{ form_data.owner }}"></div>
                <div class="row"><label>位置</label><input type="text" name="location" value="{{ form_data.location }}"></div>
                <div class="row"><label>时间</label><input type="date" name="asset_date" value="{{ form_data.asset_date }}"></div>
                <div class="row">
                    <label>状态</label>
                    <select name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s.dict_value }}" {% if form_data.status == s.dict_value %}selected{% endif %}>{{ s.dict_value }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" onclick="openUploadChooser('device-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="device-image-files-text" class="file-list" data-inputs="device-camera-files,device-file-files">未选择图片</div>
                    <div id="device-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('device-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input type="file" id="device-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('device-camera-files', 'device-image-files-text', 'device-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input type="file" id="device-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('device-file-files', 'device-image-files-text', 'device-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('device-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            <div class="row"><label>备注</label><textarea name="remark">{{ form_data.remark }}</textarea></div>
            <button type="submit">保存</button>
        </form>
    </div>
</div>
</body>
</html>
"""

ASSET_DETAIL_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>主设备详情</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:1100px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.btn-red{background:#dc3545;}
.btn-green{background:#198754;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.table-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;min-width:900px;}
th,td{border:1px solid #ddd;padding:10px;text-align:left;}
th{background:#f0f3f8;}
.msg{color:green;margin-bottom:10px;}
.err{color:red;margin-bottom:10px;}
.readonly{background:#e9ecef;color:#666;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
function validateAssetGroupNoValue(value){
    value = (value || '').trim();
    if(!value){
        return '';
    }
    return /^\\d{18}$/.test(value) ? '' : '集团编号必须为18位数字';
}

function confirmAssetSave(form){
    const groupNo = ((form.querySelector('[name="group_no"]') || {}).value || '').trim();
    const message = validateAssetGroupNoValue(groupNo);
    if(message){
        alert(message);
        return false;
    }
    return confirm('确认保存吗？');
}

function enableEdit(formId){
    const form = document.getElementById(formId);
    const fields = form.querySelectorAll('.edit-field');
    fields.forEach(el => {
        el.disabled = false;
        el.classList.remove('readonly');
    });
    document.getElementById(formId + '-save').style.display = 'inline-block';
    document.getElementById(formId + '-edit').style.display = 'none';
}

function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.add('show');
    }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.remove('show');
    }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){
        return;
    }

    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });

    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }

    if(dialogId){
        closeUploadChooser(dialogId);
    }
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card"><a href="/">返回查询页</a></div>

    <div class="card">
        <h2>主设备详情</h2>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form id="asset-form" method="post" enctype="multipart/form-data" onsubmit="return confirmAssetSave(this)">
            <input type="hidden" name="action" value="save_asset">
            <div class="grid">
                <div class="row"><label>集团编号</label><input class="edit-field readonly" disabled type="text" name="group_no" value="{{ asset.group_no or '' }}"></div>
                <div class="row"><label>内部编号</label><input class="edit-field readonly" disabled type="text" name="internal_no" value="{{ asset.internal_no or '' }}"></div>
                <div class="row"><label>名称</label><input class="edit-field readonly" disabled type="text" name="name" value="{{ asset.name }}"></div>
                <div class="row"><label>型号</label><input class="edit-field readonly" disabled type="text" name="model" value="{{ asset.model or '' }}"></div>
                <div class="row"><label>责任人</label><input class="edit-field readonly" disabled type="text" name="owner" value="{{ asset.owner or '' }}"></div>
                <div class="row"><label>位置</label><input class="edit-field readonly" disabled type="text" name="location" value="{{ asset.location or '' }}"></div>
                <div class="row"><label>时间</label><input class="edit-field readonly" disabled type="date" name="asset_date" value="{{ asset.asset_date.isoformat() if asset.asset_date else today }}"></div>
                <div class="row">
                    <label>状态</label>
                    <select class="edit-field readonly" disabled name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s.dict_value }}" {% if asset.status == s.dict_value %}selected{% endif %}>{{ s.dict_value }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" class="edit-field readonly" disabled onclick="openUploadChooser('asset-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="asset-image-files-text" class="file-list" data-inputs="asset-camera-files,asset-file-files">未选择图片</div>
                    <div id="asset-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('asset-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input class="edit-field readonly" disabled type="file" id="asset-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('asset-camera-files', 'asset-image-files-text', 'asset-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input class="edit-field readonly" disabled type="file" id="asset-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('asset-file-files', 'asset-image-files-text', 'asset-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('asset-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <div class="row">
                <label>当前图片</label>
                {% if images %}
                    {% for img in images %}
                    <div class="image-row">
                        <a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a>
                        <label style="display:flex;align-items:center;gap:6px;font-weight:normal;">
                            <input class="edit-field readonly" disabled type="checkbox" name="delete_asset_image_ids" value="{{ img.id }}">
                            删除
                        </label>
                    </div>
                    {% endfor %}
                    <div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“保存”。</div>
                {% else %}
                    <div>暂无图片</div>
                {% endif %}
            </div>

            <div class="row"><label>备注</label><textarea class="edit-field readonly" disabled name="remark">{{ asset.remark or '' }}</textarea></div>

            <button type="button" id="asset-form-edit" onclick="enableEdit('asset-form')">修改</button>
            <button type="submit" id="asset-form-save" style="display:none;">保存</button>
        </form>

        <form method="post" action="/asset/{{ asset.id }}/delete" onsubmit="return confirm('确认删除该主设备及其所有配件？');" style="margin-top:12px;">
            <button class="btn-red" type="submit">删除主设备</button>
        </form>
    </div>


    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2 style="margin:0;">配件列表</h2>
            <a href="/device/new?type=配件&parent_asset_id={{ asset.id }}"><button type="button" class="btn-green" style="width:auto;">+</button></a>
        </div>

        <div class="table-wrap" style="margin-top:12px;">
            <table>
                <thead>
                    <tr>
                        <th>内部编号</th>
                        <th>集团编号</th>
                        <th>名称</th>
                        <th>型号</th>
                        <th>状态</th>
                        <th>责任人</th>
                        <th>图片</th>
                    </tr>
                </thead>
                <tbody>
                    {% for item in accessories %}
                    <tr>
                        <td><a href="/accessory/{{ item.id }}">{{ item.sub_internal_no }}</a></td>
                        <td><a href="/accessory/{{ item.id }}">{{ item.sub_group_no }}</a></td>
                        <td>{{ item.name }}</td>
                        <td>{{ item.model or '' }}</td>
                        <td>{{ item.status or '' }}</td>
                        <td>{{ item.owner or '' }}</td>
                        <td>{{ item.images|length }}</td>
                    </tr>
                    {% endfor %}
                    {% if not accessories %}
                    <tr><td colspan="7">暂无配件</td></tr>
                    {% endif %}
                </tbody>
            </table>
        </div>
    </div>
</div>
</body>
</html>
"""

ACCESSORY_DETAIL_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>配件详情</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:900px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.btn-red{background:#dc3545;}
.readonly{background:#e9ecef;color:#666;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.msg{color:green;margin-bottom:10px;}
.err{color:red;margin-bottom:10px;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
function validateAccessoryGroupNoValue(value){
    value = (value || '').trim();
    if(!value){
        return '';
    }
    return /^\\d{18}-\\d+$/.test(value) ? '' : '附属资产集团编号必须为18位主编号-序号';
}

function confirmAccessorySave(form){
    const groupNo = ((form.querySelector('[name="sub_group_no"]') || {}).value || '').trim();
    const message = validateAccessoryGroupNoValue(groupNo);
    if(message){
        alert(message);
        return false;
    }
    return confirm('确认保存吗？');
}

function enableEdit(formId){
    const form = document.getElementById(formId);
    const fields = form.querySelectorAll('.edit-field');
    fields.forEach(el => {
        el.disabled = false;
        el.classList.remove('readonly');
    });
    document.getElementById(formId + '-save').style.display = 'inline-block';
    document.getElementById(formId + '-edit').style.display = 'none';
}

function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.add('show');
    }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){
        dialog.classList.remove('show');
    }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){
        return;
    }

    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });

    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }

    if(dialogId){
        closeUploadChooser(dialogId);
    }
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card"><a href="{{ back_url }}">返回</a></div>

    <div class="card">
        <h2>配件详情</h2>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form id="accessory-form" method="post" enctype="multipart/form-data" onsubmit="return confirmAccessorySave(this)">
            <div class="grid">
                <div class="row"><label>附属资产内部编号</label><input class="edit-field readonly" disabled type="text" name="sub_internal_no" value="{{ accessory.sub_internal_no or '' }}"></div>
                <div class="row"><label>附属资产集团编号</label><input class="edit-field readonly" disabled type="text" name="sub_group_no" value="{{ accessory.sub_group_no or '' }}"></div>
                <div class="row"><label>名称</label><input class="edit-field readonly" disabled type="text" name="name" value="{{ accessory.name }}"></div>
                <div class="row"><label>型号</label><input class="edit-field readonly" disabled type="text" name="model" value="{{ accessory.model or '' }}"></div>
                <div class="row"><label>责任人</label><input class="edit-field readonly" disabled type="text" name="owner" value="{{ accessory.owner or '' }}"></div>
                <div class="row"><label>位置</label><input class="edit-field readonly" disabled type="text" name="location" value="{{ accessory.location or '' }}"></div>
                <div class="row"><label>时间</label><input class="edit-field readonly" disabled type="date" name="asset_date" value="{{ accessory.asset_date.isoformat() if accessory.asset_date else '' }}"></div>
                <div class="row">
                    <label>状态</label>
                    <select class="edit-field readonly" disabled name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s.dict_value }}" {% if accessory.status == s.dict_value %}selected{% endif %}>{{ s.dict_value }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" class="edit-field readonly" disabled onclick="openUploadChooser('accessory-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="accessory-image-files-text" class="file-list" data-inputs="accessory-camera-files,accessory-file-files">未选择图片</div>
                    <div id="accessory-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('accessory-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input class="edit-field readonly" disabled type="file" id="accessory-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('accessory-camera-files', 'accessory-image-files-text', 'accessory-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input class="edit-field readonly" disabled type="file" id="accessory-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('accessory-file-files', 'accessory-image-files-text', 'accessory-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('accessory-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <div class="row">
                <label>当前图片</label>
                {% if images %}
                    {% for img in images %}
                    <div class="image-row">
                        <a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a>
                        <label style="display:flex;align-items:center;gap:6px;font-weight:normal;">
                            <input class="edit-field readonly" disabled type="checkbox" name="delete_accessory_image_ids" value="{{ img.id }}">
                            删除
                        </label>
                    </div>
                    {% endfor %}
                    <div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“保存”。</div>
                {% else %}
                    <div>暂无图片</div>
                {% endif %}
            </div>

            <div class="row"><label>备注</label><textarea class="edit-field readonly" disabled name="remark">{{ accessory.remark or '' }}</textarea></div>

            <button type="button" id="accessory-form-edit" onclick="enableEdit('accessory-form')">修改</button>
            <button type="submit" id="accessory-form-save" style="display:none;">保存</button>
        </form>

        <form method="post" action="/accessory/{{ accessory.id }}/delete" onsubmit="return confirm('确认删除该配件？');" style="margin-top:12px;">
            <button class="btn-red" type="submit">删除配件</button>
        </form>
    </div>
</div>
</body>
</html>
"""
