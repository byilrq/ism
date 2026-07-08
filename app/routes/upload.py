"""upload.py — file upload / import / image handling for ISM."""

from datetime import datetime, date
from io import BytesIO
import os
import random
import re
import shutil
import string
import uuid

from flask import request, redirect, url_for, send_from_directory, abort, current_app
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_, func

from app import db, FlaskConfig as Config
from app.models import AssetImage, AccessoryImage, Asset

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
IMPORT_EXPORT_HEADERS = ["类型", "集团编号", "内部编号", "名称", "型号", "责任人", "位置", "时间", "状态", "备注"]
IMPORT_LOG_SUBDIR = "import_logs"
RECYCLE_SUBDIR = "recycle"


# ---------------------------------------------------------------------------
# Minimal normalize helpers (self-contained to avoid circular import)
# ---------------------------------------------------------------------------

def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_empty_to_none(value):
    value = _normalize_text(value)
    return value if value else None


def _normalize_status_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_location(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return _normalize_location(value[0]) if value else ""
    value = _normalize_text(value)
    if not value:
        return ""
    tuple_match = re.fullmatch(r"\(\s*[\'\"](.+?)[\'\"]\s*,\s*\)", value)
    if tuple_match:
        value = _normalize_text(tuple_match.group(1))
    return value.upper()


def _location_equals(column, value):
    value = _normalize_location(value)
    return func.upper(column) == value


def _prefer_new_value(new_value, old_value=None):
    new_value = _normalize_text(new_value)
    if new_value:
        return new_value
    return _normalize_text(old_value)


def _prefer_new_date(new_value, old_value=None):
    return new_value if new_value else old_value


def _split_accessory_number(value):
    value = _normalize_text(value)
    if not value or "-" not in value:
        return "", ""
    main_no, suffix = value.rsplit("-", 1)
    return _normalize_text(main_no), _normalize_text(suffix)


def _normalize_group_accessory_suffix(suffix):
    suffix = _normalize_text(suffix)
    if suffix.isdigit():
        return str(int(suffix))
    return suffix


def _normalize_internal_accessory_suffix(suffix):
    suffix = _normalize_text(suffix)
    if suffix.isdigit():
        return suffix.zfill(3)
    return suffix


def extract_main_number(value):
    value = _normalize_text(value)
    if not value:
        return ""
    return value.split("-", 1)[0]


def sanitize_image_prefix(value):
    value = extract_main_number(value)
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("._-")
    return value or "asset"


# ---------------------------------------------------------------------------
# AssetLocationImage model
# ---------------------------------------------------------------------------

class AssetLocationImage(db.Model):
    __tablename__ = "asset_location_image"
    id = db.Column(db.Integer, primary_key=True)
    location_name = db.Column(db.String(255), nullable=False, index=True)
    image_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


# ---------------------------------------------------------------------------
# Image file handling
# ---------------------------------------------------------------------------

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def build_image_filename_prefix(group_no="", internal_no="", parent_asset=None):
    if parent_asset:
        parent_group_no = sanitize_image_prefix(getattr(parent_asset, "group_no", ""))
        if parent_group_no != "asset":
            return parent_group_no
        parent_internal_no = sanitize_image_prefix(getattr(parent_asset, "internal_no", ""))
        if parent_internal_no != "asset":
            return parent_internal_no
    group_prefix = sanitize_image_prefix(group_no)
    if group_prefix != "asset":
        return group_prefix
    internal_prefix = sanitize_image_prefix(internal_no)
    if internal_prefix != "asset":
        return internal_prefix
    return "asset"


def save_uploaded_image(file_storage, subdir, filename_prefix="asset"):
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        return None
    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    date_part = datetime.now().strftime("%Y.%m.%d")
    random_part = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
    safe_prefix = sanitize_image_prefix(filename_prefix)
    filename = f"{safe_prefix}.{date_part}.{random_part}.{ext}"
    folder = os.path.join(Config.UPLOAD_FOLDER, subdir)
    os.makedirs(folder, exist_ok=True)
    abs_path = os.path.join(folder, filename)
    file_storage.save(abs_path)
    return f"{subdir}/{filename}"


def move_image_file_to_recycle(relative_path):
    if not relative_path:
        return
    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return
    if safe_relative_path == RECYCLE_SUBDIR or safe_relative_path.startswith(f"{RECYCLE_SUBDIR}/"):
        return
    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if not os.path.exists(abs_path) or not os.path.isfile(abs_path):
        return
    recycle_dir = os.path.join(Config.UPLOAD_FOLDER, RECYCLE_SUBDIR, os.path.dirname(safe_relative_path))
    os.makedirs(recycle_dir, exist_ok=True)
    recycle_filename = os.path.basename(safe_relative_path)
    target_path = os.path.join(recycle_dir, recycle_filename)
    if os.path.exists(target_path):
        name, ext = os.path.splitext(recycle_filename)
        target_path = os.path.join(recycle_dir, f"{name}.{uuid.uuid4().hex[:8]}{ext}")
    shutil.move(abs_path, target_path)


def delete_image_file(relative_path):
    move_image_file_to_recycle(relative_path)


def permanent_delete_image_file(relative_path):
    if not relative_path:
        return
    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return
    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if os.path.exists(abs_path) and os.path.isfile(abs_path):
        os.remove(abs_path)


def trim_asset_images(asset):
    images = AssetImage.query.filter_by(asset_id=asset.id).order_by(
        AssetImage.created_at.asc(), AssetImage.id.asc()
    ).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def trim_accessory_images(accessory):
    images = AccessoryImage.query.filter_by(accessory_id=accessory.id).order_by(
        AccessoryImage.created_at.asc(), AccessoryImage.id.asc()
    ).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def get_asset_location_images(location):
    location = _normalize_location(location)
    if not location:
        return []
    return AssetLocationImage.query.filter(
        _location_equals(AssetLocationImage.location_name, location)
    ).order_by(AssetLocationImage.created_at.asc(), AssetLocationImage.id.asc()).all()


def trim_asset_location_images(location):
    images = get_asset_location_images(location)
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def delete_accessory_with_files(accessory):
    if not accessory:
        return
    accessory.previous_status = accessory.status
    accessory.status = "已删除"
    accessory.deleted_at = datetime.now()


def delete_asset_with_files(asset, get_asset_related_accessories=None):
    if not asset:
        return
    if get_asset_related_accessories:
        for accessory in get_asset_related_accessories(asset):
            delete_accessory_with_files(accessory)
    asset.previous_status = asset.status
    asset.status = "已删除"
    asset.deleted_at = datetime.now()


def restore_asset(asset):
    if not asset or not asset.deleted_at:
        return False
    asset.status = asset.previous_status if asset.previous_status else "正常"
    asset.deleted_at = None
    asset.previous_status = None
    return True


def restore_accessory(accessory):
    if not accessory or not accessory.deleted_at:
        return False
    accessory.status = accessory.previous_status if accessory.previous_status else "正常"
    accessory.deleted_at = None
    accessory.previous_status = None
    return True


def permanent_delete_asset(asset):
    if not asset:
        return
    for img in AssetImage.query.filter_by(asset_id=asset.id).all():
        delete_image_file(img.image_path)
    db.session.delete(asset)


def permanent_delete_accessory(accessory):
    if not accessory:
        return
    for img in AccessoryImage.query.filter_by(accessory_id=accessory.id).all():
        delete_image_file(img.image_path)
    db.session.delete(accessory)


# ---------------------------------------------------------------------------
# Import helpers — 类型识别、挂靠匹配、编号补齐、校验、更新
# 完全遵循"资产上传与更新规则（完整版）"
# ---------------------------------------------------------------------------

def _infer_device_type(group_no="", internal_no=""):
    """按规格自动判定资产类型，不依赖 Excel 类型列。

    优先级（同规则表）：
    1. 集团编号有 - 后缀 → 配件
    2. 内部编号有 - 后缀 → 配件
    3. 其他 → 主设备
    """
    group_no = _normalize_text(group_no)
    internal_no = _normalize_text(internal_no)

    if group_no and "-" in group_no:
        return "配件"
    if internal_no and "-" in internal_no:
        return "配件"

    return "主设备"


def _is_group_no_18(value):
    """是否为18位纯数字集团编号"""
    return bool(re.fullmatch(r"\d{18}", _normalize_text(value)))


def _validate_asset_group_no(group_no):
    """主设备集团编号校验：18位纯数字"""
    group_no = _normalize_text(group_no)
    if group_no and not _is_group_no_18(group_no):
        return "集团编号必须为18位纯数字"
    return ""


def _validate_accessory_group_no(group_no, internal_no=""):
    """配件集团编号校验：18位主编号-序号"""
    group_no = _normalize_text(group_no)
    if not group_no:
        return ""
    internal_no = _normalize_text(internal_no)
    if internal_no and "-" in internal_no:
        return ""
    if not re.fullmatch(r"\d{18}-\d+", group_no):
        return "附属资产集团编号必须为18位主编号-序号"
    return ""


def _resolve_parent_asset_id(internal_no="", group_no=""):
    """配件挂靠主设备的双向匹配（规则第三节）

    优先级：
    1. 从配件集团编号提取主部分 → 匹配 Asset.group_no
    2. 从配件内部编号提取主部分 → 匹配 Asset.internal_no
    如果两个主部分匹配到不同主设备 → 报错
    """
    from app.models import Asset

    internal_no = _normalize_text(internal_no)
    group_no = _normalize_text(group_no)

    base_group = _normalize_text(group_no.rsplit("-", 1)[0]) if "-" in group_no else ""
    base_internal = _normalize_text(internal_no.rsplit("-", 1)[0]) if "-" in internal_no else ""

    parent_by_group = Asset.query.filter_by(group_no=base_group).first() if base_group else None
    parent_by_internal = Asset.query.filter_by(internal_no=base_internal).first() if base_internal else None

    if parent_by_group and parent_by_internal and parent_by_group.id != parent_by_internal.id:
        raise ValueError(f"配件集团编号主部分({base_group})与内部编号主部分({base_internal})对应到不同主设备，无法挂靠")

    return parent_by_group or parent_by_internal


def _auto_complete_numbers(internal_no="", group_no="", parent_asset=None):
    """编号自动补齐（规则第四节）"""
    internal_no = _normalize_text(internal_no)
    group_no = _normalize_text(group_no)
    if not parent_asset:
        return internal_no, group_no

    # 有内部编号，缺集团编号
    if internal_no and not group_no and "-" in internal_no:
        suffix = internal_no.rsplit("-", 1)[1]
        if parent_asset.group_no:
            group_suffix = _normalize_group_accessory_suffix(suffix)
            group_no = f"{parent_asset.group_no}-{group_suffix}"

    # 有集团编号，缺内部编号
    if group_no and not internal_no and "-" in group_no:
        suffix = group_no.rsplit("-", 1)[1]
        if parent_asset.internal_no:
            internal_suffix = _normalize_internal_accessory_suffix(suffix)
            internal_no = f"{parent_asset.internal_no}-{internal_suffix}"

    return internal_no, group_no


def _find_existing_asset(internal_no="", group_no="", exclude_id=None):
    """通过编号查找已存在的 Asset（主设备）

    规则五主设备唯一性约束：集团编号或内部编号匹配即视为已有设备。
    """
    from app.models import Asset
    conditions = []
    if internal_no:
        conditions.append(Asset.internal_no == internal_no)
    if group_no:
        conditions.append(Asset.group_no == group_no)
    if not conditions:
        return None
    query = Asset.query.filter(or_(*conditions))
    if exclude_id is not None:
        query = query.filter(Asset.id != exclude_id)
    return query.first()


def _find_existing_accessory(internal_no="", group_no="", exclude_id=None):
    """通过编号查找已存在的 Accessory（配件）"""
    from app.models import Accessory
    conditions = []
    if internal_no:
        conditions.append(Accessory.sub_internal_no == internal_no)
    if group_no:
        conditions.append(Accessory.sub_group_no == group_no)
    if not conditions:
        return None
    query = Accessory.query.filter(or_(*conditions))
    if exclude_id is not None:
        query = query.filter(Accessory.id != exclude_id)
    return query.first()


def _upsert_asset_from_row(row_data):
    """主设备新增/更新（规则六覆盖策略）"""
    from app.models import Asset

    internal_no = _normalize_text(row_data.get("内部编号"))
    group_no = _normalize_text(row_data.get("集团编号"))
    name = _normalize_text(row_data.get("名称"))
    model = _normalize_text(row_data.get("型号"))
    owner = _normalize_text(row_data.get("责任人"))
    location = _normalize_location(row_data.get("位置"))
    status = _normalize_status_value(row_data.get("状态"))
    remark = _normalize_text(row_data.get("备注"))
    asset_date = parse_import_date(row_data.get("时间"))

    if not group_no and not internal_no:
        raise ValueError("集团编号和内部编号至少填写一个")

    if group_no:
        err = _validate_asset_group_no(group_no)
        if err:
            raise ValueError(err)

    obj = _find_existing_asset(internal_no=internal_no, group_no=group_no)
    if obj:
        # 编号字段：Excel有值才覆盖，空白保留现有值
        if group_no:
            obj.group_no = group_no
        if internal_no:
            obj.internal_no = internal_no
        obj.name = _prefer_new_value(name, obj.name) or obj.internal_no or obj.group_no or obj.name
        obj.model = _prefer_new_value(model, obj.model)
        obj.owner = _prefer_new_value(owner, obj.owner)
        obj.location = _prefer_new_value(location, obj.location)
        obj.status = _prefer_new_value(status, obj.status)
        obj.remark = _prefer_new_value(remark, obj.remark)
        obj.asset_date = _prefer_new_date(asset_date, obj.asset_date)
    else:
        obj = Asset(
            group_no=_normalize_empty_to_none(group_no),
            internal_no=_normalize_empty_to_none(internal_no),
            name=name or internal_no or group_no,
            model=model, owner=owner, location=location,
            status=status, remark=remark,
            asset_date=asset_date or date.today(),
        )
        db.session.add(obj)

    db.session.flush()
    return obj


def _upsert_accessory_from_row(row_data, skip_auto_complete=False):
    """配件新增/更新（规则三、四、六）

    skip_auto_complete: 混合上传时跳过编号补齐
    """
    from app.models import Accessory, Asset

    internal_no = _normalize_text(row_data.get("内部编号"))
    group_no = _normalize_text(row_data.get("集团编号"))
    name = _normalize_text(row_data.get("名称"))
    model = _normalize_text(row_data.get("型号"))
    owner = _normalize_text(row_data.get("责任人"))
    location = _normalize_location(row_data.get("位置"))
    status = _normalize_status_value(row_data.get("状态"))
    remark = _normalize_text(row_data.get("备注"))
    asset_date = parse_import_date(row_data.get("时间"))

    if not group_no and not internal_no:
        raise ValueError("集团编号和内部编号至少填写一个")

    if group_no:
        err = _validate_accessory_group_no(group_no, internal_no)
        if err:
            raise ValueError(err)

    # 双向匹配主设备（规则三）
    parent = _resolve_parent_asset_id(internal_no=internal_no, group_no=group_no)
    parent_asset_id = parent.id if parent else None

    if not parent_asset_id:
        raise ValueError("未找到匹配的主设备，无法挂靠配件")

    # 编号自动补齐（规则四）
    if not skip_auto_complete:
        internal_no, group_no = _auto_complete_numbers(
        internal_no=internal_no, group_no=group_no, parent_asset=parent
    )
    row_data["内部编号"] = internal_no

    # 主部分一致性校验（规则五末尾）
    if internal_no and "-" in internal_no and group_no and "-" in group_no:
        base_internal = internal_no.rsplit("-", 1)[0]
        base_group = group_no.rsplit("-", 1)[0]
        if base_internal != base_group:
            parent_i = Asset.query.filter_by(internal_no=base_internal).first()
            parent_g = Asset.query.filter_by(group_no=base_group).first()
            if parent_i and parent_g and parent_i.id != parent_g.id:
                raise ValueError(f"配件内部编号主部分({base_internal})与集团编号主部分({base_group})对应到不同主设备")

    obj = _find_existing_accessory(internal_no=internal_no, group_no=group_no)
    if obj:
        if group_no:
            obj.sub_group_no = group_no
        if internal_no:
            obj.sub_internal_no = internal_no
        obj.parent_asset_id = parent_asset_id or obj.parent_asset_id
        obj.name = _prefer_new_value(name, obj.name) or obj.sub_internal_no or obj.sub_group_no or obj.name
        obj.model = _prefer_new_value(model, obj.model)
        obj.owner = _prefer_new_value(owner, obj.owner)
        obj.location = _prefer_new_value(location, obj.location)
        obj.status = _prefer_new_value(status, obj.status)
        obj.remark = _prefer_new_value(remark, obj.remark)
        obj.asset_date = _prefer_new_date(asset_date, obj.asset_date)
    else:
        obj = Accessory(
            parent_asset_id=parent_asset_id,
            sub_group_no=_normalize_empty_to_none(group_no),
            sub_internal_no=_normalize_empty_to_none(internal_no),
            name=name or internal_no or group_no,
            model=model, owner=owner, location=location,
            status=status, remark=remark,
            asset_date=asset_date or date.today(),
        )
        db.session.add(obj)

    db.session.flush()
    return obj


# ---------------------------------------------------------------------------
# Excel 解析基础函数
# ---------------------------------------------------------------------------

def build_import_column_map(ws):
    header_values = [_normalize_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    expected_headers = IMPORT_EXPORT_HEADERS
    actual_headers = header_values[:len(expected_headers)]
    if actual_headers != expected_headers:
        raise ValueError(
            "Excel表头必须与导出格式完全一致："
            + "、".join(expected_headers)
            + "；当前表头："
            + "、".join(actual_headers)
        )
    return {header: idx for idx, header in enumerate(expected_headers)}, 2


def get_import_row_value(row, column_map, column_name):
    idx = column_map.get(column_name)
    if idx is None or idx >= len(row):
        return ""
    return _normalize_text(row[idx])


def parse_import_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_text(value)
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def build_import_row_data(row, column_map):
    time_idx = column_map.get("时间")
    time_value = row[time_idx] if time_idx is not None and time_idx < len(row) else None
    return {
        "类型": get_import_row_value(row, column_map, "类型"),
        "内部编号": get_import_row_value(row, column_map, "内部编号"),
        "集团编号": get_import_row_value(row, column_map, "集团编号"),
        "名称": get_import_row_value(row, column_map, "名称"),
        "型号": get_import_row_value(row, column_map, "型号"),
        "责任人": get_import_row_value(row, column_map, "责任人"),
        "位置": get_import_row_value(row, column_map, "位置"),
        "时间": time_value,
        "状态": get_import_row_value(row, column_map, "状态"),
        "备注": get_import_row_value(row, column_map, "备注"),
    }


def is_import_row_empty(row_data):
    return not any(
        _normalize_text(row_data.get(column_name))
        for column_name in ["类型", "内部编号", "集团编号", "名称", "型号", "责任人", "位置", "状态", "备注"]
    )


# ---------------------------------------------------------------------------
# Import entry points
# ---------------------------------------------------------------------------

def create_import_failure_log(failure_rows):
    if not failure_rows:
        return ""
    folder = os.path.join(Config.UPLOAD_FOLDER, IMPORT_LOG_SUBDIR)
    os.makedirs(folder, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "导入失败日志"
    ws.append(["Excel行号", "错误原因"] + IMPORT_EXPORT_HEADERS)
    for item in failure_rows:
        row_data = item.get("row_data", {})
        ws.append([
            item.get("row_index"),
            item.get("error"),
            _normalize_text(row_data.get("类型")),
            _normalize_text(row_data.get("集团编号")),
            _normalize_text(row_data.get("内部编号")),
            _normalize_text(row_data.get("名称")),
            _normalize_text(row_data.get("型号")),
            _normalize_text(row_data.get("责任人")),
            _normalize_text(row_data.get("位置")),
            parse_import_date(row_data.get("时间")).isoformat() if parse_import_date(row_data.get("时间")) else _normalize_text(row_data.get("时间")),
            _normalize_text(row_data.get("状态")),
            _normalize_text(row_data.get("备注")),
        ])
    filename = f"import_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(os.path.join(folder, filename))
    return filename


def import_devices_from_excel(file_storage):
    """混合上传：自动识别每行类型，主设备/配件统一处理。"""
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    column_map, data_start_row = build_import_column_map(ws)
    success_count = 0
    failure_rows = []
    for row_index, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        row_data = build_import_row_data(row, column_map)
        if is_import_row_empty(row_data):
            continue
        group_no = _normalize_text(row_data.get("集团编号"))
        internal_no = _normalize_text(row_data.get("内部编号"))
        device_type = _infer_device_type(group_no=group_no, internal_no=internal_no)
        row_data["类型"] = device_type
        try:
            with db.session.begin_nested():
                if device_type == "主设备":
                    _upsert_asset_from_row(row_data)
                else:
                    _upsert_accessory_from_row(row_data, skip_auto_complete=True)
            success_count += 1
        except Exception as e:
            failure_rows.append({"row_index": row_index, "error": str(e), "row_data": row_data})
    log_filename = create_import_failure_log(failure_rows)
    return {"success_count": success_count, "failed_count": len(failure_rows), "failure_rows": failure_rows, "log_filename": log_filename}


def import_accessories_from_excel(file_storage):
    """纯配件上传：规则二-2, 只导入配件，集团编号必须为 18位数字-序号。"""
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    column_map, data_start_row = build_import_column_map(ws)
    success_count = 0
    failure_rows = []
    for row_index, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        row_data = build_import_row_data(row, column_map)
        if is_import_row_empty(row_data):
            continue
        group_no = _normalize_text(row_data.get("集团编号"))
        internal_no = _normalize_text(row_data.get("内部编号"))
        try:
            if not re.fullmatch(r"\d{18}-\d+", group_no):
                raise ValueError("集团编号格式不正确，必须为18位主编号-序号")
            parent_base = _normalize_text(group_no.rsplit("-", 1)[0])
            suffix = _normalize_text(group_no.rsplit("-", 1)[1])
            parent = Asset.query.filter_by(group_no=parent_base).first()
            if not parent:
                failure_rows.append({"row_index": row_index, "error": f"未找到集团编号 {parent_base} 对应的主资产，已跳过", "row_data": row_data})
                continue
            if parent.internal_no and not internal_no:
                internal_suffix = _normalize_internal_accessory_suffix(suffix)
                internal_no = f"{parent.internal_no}-{internal_suffix}"
            row_data["内部编号"] = internal_no
            with db.session.begin_nested():
                _upsert_accessory_from_row(row_data)
            success_count += 1
        except Exception as e:
            failure_rows.append({"row_index": row_index, "error": str(e), "row_data": row_data})
    log_filename = create_import_failure_log(failure_rows)
    return {"success_count": success_count, "failed_count": len(failure_rows), "failure_rows": failure_rows, "log_filename": log_filename}


# ---------------------------------------------------------------------------
# Route registration
# ---------------------------------------------------------------------------

def register_upload_routes(app):
    with app.app_context():
        AssetLocationImage.__table__.create(bind=db.engine, checkfirst=True)

    from app.routes import ensure_read_access, ensure_manage_access

    @app.route("/uploads/<path:filename>")
    def uploaded_file(filename):
        safe_path = os.path.normpath(filename)
        if safe_path.startswith(".."):
            abort(404)
        return send_from_directory(Config.UPLOAD_FOLDER, safe_path)

    @app.route("/import_logs/<path:filename>")
    def download_import_log(filename):
        guard = ensure_read_access()
        if guard:
            return guard
        safe_filename = os.path.basename(filename)
        if not safe_filename:
            abort(404)
        return send_from_directory(
            os.path.join(Config.UPLOAD_FOLDER, IMPORT_LOG_SUBDIR),
            safe_filename, as_attachment=True
        )

    @app.route("/import_devices", methods=["POST"])
    def import_devices():
        guard = ensure_manage_access()
        if guard:
            return guard
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            return redirect(url_for("search_assets"))
        try:
            result = import_devices_from_excel(excel_file)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            error_message = str(e)
            if "Excel表头必须" in error_message:
                error_message = "表头格式及顺序不正确。\n" + error_message
            else:
                error_message = "上传设备失败：" + error_message
            return redirect(url_for("search_assets", import_error=error_message))
        redirect_args = {"import_success": result.get("success_count", 0), "import_failed": result.get("failed_count", 0)}
        if result.get("log_filename"):
            redirect_args["import_log"] = result["log_filename"]
        return redirect(url_for("search_assets", **redirect_args))

    @app.route("/import_accessories", methods=["POST"])
    def import_accessories():
        guard = ensure_manage_access()
        if guard:
            return guard
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            return redirect(url_for("search_assets"))
        try:
            result = import_accessories_from_excel(excel_file)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            error_message = str(e)
            if "Excel表头必须" in error_message:
                error_message = "表头格式及顺序不正确。\n" + error_message
            else:
                error_message = "上传配件失败：" + error_message
            return redirect(url_for("search_assets", import_error=error_message))
        redirect_args = {"import_success": result.get("success_count", 0), "import_failed": result.get("failed_count", 0)}
        if result.get("log_filename"):
            redirect_args["import_log"] = result["log_filename"]
        return redirect(url_for("search_assets", **redirect_args))
