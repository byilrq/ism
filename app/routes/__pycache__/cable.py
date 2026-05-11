from collections import OrderedDict
from datetime import datetime
from io import BytesIO
import os
import re
import shutil
import uuid

from flask import request, redirect, url_for, render_template_string, send_file
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_

from app import db
from config import Config


ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
RECYCLE_SUBDIR = "recycle"
CABLE_SPECS = ["高频", "低频", "其它"]
CABLE_STATUSES = ["在库", "借出", "其它"]
CABLE_ADD_TYPES = ["电缆", "货架"]
SCAN_PARAM_NAMES = ["keyword", "code", "barcode", "text", "result", "data", "content", "q"]
EMPTY_CABLE_NO_PREFIX = "__EMPTY_CABLE_NO__"


class Cable(db.Model):
    __tablename__ = "cable"

    id = db.Column(db.Integer, primary_key=True)
    cable_no = db.Column(db.String(128), unique=True, nullable=False, index=True)
    name = db.Column(db.String(255), nullable=True)
    spec = db.Column(db.String(32), nullable=True)
    owner = db.Column(db.String(128), nullable=True, index=True)
    location = db.Column(db.String(255), nullable=True, index=True)
    status = db.Column(db.String(32), nullable=True, index=True)
    remark = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class CableImage(db.Model):
    __tablename__ = "cable_image"

    id = db.Column(db.Integer, primary_key=True)
    cable_id = db.Column(db.Integer, db.ForeignKey("cable.id"), nullable=False, index=True)
    image_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class CableShelf(db.Model):
    __tablename__ = "cable_shelf"

    id = db.Column(db.Integer, primary_key=True)
    shelf_name = db.Column(db.String(255), unique=True, nullable=False, index=True)
    remark = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class CableShelfImage(db.Model):
    __tablename__ = "cable_shelf_image"

    id = db.Column(db.Integer, primary_key=True)
    shelf_id = db.Column(db.Integer, db.ForeignKey("cable_shelf.id"), nullable=False, index=True)
    image_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_empty_to_none(value):
    value = normalize_text(value)
    return value if value else None


def normalize_location(value):
    if value is None:
        return ""

    if isinstance(value, (list, tuple)):
        if not value:
            return ""
        return normalize_location(value[0])

    value = normalize_text(value)
    if not value:
        return ""

    tuple_match = re.fullmatch(r"\(\s*['\"](.+?)['\"]\s*,\s*\)", value)
    if tuple_match:
        value = normalize_text(tuple_match.group(1))

    return value


def is_virtual_empty_cable_no(value):
    value = normalize_text(value)
    return bool(value) and value.startswith(EMPTY_CABLE_NO_PREFIX)


def get_display_cable_no(value):
    value = normalize_text(value)
    return "" if is_virtual_empty_cable_no(value) else value


def make_stored_cable_no(value="", existing_value=""):
    value = normalize_text(value)
    existing_value = normalize_text(existing_value)
    if value:
        return value
    if is_virtual_empty_cable_no(existing_value):
        return existing_value
    return f"{EMPTY_CABLE_NO_PREFIX}{uuid.uuid4().hex}"


def build_cable_filename_prefix(cable_no="", name="", location=""):
    return (
        get_display_cable_no(cable_no)
        or normalize_text(name)
        or normalize_location(location)
        or "cable"
    )


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def sanitize_image_prefix(value):
    value = normalize_text(value)
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("._-")
    return value or "cable"


def save_uploaded_image(file_storage, subdir, filename_prefix="cable"):
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        return None

    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    date_part = datetime.now().strftime("%Y.%m.%d")
    random_part = uuid.uuid4().hex[:16]
    safe_prefix = sanitize_image_prefix(filename_prefix)
    filename = f"{safe_prefix}.{date_part}.{random_part}.{ext}"

    folder = os.path.join(Config.UPLOAD_FOLDER, subdir)
    os.makedirs(folder, exist_ok=True)

    abs_path = os.path.join(folder, filename)
    file_storage.save(abs_path)
    return f"{subdir}/{filename}"


def move_image_file_to_recycle(relative_path):
    if not relative_path:
        return

    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return
    if safe_relative_path == RECYCLE_SUBDIR or safe_relative_path.startswith(f"{RECYCLE_SUBDIR}/"):
        return

    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if not os.path.exists(abs_path) or not os.path.isfile(abs_path):
        return

    recycle_dir = os.path.join(Config.UPLOAD_FOLDER, RECYCLE_SUBDIR, os.path.dirname(safe_relative_path))
    os.makedirs(recycle_dir, exist_ok=True)

    recycle_filename = os.path.basename(safe_relative_path)
    target_path = os.path.join(recycle_dir, recycle_filename)
    if os.path.exists(target_path):
        name, ext = os.path.splitext(recycle_filename)
        target_path = os.path.join(recycle_dir, f"{name}.{uuid.uuid4().hex[:8]}{ext}")

    shutil.move(abs_path, target_path)


def delete_image_file(relative_path):
    move_image_file_to_recycle(relative_path)


def permanent_delete_image_file(relative_path):
    if not relative_path:
        return

    safe_relative_path = os.path.normpath(relative_path).replace("\\", "/")
    if safe_relative_path.startswith(".."):
        return

    abs_path = os.path.join(Config.UPLOAD_FOLDER, safe_relative_path)
    if os.path.exists(abs_path) and os.path.isfile(abs_path):
        os.remove(abs_path)


def trim_cable_images(cable):
    images = CableImage.query.filter_by(cable_id=cable.id).order_by(CableImage.created_at.asc(), CableImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def trim_shelf_images(shelf):
    images = CableShelfImage.query.filter_by(shelf_id=shelf.id).order_by(CableShelfImage.created_at.asc(), CableShelfImage.id.asc()).all()
    while len(images) > 5:
        old = images.pop(0)
        permanent_delete_image_file(old.image_path)
        db.session.delete(old)


def delete_cable_with_files(cable):
    if not cable:
        return

    image_rows = CableImage.query.filter_by(cable_id=cable.id).all()
    for img in image_rows:
        delete_image_file(img.image_path)
        db.session.delete(img)

    if image_rows:
        db.session.flush()

    db.session.delete(cable)


def delete_shelf_with_files(shelf):
    if not shelf:
        return

    image_rows = CableShelfImage.query.filter_by(shelf_id=shelf.id).all()
    for img in image_rows:
        delete_image_file(img.image_path)
        db.session.delete(img)

    if image_rows:
        db.session.flush()

    db.session.delete(shelf)


def get_cable_images(cable_id):
    return CableImage.query.filter_by(cable_id=cable_id).order_by(CableImage.created_at.asc(), CableImage.id.asc()).all()


def get_shelf_images(shelf_id):
    return CableShelfImage.query.filter_by(shelf_id=shelf_id).order_by(CableShelfImage.created_at.asc(), CableShelfImage.id.asc()).all()


def get_shelf_by_location(location):
    location = normalize_location(location)
    if not location:
        return None
    return CableShelf.query.filter_by(shelf_name=location).first()


def ensure_shelf(location, auto_create=False):
    location = normalize_location(location)
    if not location:
        return None
    shelf = CableShelf.query.filter_by(shelf_name=location).first()
    if shelf or not auto_create:
        return shelf
    shelf = CableShelf(shelf_name=location)
    db.session.add(shelf)
    db.session.flush()
    return shelf


def get_all_shelf_locations():
    rows = CableShelf.query.order_by(CableShelf.shelf_name.asc(), CableShelf.id.asc()).all()
    values = []
    seen = set()
    for row in rows:
        location = normalize_location(row.shelf_name)
        if not location or location in seen:
            continue
        values.append(location)
        seen.add(location)
    return values


def build_location_shelf_lookup(locations):
    normalized_locations = [normalize_location(item) for item in locations if normalize_location(item)]
    if not normalized_locations:
        return {}
    shelves = CableShelf.query.filter(CableShelf.shelf_name.in_(normalized_locations)).all()
    return {normalize_location(item.shelf_name): item for item in shelves}


def backfill_cable_shelves():
    changed = False

    for cable in Cable.query.filter(Cable.location.isnot(None)).all():
        normalized = normalize_location(cable.location)
        if cable.location != normalized:
            cable.location = normalized or None
            changed = True

    shelf_rows = CableShelf.query.order_by(CableShelf.id.asc()).all()
    existing = {}
    for shelf in shelf_rows:
        normalized = normalize_location(shelf.shelf_name)
        if not normalized:
            continue

        keeper = existing.get(normalized)
        if keeper and keeper.id != shelf.id:
            for img in CableShelfImage.query.filter_by(shelf_id=shelf.id).all():
                img.shelf_id = keeper.id
                changed = True
            if not normalize_text(keeper.remark) and normalize_text(shelf.remark):
                keeper.remark = shelf.remark
                changed = True
            db.session.delete(shelf)
            changed = True
            continue

        if shelf.shelf_name != normalized:
            shelf.shelf_name = normalized
            changed = True
        existing[normalized] = shelf

    locations = (
        db.session.query(Cable.location)
        .filter(Cable.location.isnot(None))
        .distinct()
        .all()
    )
    for row in locations:
        location = normalize_location(row[0] if isinstance(row, (list, tuple)) else row)
        if location and location not in existing:
            shelf = CableShelf(shelf_name=location)
            db.session.add(shelf)
            existing[location] = shelf
            changed = True

    if changed:
        db.session.commit()


def extract_scan_prefill_code(scan_code=""):
    scan_code = normalize_text(scan_code)
    if scan_code:
        return scan_code
    for name in SCAN_PARAM_NAMES:
        value = normalize_text(request.args.get(name))
        if value:
            return value
    return ""


def extract_scan_keyword(scan_code=""):
    scan_code = normalize_text(scan_code)
    if scan_code:
        return scan_code
    keyword = normalize_text(request.args.get("keyword"))
    if keyword:
        return keyword
    return extract_scan_prefill_code("")


def build_cable_rows(keyword="", searched=False):
    if not searched:
        return []

    keyword = normalize_text(keyword)
    query = Cable.query
    if keyword:
        query = query.filter(
            or_(
                Cable.cable_no.like(f"%{keyword}%"),
                Cable.name.like(f"%{keyword}%"),
                Cable.owner.like(f"%{keyword}%"),
                Cable.location.like(f"%{keyword}%"),
                Cable.remark.like(f"%{keyword}%"),
            )
        )

    items = query.order_by(Cable.location.asc(), Cable.cable_no.asc(), Cable.id.asc()).all()
    shelf_lookup = build_location_shelf_lookup([item.location for item in items])

    rows = []
    for item in items:
        location = item.location or ""
        shelf = shelf_lookup.get(normalize_location(location)) if location else None
        rows.append({
            "row_type": "cable",
            "id": item.id,
            "selected_value": f"cable:{item.id}",
            "cable_no": get_display_cable_no(item.cable_no),
            "name": item.name or "",
            "spec": item.spec or "",
            "owner": item.owner or "",
            "location": location,
            "status": item.status or "",
            "remark": item.remark or "",
            "detail_url": url_for("cable_detail", cable_id=item.id),
            "location_detail_url": url_for("cable_location_detail", shelf_id=shelf.id) if shelf else "",
        })

    occupied_locations = {
        normalize_location(row[0])
        for row in db.session.query(Cable.location).filter(Cable.location.isnot(None)).distinct().all()
        if normalize_location(row[0])
    }

    shelf_query = CableShelf.query
    if keyword:
        shelf_query = shelf_query.filter(
            or_(
                CableShelf.shelf_name.like(f"%{keyword}%"),
                CableShelf.remark.like(f"%{keyword}%"),
            )
        )

    for shelf in shelf_query.order_by(CableShelf.shelf_name.asc(), CableShelf.id.asc()).all():
        shelf_name = normalize_location(shelf.shelf_name)
        if not shelf_name or shelf_name in occupied_locations:
            continue
        rows.append({
            "row_type": "shelf",
            "id": shelf.id,
            "selected_value": f"shelf:{shelf.id}",
            "cable_no": "",
            "name": "空货架",
            "spec": "",
            "owner": "",
            "location": normalize_location(shelf.shelf_name),
            "status": "",
            "remark": shelf.remark or "",
            "detail_url": url_for("cable_location_detail", shelf_id=shelf.id),
            "location_detail_url": url_for("cable_location_detail", shelf_id=shelf.id),
        })

    rows.sort(key=lambda x: (normalize_location(x.get("location")), 0 if x.get("row_type") == "shelf" else 1, normalize_text(x.get("cable_no")), x.get("id") or 0))
    return rows


def import_cables_from_excel(file_storage):
    wb = load_workbook(file_storage)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cable_no = normalize_text(row[0] if len(row) > 0 else "")
        name = normalize_text(row[1] if len(row) > 1 else "")
        spec = normalize_text(row[2] if len(row) > 2 else "")
        owner = normalize_text(row[3] if len(row) > 3 else "")
        location = normalize_text(row[4] if len(row) > 4 else "")
        status = normalize_text(row[5] if len(row) > 5 else "")
        remark = normalize_text(row[6] if len(row) > 6 else "")

        if not cable_no and not name and not spec and not owner and not location and not status and not remark:
            continue

        if spec and spec not in CABLE_SPECS:
            raise ValueError(f"第{row_idx}行规格不合法：{spec}")
        if status and status not in CABLE_STATUSES:
            raise ValueError(f"第{row_idx}行状态不合法：{status}")

        if location:
            ensure_shelf(location, auto_create=True)

        obj = Cable.query.filter_by(cable_no=cable_no).first() if cable_no else None
        if obj:
            obj.name = name or ""
            obj.spec = normalize_empty_to_none(spec)
            obj.owner = normalize_empty_to_none(owner)
            obj.location = normalize_empty_to_none(location)
            obj.status = normalize_empty_to_none(status)
            obj.remark = normalize_empty_to_none(remark)
        else:
            obj = Cable(
                cable_no=make_stored_cable_no(cable_no),
                name=name or "",
                spec=normalize_empty_to_none(spec),
                owner=normalize_empty_to_none(owner),
                location=normalize_empty_to_none(location),
                status=normalize_empty_to_none(status),
                remark=normalize_empty_to_none(remark),
            )
            db.session.add(obj)


CABLE_SEARCH_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>电缆查询</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:1200px;margin:auto;padding:12px;}
.card{background:#fff;border-radius:10px;padding:14px;margin-bottom:14px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.topbar{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;}
.switch-row{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
input,select,button{padding:10px;font-size:16px;border-radius:6px;border:1px solid #ccc;box-sizing:border-box;}
input[type=text]{min-width:260px;}
button{background:#0d6efd;color:#fff;border:none;cursor:pointer;}
.btn-green{background:#198754;}
.btn-gray{background:#6c757d;}
.btn-orange{background:#fd7e14;}
.btn-red{background:#dc3545;}
.table-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;min-width:920px;}
th,td{border:1px solid #ccc;padding:8px;text-align:left;}
th{background:#f0f3f8;}
.err{color:red;margin-top:10px;}
.muted{color:#666;font-size:14px;}
.num-link{color:#0d6efd;text-decoration:underline;}
.action-bar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;}
.pagination{margin-top:12px;display:flex;gap:10px;flex-wrap:wrap;}
@media (max-width:768px){input[type=text]{width:100%;}}
</style>
<script>
function toggleSelectCurrentPage(source){
    const items = document.querySelectorAll('input[name="selected_items"]');
    items.forEach(item => { item.checked = source.checked; });
}

function confirmDeleteSelected(){
    const checkedItems = Array.from(document.querySelectorAll('input[name="selected_items"]:checked'));
    if(checkedItems.length === 0){
        alert('请先勾选要删除的电缆或空货架');
        return false;
    }
    const cableCount = checkedItems.filter(item => item.dataset.rowType === 'cable').length;
    const shelfCount = checkedItems.filter(item => item.dataset.rowType === 'shelf').length;
    if(!confirm(`将删除 ${cableCount} 条电缆、${shelfCount} 个空货架，是否继续？`)){
        return false;
    }
    const pin = prompt('请输入4位Pin码确认批量删除');
    if(pin === null){
        return false;
    }
    if(pin !== '0819'){
        alert('Pin码错误，已取消批量删除');
        return false;
    }
    document.getElementById('delete_pin').value = pin;
    return true;
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="topbar">
            <div><strong>当前用户：</strong>{{ current_user.username if current_user.is_authenticated else "访客模式" }}</div>
            <div>{% if current_user.is_authenticated %}<a href="/logout">退出登录</a>{% else %}<span style="color:#666;">免登录访问</span>{% endif %}</div>
        </div>
    </div>

    <div class="card">
        <div class="switch-row">
            <a href="/"><button type="button" class="btn-gray">资产查询</button></a>
            <a href="/cable"><button type="button">电缆查询</button></a>
        </div>
        <h2>电缆查询</h2>
        {% if scan_code %}<div style="color:#666;margin-bottom:10px;word-break:break-all;">当前来自扫码访问，已自动带入电缆编号：<strong>{{ scan_code }}</strong></div>{% endif %}
        <div class="muted" style="margin-bottom:10px;">可搜索：电缆编号、电缆名称、责任人、位置、备注、空货架。位置可点击进入货架详情。</div>
        <form method="get" action="/cable">
            <input type="hidden" name="searched" value="1">
            <div class="action-bar">
                <input type="text" name="keyword" value="{{ keyword }}" placeholder="电缆编号、电缆名称、责任人、位置、空货架">
                <select name="status_filter">
                    <option value="">全部状态</option>
                    {% for s in statuses %}
                    <option value="{{ s }}" {% if status_filter == s %}selected{% endif %}>{{ s }}</option>
                    {% endfor %}
                </select>
                <select name="per_page">
                    <option value="30" {% if per_page == 30 %}selected{% endif %}>每页30条</option>
                    <option value="50" {% if per_page == 50 %}selected{% endif %}>每页50条</option>
                    <option value="100" {% if per_page == 100 %}selected{% endif %}>每页100条</option>
                </select>
                <button type="submit">搜索</button>
                <a href="/cable"><button type="button" class="btn-gray">重置</button></a>
                <a href="/cable/new{% if prefill_cable_no %}?cable_no={{ prefill_cable_no|urlencode }}{% endif %}"><button type="button" class="btn-green">+</button></a>
            </div>
        </form>

        <form method="post" action="/cable/import" enctype="multipart/form-data" style="margin-top:12px;">
            <div class="action-bar">
                <input type="file" name="excel_file" accept=".xlsx,.xlsm,.xltx,.xltm">
                <button type="submit" class="btn-orange">批量上传电缆</button>
            </div>
            <div class="muted" style="margin-top:8px;">Excel格式：电缆编号、电缆名称、规格、责任人、位置、状态、备注。批量导入时若位置不存在，将自动补建货架。</div>
        </form>

        {% if error %}<div class="err">{{ error }}</div>{% endif %}
    </div>

    {% if searched %}
    <div class="card">
        <form method="post" action="/cable/export">
            <input type="hidden" name="keyword" value="{{ keyword }}">
            <input type="hidden" name="status_filter" value="{{ status_filter }}">
            <input type="hidden" name="per_page" value="{{ per_page }}">
            <input type="hidden" name="delete_pin" id="delete_pin" value="">

            <div class="action-bar" style="margin-bottom:10px;">
                <button type="submit" class="btn-green">导出选中</button>
                <button type="submit" class="btn-red" formaction="/cable/delete_selected" formmethod="post" onclick="return confirmDeleteSelected();">删除选中</button>
                <div>当前总数：{{ total }}</div>
            </div>

            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th><input type="checkbox" onclick="toggleSelectCurrentPage(this)" title="选取当前页"></th>
                            <th>编号</th>
                            <th>名称</th>
                            <th>位置</th>
                            <th>规格</th>
                            <th>责任</th>
                            <th>状态</th>
                            <th>备注</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for item in rows %}
                        <tr>
                            <td><input type="checkbox" name="selected_items" value="{{ item.selected_value }}" data-row-type="{{ item.row_type }}"></td>
                            <td>{% if item.row_type == 'shelf' %}<a class="num-link" href="{{ item.detail_url }}">（空）</a>{% else %}<a class="num-link" href="{{ item.detail_url }}">{{ item.cable_no }}</a>{% endif %}</td>
                            <td>{{ item.name }}</td>
                            <td>
                                {% if item.location_detail_url %}
                                <a class="num-link" href="{{ item.location_detail_url }}">{{ item.location }}</a>
                                {% else %}
                                {{ item.location }}
                                {% endif %}
                            </td>
                            <td>{{ item.spec }}</td>
                            <td>{{ item.owner }}</td>
                            <td>{{ item.status }}</td>
                            <td>{{ item.remark }}</td>
                        </tr>
                        {% endfor %}
                        {% if not rows %}
                        <tr><td colspan="8">暂无数据</td></tr>
                        {% endif %}
                    </tbody>
                </table>
            </div>
        </form>

        <div class="pagination">
            {% if page > 1 %}
                <a href="/cable?searched=1&keyword={{ keyword }}&status_filter={{ status_filter }}&per_page={{ per_page }}&page={{ page - 1 }}">上一页</a>
            {% endif %}
            <span>第 {{ page }} / {{ total_pages }} 页</span>
            {% if page < total_pages %}
                <a href="/cable?searched=1&keyword={{ keyword }}&status_filter={{ status_filter }}&per_page={{ per_page }}&page={{ page + 1 }}">下一页</a>
            {% endif %}
        </div>
    </div>
    {% endif %}
</div>
</body>
</html>
"""


CABLE_NEW_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>新增电缆/货架</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:960px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.switch-row{display:flex;gap:8px;flex-wrap:wrap;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.err{color:red;margin-bottom:10px;}
.msg{color:green;margin-bottom:10px;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
.muted{color:#666;font-size:14px;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.add('show'); }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.remove('show'); }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });
    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }
    if(dialogId){ closeUploadChooser(dialogId); }
}

function toggleAddType(){
    const addType = document.getElementById('add_type').value;
    const cableFields = document.querySelectorAll('.cable-field');
    const shelfFields = document.querySelectorAll('.shelf-field');
    cableFields.forEach(el => { el.style.display = addType === '电缆' ? 'block' : 'none'; });
    shelfFields.forEach(el => { el.style.display = addType === '货架' ? 'block' : 'none'; });

    const note = document.getElementById('page-note');
    if(note){
        if(addType === '货架'){
            note.textContent = '先创建货架（位置）后，再在该位置下继续新增电缆。货架也支持单独上传图片。';
        }else{
            note.textContent = '新增电缆时，电缆编号可不填；若位置已存在将自动归入该货架；若位置不存在，请先切换为“货架”创建位置。';
        }
    }
}
</script>
</head>
<body onload="toggleAddType()">
<div class="wrap">
    <div class="card">
        <div class="switch-row">
            <a href="/"><button type="button" class="btn-gray">资产查询</button></a>
            <a href="/cable"><button type="button">电缆查询</button></a>
        </div>
    </div>

    <div class="card"><a href="/cable">返回电缆查询页</a></div>

    <div class="card">
        <h2>新增电缆/货架</h2>
        <div id="page-note" class="muted" style="margin-bottom:10px;">新增电缆时，电缆编号可不填；若位置已存在将自动归入该货架；若位置不存在，请先切换为“货架”创建位置。</div>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row">
                    <label>新增类型</label>
                    <select name="add_type" id="add_type" onchange="toggleAddType()">
                        {% for item in add_types %}
                        <option value="{{ item }}" {% if form_data.add_type == item %}selected{% endif %}>{{ item }}</option>
                        {% endfor %}
                    </select>
                </div>

                <div class="row cable-field"><label>电缆编号（可空）</label><input type="text" name="cable_no" value="{{ form_data.cable_no }}"></div>
                <div class="row cable-field"><label>电缆名称</label><input type="text" name="name" value="{{ form_data.name }}"></div>

                <div class="row cable-field">
                    <label>规格</label>
                    <select name="spec">
                        <option value="">请选择</option>
                        {% for s in specs %}
                        <option value="{{ s }}" {% if form_data.spec == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>

                <div class="row cable-field"><label>责任人</label><input type="text" name="owner" value="{{ form_data.owner }}"></div>

                <div class="row cable-field">
                    <label>位置</label>
                    <input type="text" name="location" value="{{ form_data.location }}" list="cable-location-options">
                    <datalist id="cable-location-options">
                        {% for item in shelf_locations %}
                        <option value="{{ item }}"></option>
                        {% endfor %}
                    </datalist>
                </div>

                <div class="row cable-field">
                    <label>状态</label>
                    <select name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s }}" {% if form_data.status == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>

                <div class="row shelf-field"><label>货架位置</label><input type="text" name="shelf_location" value="{{ form_data.shelf_location }}" list="cable-location-options"></div>

                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" onclick="openUploadChooser('new-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="new-image-files-text" class="file-list" data-inputs="new-camera-files,new-file-files">未选择图片</div>
                    <div id="new-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('new-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input type="file" id="new-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('new-camera-files', 'new-image-files-text', 'new-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input type="file" id="new-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('new-file-files', 'new-image-files-text', 'new-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('new-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <div class="row"><label>备注</label><textarea name="remark">{{ form_data.remark }}</textarea></div>
            <button type="submit">保存</button>
        </form>
    </div>
</div>
</body>
</html>
"""


CABLE_FORM_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ page_title }}</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:950px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.switch-row{display:flex;gap:8px;flex-wrap:wrap;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,select,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.btn-gray{background:#6c757d;}
.btn-green{background:#198754;}
.readonly{background:#e9ecef;color:#666;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.msg{color:green;margin-bottom:10px;}
.err{color:red;margin-bottom:10px;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
.detail-actions{display:flex;gap:82px;flex-wrap:wrap;align-items:center;margin-top:12px;}
.detail-actions button{width:auto;min-width:120px;}
.link-row{margin-top:-4px;margin-bottom:12px;color:#666;font-size:14px;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.add('show'); }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.remove('show'); }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });
    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }
    if(dialogId){ closeUploadChooser(dialogId); }
}

function enableEdit(formId){
    const form = document.getElementById(formId);
    const fields = form.querySelectorAll('.edit-field');
    fields.forEach(el => {
        el.disabled = false;
        el.classList.remove('readonly');
    });
    document.getElementById(formId + '-save').style.display = 'inline-block';
    document.getElementById(formId + '-edit').style.display = 'none';
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="switch-row">
            <a href="/"><button type="button" class="btn-gray">资产查询</button></a>
            <a href="/cable"><button type="button">电缆查询</button></a>
        </div>
    </div>

    <div class="card"><a href="/cable">返回电缆查询页</a></div>

    <div class="card">
        <h2>{{ page_title }}</h2>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form id="cable-form" method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row"><label>电缆编号（可空）</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="cable_no" value="{{ form_data.cable_no }}"></div>
                <div class="row"><label>电缆名称</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="name" value="{{ form_data.name }}"></div>
                <div class="row">
                    <label>规格</label>
                    <select class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="spec">
                        <option value="">请选择</option>
                        {% for s in specs %}
                        <option value="{{ s }}" {% if form_data.spec == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="row"><label>责任人</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="owner" value="{{ form_data.owner }}"></div>
                <div class="row">
                    <label>位置</label>
                    <input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="location" value="{{ form_data.location }}" list="detail-location-options">
                    <datalist id="detail-location-options">
                        {% for item in shelf_locations %}
                        <option value="{{ item }}"></option>
                        {% endfor %}
                    </datalist>
                </div>
                <div class="row">
                    <label>状态</label>
                    <select class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="status">
                        <option value="">请选择</option>
                        {% for s in statuses %}
                        <option value="{{ s }}" {% if form_data.status == s %}selected{% endif %}>{{ s }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="row">
                    <label>上传图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} onclick="openUploadChooser('cable-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="cable-image-files-text" class="file-list" data-inputs="cable-camera-files,cable-file-files">未选择图片</div>
                    <div id="cable-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('cable-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="cable-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('cable-camera-files', 'cable-image-files-text', 'cable-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="cable-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('cable-file-files', 'cable-image-files-text', 'cable-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('cable-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            {% if location_detail_url %}
            <div class="link-row">当前位置对应货架：<a href="{{ location_detail_url }}">{{ form_data.location }}</a></div>
            {% endif %}

            {% if images is not none %}
            <div class="row">
                <label>当前图片</label>
                {% if images %}
                    {% for img in images %}
                    <div class="image-row">
                        <a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a>
                        <label style="display:flex;align-items:center;gap:6px;font-weight:normal;">
                            <input class="edit-field readonly" disabled type="checkbox" name="delete_cable_image_ids" value="{{ img.id }}">
                            删除
                        </label>
                    </div>
                    {% endfor %}
                    <div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“确认”。</div>
                {% else %}
                    <div>暂无图片</div>
                {% endif %}
            </div>
            {% endif %}

            <div class="row"><label>备注</label><textarea class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="remark">{{ form_data.remark }}</textarea></div>

            {% if readonly %}
            <div class="detail-actions">
                <button type="button" id="cable-form-edit" onclick="enableEdit('cable-form')">修改</button>
                <button type="submit" id="cable-form-save" style="display:none;">确认</button>
            </div>
            {% else %}
            <div class="detail-actions">
                <button type="submit" class="btn-green">保存</button>
                <a href="/cable"><button type="button" class="btn-gray">返回</button></a>
            </div>
            {% endif %}
        </form>
    </div>
</div>
</body>
</html>
"""


SHELF_DETAIL_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>货架详情</title>
<style>
body{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;}
.wrap{max-width:1100px;margin:auto;padding:14px;}
.card{background:#fff;border-radius:12px;padding:16px;margin-bottom:16px;box-shadow:0 1px 6px rgba(0,0,0,.08);}
.switch-row{display:flex;gap:8px;flex-wrap:wrap;}
.switch-row a{flex:1 1 160px;}
.switch-row a button{width:100%;}
.row{margin-bottom:12px;}
label{display:block;margin-bottom:6px;font-weight:bold;}
input,textarea,button{width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;}
textarea{min-height:90px;}
button{background:#0d6efd;color:#fff;border:none;}
.btn-gray{background:#6c757d;}
.btn-green{background:#198754;}
.readonly{background:#e9ecef;color:#666;}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.msg{color:green;margin-bottom:10px;}
.err{color:red;margin-bottom:10px;}
.image-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.upload-actions{display:flex;gap:8px;flex-wrap:wrap;}
.upload-actions button{width:auto;min-width:120px;}
.file-list{margin-top:8px;color:#666;font-size:14px;word-break:break-all;}
.upload-dialog{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
.upload-dialog.show{display:flex;}
.upload-dialog-card{width:100%;max-width:320px;background:#fff;border-radius:14px;padding:16px;box-shadow:0 8px 24px rgba(0,0,0,.18);}
.upload-dialog-title{font-size:18px;font-weight:bold;margin-bottom:12px;text-align:center;}
.upload-dialog-actions{display:flex;flex-direction:column;gap:10px;}
.upload-dialog-actions button{width:100%;}
.upload-choice-file{position:relative;display:block;width:100%;box-sizing:border-box;padding:10px;font-size:16px;border-radius:8px;border:1px solid #ccc;background:#0d6efd;color:#fff;text-align:center;overflow:hidden;}
.upload-choice-file input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer;}
.upload-dialog-actions .btn-cancel{background:#6c757d;}
.table-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;min-width:780px;}
th,td{border:1px solid #ddd;padding:10px;text-align:left;}
th{background:#f0f3f8;}
.detail-actions{display:flex;gap:82px;flex-wrap:wrap;align-items:center;margin-top:12px;}
.detail-actions button{width:auto;min-width:120px;}
.muted{color:#666;font-size:14px;}
@media (max-width:768px){.grid{grid-template-columns:1fr;}}
</style>
<script>
function openUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.add('show'); }
}

function closeUploadChooser(dialogId){
    const dialog = document.getElementById(dialogId);
    if(dialog){ dialog.classList.remove('show'); }
}

function updateSelectedFiles(inputId, textId, dialogId){
    const textEl = document.getElementById(textId);
    if(!textEl){ return; }
    const inputIds = (textEl.dataset.inputs || inputId || '').split(',').map(item => item.trim()).filter(Boolean);
    const files = [];
    inputIds.forEach(id => {
        const input = document.getElementById(id);
        if(input && input.files){
            Array.from(input.files).forEach(file => files.push(file));
        }
    });
    if(files.length > 0){
        const names = files.map(file => file.name).join('，');
        textEl.textContent = `已选择 ${files.length} 张：${names}`;
    }else{
        textEl.textContent = '未选择图片';
    }
    if(dialogId){ closeUploadChooser(dialogId); }
}

function enableEdit(formId){
    const form = document.getElementById(formId);
    const fields = form.querySelectorAll('.edit-field');
    fields.forEach(el => {
        el.disabled = false;
        el.classList.remove('readonly');
    });
    document.getElementById(formId + '-save').style.display = 'inline-block';
    document.getElementById(formId + '-edit').style.display = 'none';
}
</script>
</head>
<body>
<div class="wrap">
    <div class="card">
        <div class="switch-row">
            <a href="/"><button type="button" class="btn-gray">资产查询</button></a>
            <a href="/cable"><button type="button">电缆查询</button></a>
        </div>
    </div>

    <div class="card"><a href="/cable">返回电缆查询页</a></div>

    <div class="card">
        <h2>货架详情</h2>
        <div class="muted" style="margin-bottom:10px;">该货架就是电缆位置。修改位置名称时，会同步更新该位置下的所有电缆。</div>
        {% if message %}<div class="msg">{{ message }}</div>{% endif %}
        {% if error %}<div class="err">{{ error }}</div>{% endif %}

        <form id="shelf-form" method="post" enctype="multipart/form-data">
            <div class="grid">
                <div class="row"><label>货架位置</label><input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="text" name="location" value="{{ form_data.location }}"></div>
                <div class="row">
                    <label>上传货架图片（最多5张）</label>
                    <div class="upload-actions">
                        <button type="button" class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} onclick="openUploadChooser('shelf-upload-choice-dialog')">上传图片</button>
                    </div>
                    <div id="shelf-image-files-text" class="file-list" data-inputs="shelf-camera-files,shelf-file-files">未选择图片</div>
                    <div id="shelf-upload-choice-dialog" class="upload-dialog" onclick="if(event.target === this){closeUploadChooser('shelf-upload-choice-dialog');}">
                        <div class="upload-dialog-card">
                            <div class="upload-dialog-title">请选择上传方式</div>
                            <div class="upload-dialog-actions">
                                <label class="upload-choice-file">拍照
                                    <input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="shelf-camera-files" name="image_files" accept="image/*" capture="environment" multiple onchange="updateSelectedFiles('shelf-camera-files', 'shelf-image-files-text', 'shelf-upload-choice-dialog')">
                                </label>
                                <label class="upload-choice-file">本地上传
                                    <input class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} type="file" id="shelf-file-files" name="image_files" accept="image/*" multiple onchange="updateSelectedFiles('shelf-file-files', 'shelf-image-files-text', 'shelf-upload-choice-dialog')">
                                </label>
                                <button type="button" class="btn-cancel" onclick="closeUploadChooser('shelf-upload-choice-dialog')">取消</button>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <div class="row">
                <label>当前货架图片</label>
                {% if images %}
                    {% for img in images %}
                    <div class="image-row">
                        <a href="/uploads/{{ img.image_path }}" target="_blank">图片{{ loop.index }}</a>
                        <label style="display:flex;align-items:center;gap:6px;font-weight:normal;">
                            <input class="edit-field readonly" disabled type="checkbox" name="delete_shelf_image_ids" value="{{ img.id }}">
                            删除
                        </label>
                    </div>
                    {% endfor %}
                    <div style="color:#666;font-size:13px;">先点击“修改”，勾选要删除的图片，再点击“确认”。</div>
                {% else %}
                    <div>暂无图片</div>
                {% endif %}
            </div>

            <div class="row"><label>备注</label><textarea class="{{ 'edit-field readonly' if readonly else '' }}" {% if readonly %}disabled{% endif %} name="remark">{{ form_data.remark }}</textarea></div>

            <div class="detail-actions">
                <button type="button" id="shelf-form-edit" onclick="enableEdit('shelf-form')">修改</button>
                <button type="submit" id="shelf-form-save" style="display:none;">确认</button>
            </div>
        </form>
    </div>

    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2 style="margin:0;">该位置下的电缆</h2>
            <a href="/cable/new?add_type=电缆&location={{ form_data.location|urlencode }}"><button type="button" class="btn-green" style="width:auto;">+</button></a>
        </div>

        <div class="table-wrap" style="margin-top:12px;">
            <table>
                <thead>
                    <tr>
                        <th>编号</th>
                        <th>名称</th>
                        <th>规格</th>
                        <th>责任</th>
                        <th>状态</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    {% for item in cable_rows %}
                    <tr>
                        <td><a href="{{ item.detail_url }}">{{ item.cable_no }}</a></td>
                        <td>{{ item.name }}</td>
                        <td>{{ item.spec }}</td>
                        <td>{{ item.owner }}</td>
                        <td>{{ item.status }}</td>
                        <td>{{ item.remark }}</td>
                    </tr>
                    {% endfor %}
                    {% if not cable_rows %}
                    <tr><td colspan="6">该位置下暂无电缆</td></tr>
                    {% endif %}
                </tbody>
            </table>
        </div>
    </div>
</div>
</body>
</html>
"""


def register_cable_routes(app):
    with app.app_context():
        Cable.__table__.create(bind=db.engine, checkfirst=True)
        CableImage.__table__.create(bind=db.engine, checkfirst=True)
        CableShelf.__table__.create(bind=db.engine, checkfirst=True)
        CableShelfImage.__table__.create(bind=db.engine, checkfirst=True)
        backfill_cable_shelves()

    @app.route("/cable", defaults={"scan_code": ""}, methods=["GET"])
    @app.route("/cable<scan_code>", methods=["GET"])
    def search_cables(scan_code=""):
        path_scan_code = normalize_text(scan_code)
        prefill_cable_no = extract_scan_prefill_code(path_scan_code)
        keyword = normalize_text(request.args.get("keyword"))
        if not keyword:
            keyword = prefill_cable_no or extract_scan_keyword(path_scan_code)

        status_filter = normalize_text(request.args.get("status_filter"))
        searched = True if path_scan_code else (normalize_text(request.args.get("searched")) == "1")
        if keyword and not searched:
            searched = True

        per_page = normalize_text(request.args.get("per_page")) or "30"
        page = normalize_text(request.args.get("page")) or "1"

        try:
            per_page = int(per_page)
        except Exception:
            per_page = 30
        if per_page not in [30, 50, 100]:
            per_page = 30

        try:
            page = int(page)
        except Exception:
            page = 1

        all_rows = build_cable_rows(keyword=keyword, searched=searched)
        if searched and status_filter:
            all_rows = [row for row in all_rows if normalize_text(row.get("status")) == status_filter]

        total = len(all_rows)
        start = (page - 1) * per_page
        end = start + per_page
        rows = all_rows[start:end]
        total_pages = (total + per_page - 1) // per_page if total else 1

        error = ""
        if searched and total == 0:
            if keyword and status_filter:
                error = "未找到符合筛选条件的电缆"
            elif keyword:
                error = "未找到对应电缆"
            elif status_filter:
                error = "未找到符合筛选条件的电缆"

        return render_template_string(
            CABLE_SEARCH_HTML,
            current_user=current_user,
            keyword=keyword,
            status_filter=status_filter,
            statuses=CABLE_STATUSES,
            searched=searched,
            per_page=per_page,
            page=page,
            total=total,
            total_pages=total_pages,
            rows=rows,
            error=error,
            scan_code=path_scan_code,
            prefill_cable_no=prefill_cable_no,
        )

    @app.route("/cable/import", methods=["POST"])
    def import_cables():
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            return redirect(url_for("search_cables"))

        try:
            import_cables_from_excel(excel_file)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量上传失败：{str(e)}"

        return redirect(url_for("search_cables"))

    @app.route("/cable/export", methods=["POST"])
    def export_selected_cables():
        selected_items = request.form.getlist("selected_items")
        wb = Workbook()
        ws = wb.active
        ws.title = "电缆导出"
        ws.append(["电缆编号", "电缆名称", "规格", "责任人", "位置", "状态", "备注"])

        for item in selected_items:
            row_type = "cable"
            row_id_text = normalize_text(item)
            if ":" in row_id_text:
                row_type, row_id_text = row_id_text.split(":", 1)
            if row_type != "cable":
                continue
            try:
                cable_id = int(row_id_text)
            except Exception:
                continue
            obj = Cable.query.get(cable_id)
            if obj:
                ws.append([
                    get_display_cable_no(obj.cable_no),
                    obj.name or "",
                    obj.spec or "",
                    obj.owner or "",
                    obj.location or "",
                    obj.status or "",
                    obj.remark or "",
                ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"cable_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/cable/delete_selected", methods=["POST"])
    def delete_selected_cables():
        selected_items = request.form.getlist("selected_items")
        delete_pin = normalize_text(request.form.get("delete_pin"))
        keyword = normalize_text(request.form.get("keyword"))
        status_filter = normalize_text(request.form.get("status_filter"))
        per_page = normalize_text(request.form.get("per_page")) or "30"

        if delete_pin != "0819":
            return "批量删除失败：Pin码错误"

        cable_ids = []
        shelf_ids = []
        for item in selected_items:
            row_type = "cable"
            row_id_text = normalize_text(item)
            if ":" in row_id_text:
                row_type, row_id_text = row_id_text.split(":", 1)
            try:
                row_id = int(row_id_text)
            except Exception:
                continue
            if row_type == "shelf":
                shelf_ids.append(row_id)
            else:
                cable_ids.append(row_id)

        try:
            cables = Cable.query.filter(Cable.id.in_(cable_ids)).order_by(Cable.id.asc()).all() if cable_ids else []
            shelves = CableShelf.query.filter(CableShelf.id.in_(shelf_ids)).order_by(CableShelf.id.asc()).all() if shelf_ids else []
            occupied_locations = {
                normalize_location(row[0])
                for row in db.session.query(Cable.location).filter(Cable.location.isnot(None)).distinct().all()
                if normalize_location(row[0])
            }

            for obj in cables:
                delete_cable_with_files(obj)

            blocked_shelves = []
            for shelf in shelves:
                shelf_name = normalize_location(shelf.shelf_name)
                if shelf_name and shelf_name in occupied_locations:
                    blocked_shelves.append(shelf_name)
                    continue
                delete_shelf_with_files(shelf)

            if blocked_shelves:
                db.session.rollback()
                return "批量删除失败：以下货架下仍有电缆，不能删除：" + "，".join(blocked_shelves)

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"批量删除失败：{str(e)}"

        return redirect(url_for("search_cables", searched=1, keyword=keyword, status_filter=status_filter, per_page=per_page, page=1))

    @app.route("/cable/new", methods=["GET", "POST"])
    def cable_new():
        error = ""
        message = ""
        prefill_cable_no = normalize_text(request.args.get("cable_no")) or extract_scan_prefill_code() or extract_scan_keyword()
        prefill_location = normalize_location(request.args.get("location"))
        default_add_type = normalize_text(request.args.get("add_type")) or "电缆"
        if default_add_type not in CABLE_ADD_TYPES:
            default_add_type = "电缆"

        form_data = {
            "add_type": default_add_type,
            "cable_no": prefill_cable_no,
            "name": "",
            "spec": "高频",
            "owner": "",
            "location": prefill_location,
            "status": "在库",
            "remark": "",
            "shelf_location": prefill_location,
        }

        if request.method == "POST":
            add_type = normalize_text(request.form.get("add_type")) or "电缆"
            if add_type not in CABLE_ADD_TYPES:
                add_type = "电缆"

            cable_no = normalize_text(request.form.get("cable_no"))
            name = normalize_text(request.form.get("name"))
            spec = normalize_text(request.form.get("spec"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_location(request.form.get("location"))
            status = normalize_text(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            shelf_location = normalize_location(request.form.get("shelf_location"))
            image_files = request.files.getlist("image_files")

            form_data = {
                "add_type": add_type,
                "cable_no": cable_no,
                "name": name,
                "spec": spec,
                "owner": owner,
                "location": location,
                "status": status,
                "remark": remark,
                "shelf_location": shelf_location,
            }

            if add_type == "货架":
                if not shelf_location:
                    error = "货架位置不能为空"
                elif CableShelf.query.filter_by(shelf_name=shelf_location).first():
                    error = "该货架位置已存在"
                else:
                    try:
                        shelf = CableShelf(
                            shelf_name=shelf_location,
                            remark=normalize_empty_to_none(remark),
                        )
                        db.session.add(shelf)
                        db.session.flush()

                        for file_storage in image_files:
                            rel = save_uploaded_image(file_storage, "cable_shelf", shelf.shelf_name)
                            if rel:
                                db.session.add(CableShelfImage(shelf_id=shelf.id, image_path=rel))

                        trim_shelf_images(shelf)
                        db.session.commit()
                        return redirect(url_for("cable_location_detail", shelf_id=shelf.id))
                    except Exception as e:
                        db.session.rollback()
                        error = f"保存失败：{str(e)}"
            else:
                if not location:
                    error = "位置不能为空，请先创建货架"
                elif spec and spec not in CABLE_SPECS:
                    error = "规格不合法"
                elif status and status not in CABLE_STATUSES:
                    error = "状态不合法"
                elif cable_no and Cable.query.filter_by(cable_no=cable_no).first():
                    error = "电缆编号已存在"
                else:
                    shelf = get_shelf_by_location(location)
                    if not shelf:
                        error = "该位置不存在，请先创建货架"
                    else:
                        try:
                            stored_cable_no = make_stored_cable_no(cable_no)
                            obj = Cable(
                                cable_no=stored_cable_no,
                                name=name or "",
                                spec=normalize_empty_to_none(spec),
                                owner=normalize_empty_to_none(owner),
                                location=normalize_location(shelf.shelf_name),
                                status=normalize_empty_to_none(status),
                                remark=normalize_empty_to_none(remark),
                            )
                            db.session.add(obj)
                            db.session.flush()

                            image_prefix = build_cable_filename_prefix(stored_cable_no, name, shelf.shelf_name)
                            for file_storage in image_files:
                                rel = save_uploaded_image(file_storage, "cable", image_prefix)
                                if rel:
                                    db.session.add(CableImage(cable_id=obj.id, image_path=rel))

                            trim_cable_images(obj)
                            db.session.commit()
                            return redirect(url_for("cable_detail", cable_id=obj.id))
                        except Exception as e:
                            db.session.rollback()
                            error = f"保存失败：{str(e)}"

        return render_template_string(
            CABLE_NEW_HTML,
            current_user=current_user,
            add_types=CABLE_ADD_TYPES,
            specs=CABLE_SPECS,
            statuses=CABLE_STATUSES,
            form_data=form_data,
            error=error,
            message=message,
            shelf_locations=get_all_shelf_locations(),
        )

    @app.route("/cable/location/<int:shelf_id>", methods=["GET", "POST"])
    def cable_location_detail(shelf_id):
        shelf = CableShelf.query.get_or_404(shelf_id)
        message = ""
        error = ""

        if request.method == "POST":
            location = normalize_location(request.form.get("location"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")
            delete_image_ids = request.form.getlist("delete_shelf_image_ids")

            if not location:
                error = "货架位置不能为空"
            else:
                existing = CableShelf.query.filter(CableShelf.shelf_name == location, CableShelf.id != shelf.id).first()
                if existing:
                    error = "该货架位置已存在"
                else:
                    old_location = shelf.shelf_name or ""
                    try:
                        shelf.shelf_name = location
                        shelf.remark = normalize_empty_to_none(remark)

                        if old_location != location:
                            Cable.query.filter_by(location=old_location).update({"location": location}, synchronize_session=False)

                        for image_id in delete_image_ids:
                            try:
                                img_id = int(image_id)
                            except Exception:
                                continue
                            img = CableShelfImage.query.filter_by(id=img_id, shelf_id=shelf.id).first()
                            if img:
                                delete_image_file(img.image_path)
                                db.session.delete(img)

                        for file_storage in image_files:
                            rel = save_uploaded_image(file_storage, "cable_shelf", shelf.shelf_name)
                            if rel:
                                db.session.add(CableShelfImage(shelf_id=shelf.id, image_path=rel))

                        trim_shelf_images(shelf)
                        db.session.commit()
                        message = "货架更新成功"
                    except Exception as e:
                        db.session.rollback()
                        error = f"更新失败：{str(e)}"

        images = get_shelf_images(shelf.id)
        cable_rows = []
        for item in Cable.query.filter_by(location=shelf.shelf_name).order_by(Cable.cable_no.asc(), Cable.id.asc()).all():
            cable_rows.append({
                "id": item.id,
                "cable_no": get_display_cable_no(item.cable_no),
                "name": item.name or "",
                "spec": item.spec or "",
                "owner": item.owner or "",
                "status": item.status or "",
                "remark": item.remark or "",
                "detail_url": url_for("cable_detail", cable_id=item.id),
            })

        form_data = OrderedDict([
            ("location", normalize_location(shelf.shelf_name)),
            ("remark", shelf.remark or ""),
        ])

        return render_template_string(
            SHELF_DETAIL_HTML,
            current_user=current_user,
            form_data=form_data,
            error=error,
            message=message,
            readonly=True,
            images=images,
            cable_rows=cable_rows,
        )

    @app.route("/cable/<int:cable_id>", methods=["GET", "POST"])
    def cable_detail(cable_id):
        cable = Cable.query.get_or_404(cable_id)
        message = ""
        error = ""

        if request.method == "POST":
            cable_no = normalize_text(request.form.get("cable_no"))
            name = normalize_text(request.form.get("name"))
            spec = normalize_text(request.form.get("spec"))
            owner = normalize_text(request.form.get("owner"))
            location = normalize_location(request.form.get("location"))
            status = normalize_text(request.form.get("status"))
            remark = normalize_text(request.form.get("remark"))
            image_files = request.files.getlist("image_files")
            delete_image_ids = request.form.getlist("delete_cable_image_ids")

            if not location:
                error = "位置不能为空，请先创建货架"
            elif spec and spec not in CABLE_SPECS:
                error = "规格不合法"
            elif status and status not in CABLE_STATUSES:
                error = "状态不合法"
            else:
                existing = Cable.query.filter(Cable.cable_no == cable_no, Cable.id != cable.id).first() if cable_no else None
                if existing:
                    error = "电缆编号已存在"
                else:
                    shelf = get_shelf_by_location(location)
                    if not shelf:
                        error = "该位置不存在，请先创建货架"
                    else:
                        try:
                            cable.cable_no = make_stored_cable_no(cable_no, cable.cable_no)
                            cable.name = name or ""
                            cable.spec = normalize_empty_to_none(spec)
                            cable.owner = normalize_empty_to_none(owner)
                            cable.location = shelf.shelf_name
                            cable.status = normalize_empty_to_none(status)
                            cable.remark = normalize_empty_to_none(remark)

                            for image_id in delete_image_ids:
                                try:
                                    img_id = int(image_id)
                                except Exception:
                                    continue
                                img = CableImage.query.filter_by(id=img_id, cable_id=cable.id).first()
                                if img:
                                    delete_image_file(img.image_path)
                                    db.session.delete(img)

                            image_prefix = build_cable_filename_prefix(cable.cable_no, cable.name, cable.location)
                            for file_storage in image_files:
                                rel = save_uploaded_image(file_storage, "cable", image_prefix)
                                if rel:
                                    db.session.add(CableImage(cable_id=cable.id, image_path=rel))

                            trim_cable_images(cable)
                            db.session.commit()
                            message = "电缆更新成功"
                        except Exception as e:
                            db.session.rollback()
                            error = f"更新失败：{str(e)}"

        images = get_cable_images(cable.id)
        shelf = get_shelf_by_location(normalize_location(cable.location))
        form_data = OrderedDict([
            ("cable_no", get_display_cable_no(cable.cable_no)),
            ("name", cable.name or ""),
            ("spec", cable.spec or ""),
            ("owner", cable.owner or ""),
            ("location", normalize_location(cable.location)),
            ("status", cable.status or ""),
            ("remark", cable.remark or ""),
        ])

        return render_template_string(
            CABLE_FORM_HTML,
            current_user=current_user,
            page_title="电缆详情",
            specs=CABLE_SPECS,
            statuses=CABLE_STATUSES,
            form_data=form_data,
            error=error,
            message=message,
            readonly=True,
            images=images,
            shelf_locations=get_all_shelf_locations(),
            location_detail_url=url_for("cable_location_detail", shelf_id=shelf.id) if shelf else "",
        )

    @app.route("/cable/<int:cable_id>/delete", methods=["POST"])
    def delete_cable(cable_id):
        cable = Cable.query.get_or_404(cable_id)
        try:
            delete_cable_with_files(cable)
            db.session.commit()
            return redirect(url_for("search_cables"))
        except Exception as e:
            db.session.rollback()
            return f"删除失败：{str(e)}"
